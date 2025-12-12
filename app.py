"""
Aplicación principal Flask para el sistema de gestión de residencias Violetas.
Implementa autenticación JWT, permisos granulares y filtrado de datos por residencia.
"""
import sys
print("=== INICIANDO IMPORTACIÓN DE APP.PY ===", file=sys.stderr, flush=True)
import os
import jwt
import re
import json
from datetime import datetime, timedelta
from functools import wraps
from collections import defaultdict
from io import BytesIO
from flask import Flask, request, jsonify, g, send_from_directory, Response
from flask_cors import CORS
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
from db_connector import get_db_connection
from storage_manager import upload_document, get_document_url, delete_document
from validators import (
    validate_residente_data, validate_cobro_data, validate_monto,
    validate_residencia_id, validate_estado, validate_metodo_pago,
    validate_text, validate_email, validate_phone, validate_personal_data,
    validate_turno_extra_data
)
import mimetypes

# Cargar variables de entorno desde .env (solo en desarrollo local)
# En Cloud Run, las variables vienen de --set-env-vars y --set-secrets
if os.path.exists('.env'):
    load_dotenv()

app = Flask(__name__, static_folder='static', template_folder='templates')
print("=== APP FLASK CREADA ===", file=sys.stderr, flush=True)
CORS(app)  # Habilitar CORS para el frontend
print("=== CORS HABILITADO ===", file=sys.stderr, flush=True)

# Configuración
# En Cloud Run, --set-secrets crea variables de entorno automáticamente
# Permitir que la aplicación se importe sin JWT_SECRET_KEY (se validará al iniciar)
JWT_SECRET_KEY = os.getenv('JWT_SECRET_KEY')

# Función para validar JWT_SECRET_KEY cuando sea necesario
def get_jwt_secret():
    """Obtiene JWT_SECRET_KEY, validando que esté disponible."""
    secret = os.getenv('JWT_SECRET_KEY')
    if not secret:
        # Intentar leer desde archivo como fallback (si se monta manualmente)
        secret_path = '/secrets/jwt-secret-key'
        if os.path.exists(secret_path):
            try:
                with open(secret_path, 'r') as f:
                    secret = f.read().strip()
            except:
                pass
    if not secret:
        raise ValueError("JWT_SECRET_KEY debe estar definida como variable de entorno o archivo secreto")
    return secret

app.config['JSON_SORT_KEYS'] = False
app.config['JSON_AS_ASCII'] = False  # Permitir caracteres Unicode en JSON (ñ, acentos, etc.)
app.config['JSONIFY_PRETTYPRINT_REGULAR'] = False

# Constantes de seguridad
ADMIN_ROLE_ID = 2  # ID fijo del rol Administrador (acceso total a todos los módulos y residencias)

# Rate limiting simple en memoria (en producción usar Redis)
login_attempts = defaultdict(list)


def require_auth(f):
    """
    Decorador para proteger rutas que requieren autenticación.
    Útil para aplicar manualmente en rutas específicas si es necesario.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # La validación se hace en before_request, este decorador solo marca la ruta
        return f(*args, **kwargs)
    
    return decorated_function


def validate_password_strength(password):
    """
    Valida que la contraseña cumpla con requisitos de seguridad.
    
    Requisitos:
    - Mínimo 8 caracteres
    - Al menos una mayúscula
    - Al menos una minúscula
    - Al menos un número
    - Al menos un carácter especial
    
    Returns:
        tuple: (is_valid, error_message)
    """
    if len(password) < 8:
        return False, "La contraseña debe tener al menos 8 caracteres"
    
    if not re.search(r'[A-Z]', password):
        return False, "La contraseña debe contener al menos una mayúscula"
    
    if not re.search(r'[a-z]', password):
        return False, "La contraseña debe contener al menos una minúscula"
    
    if not re.search(r'\d', password):
        return False, "La contraseña debe contener al menos un número"
    
    if not re.search(r'[!@#$%^&*(),.?":{}|<>]', password):
        return False, "La contraseña debe contener al menos un carácter especial"
    
    return True, None


def check_rate_limit(ip, max_attempts=5, window_minutes=1):
    """
    Verifica rate limit simple en memoria para prevenir ataques de fuerza bruta.
    
    Args:
        ip: Dirección IP del cliente
        max_attempts: Número máximo de intentos permitidos
        window_minutes: Ventana de tiempo en minutos
        
    Returns:
        bool: True si puede continuar, False si excedió el límite
    """
    now = datetime.utcnow()
    window_start = now - timedelta(minutes=window_minutes)
    
    # Limpiar intentos antiguos
    login_attempts[ip] = [t for t in login_attempts[ip] if t > window_start]
    
    # Verificar límite
    if len(login_attempts[ip]) >= max_attempts:
        return False
    
    # Registrar intento
    login_attempts[ip].append(now)
    return True


def log_security_event(tipo_evento, id_usuario=None, detalles=None):
    """
    Registra eventos de seguridad para auditoría.
    
    Args:
        tipo_evento: Tipo de evento ('login_exitoso', 'login_fallido', 'cambio_clave', etc.)
        id_usuario: ID del usuario (opcional)
        detalles: Diccionario con detalles adicionales (opcional)
    """
    try:
        ip_address = request.remote_addr
        user_agent = request.headers.get('User-Agent', '')
        
        log_data = {
            'tipo': tipo_evento,
            'id_usuario': id_usuario,
            'ip': ip_address,
            'user_agent': user_agent,
            'timestamp': datetime.utcnow().isoformat()
        }
        
        if detalles:
            log_data.update(detalles)
        
        app.logger.info(f"SECURITY: {json.dumps(log_data)}")
    except Exception as e:
        app.logger.error(f"Error al registrar evento de seguridad: {str(e)}")


def validate_residencia_access(id_residencia_from_db, allow_Administrador=True):
    """
    Valida que el id_residencia del recurso esté en la lista de acceso del usuario.
    
    Args:
        id_residencia_from_db: El id_residencia del registro obtenido de la BD
        allow_Administrador: Si True, permite acceso a Administrador (rol 1)
        
    Returns:
        tuple: (is_valid, error_response) donde error_response es None si es válido
    """
    # BYPASS para Administrador
    if allow_Administrador and g.id_rol == ADMIN_ROLE_ID:
        return True, None
    
    # Verificar que la residencia está en la lista de acceso
    if not hasattr(g, 'residencias_acceso') or id_residencia_from_db not in g.residencias_acceso:
        return False, (jsonify({'error': 'No tienes permisos para acceder a este recurso'}), 403)
    
    return True, None


def verificar_cobro_mensual_duplicado(cursor, id_residente, id_residencia, mes_pagado, concepto):
    """
    Verifica si ya existe un cobro mensual (concepto de mes) para el mismo residente y mes.
    Previene duplicados tanto para cobros pendientes como completados.
    
    Args:
        cursor: Cursor de base de datos
        id_residente: ID del residente
        id_residencia: ID de la residencia
        mes_pagado: Mes en formato 'YYYY-MM' o None
        concepto: Concepto del cobro (puede ser "Diciembre 25", "Enero 26", etc.)
    
    Returns:
        tuple: (existe_duplicado, id_pago_duplicado) o (False, None)
    """
    meses_espanol_list = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 
                          'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
    
    # Verificar si el concepto es un concepto mensual (empieza con un mes en español)
    es_concepto_mensual = concepto and any(concepto.lower().startswith(mes) for mes in meses_espanol_list)
    
    if not es_concepto_mensual:
        return False, None
    
    # Construir query de validación
    condiciones = []
    params = [id_residente, id_residencia]
    
    # Si tenemos mes_pagado, usarlo para validación más precisa
    if mes_pagado:
        condiciones.append("mes_pagado = %s")
        params.append(mes_pagado)
    
    # Validar por concepto de mes (cualquier concepto que empiece con un mes)
    condiciones_concepto = []
    for mes in meses_espanol_list:
        condiciones_concepto.append(f"concepto ILIKE '{mes} %%'")
    condiciones_concepto.append("concepto ILIKE 'Pago %%'")
    
    condiciones.append(f"({' OR '.join(condiciones_concepto)})")
    
    query = f"""
        SELECT id_pago, estado, concepto, mes_pagado
        FROM pago_residente
        WHERE id_residente = %s 
          AND id_residencia = %s
          AND {' AND '.join(condiciones)}
        LIMIT 1
    """
    
    cursor.execute(query, params)
    resultado = cursor.fetchone()
    
    if resultado:
        return True, resultado[0]  # Existe duplicado, retornar id_pago
    
    return False, None


def generar_cobros_historicos_completados(cursor, id_residente, id_residencia, fecha_ingreso, costo_habitacion, metodo_pago):
    """
    Genera cobros completados históricos desde la fecha_ingreso hasta el mes anterior al actual.
    Estos cobros aparecerán como completados en Facturación e Históricos.
    
    Args:
        cursor: Cursor de base de datos
        id_residente: ID del residente
        id_residencia: ID de la residencia
        fecha_ingreso: Fecha de ingreso del residente (datetime.date o string YYYY-MM-DD)
        costo_habitacion: Costo mensual de la habitación
        metodo_pago: Método de pago preferido
    
    Returns:
        int: Número de cobros históricos generados
    """
    if not fecha_ingreso or not costo_habitacion or costo_habitacion <= 0:
        return 0
    
    # Convertir fecha_ingreso a date si es string
    if isinstance(fecha_ingreso, str):
        try:
            fecha_ingreso = datetime.strptime(fecha_ingreso, '%Y-%m-%d').date()
        except:
            app.logger.warning(f"Fecha de ingreso inválida: {fecha_ingreso}")
            return 0
    
    hoy = datetime.now().date()
    mes_actual = datetime(hoy.year, hoy.month, 1).date()
    
    # Calcular mes de ingreso (primer día del mes de ingreso)
    mes_ingreso = datetime(fecha_ingreso.year, fecha_ingreso.month, 1).date()
    
    # Si el mes de ingreso es futuro, no generar históricos
    if mes_ingreso > mes_actual:
        return 0
    
    meses_espanol = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    
    cobros_generados = 0
    fecha_actual = mes_ingreso
    
    # Generar cobros para cada mes desde mes_ingreso hasta mes_actual (incluido)
    # Si el residente ingresó antes del mes actual, generar todos los meses hasta el actual
    # Incluir el mes actual si ya pasó el día 1 (es decir, si estamos en cualquier día del mes)
    while fecha_actual <= mes_actual:
        mes_pagado_str = fecha_actual.strftime('%Y-%m')
        nombre_mes = meses_espanol.get(fecha_actual.month, 'mes')
        año_corto = str(fecha_actual.year)[-2:]
        concepto = f"{nombre_mes.capitalize()} {año_corto}"
        
        # Verificar si ya existe un cobro para este mes
        existe_duplicado, _ = verificar_cobro_mensual_duplicado(
            cursor, id_residente, id_residencia, mes_pagado_str, concepto
        )
        
        if not existe_duplicado:
            try:
                # Crear cobro completado (con fecha_pago = día 1 del mes correspondiente)
                cursor.execute("""
                    INSERT INTO pago_residente (
                        id_residente, id_residencia, monto, fecha_pago, fecha_prevista,
                        mes_pagado, concepto, metodo_pago, estado, es_cobro_previsto
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    id_residente,
                    id_residencia,
                    costo_habitacion,
                    fecha_actual,  # fecha_pago = día 1 del mes (cobro completado)
                    fecha_actual,  # fecha_prevista también día 1 del mes
                    mes_pagado_str,
                    concepto,
                    metodo_pago or 'transferencia',
                    'cobrado',  # Estado completado
                    False  # No es cobro previsto, es histórico completado
                ))
                cobros_generados += 1
                app.logger.info(f"Cobro histórico completado generado para residente {id_residente}: €{costo_habitacion}, mes: {mes_pagado_str}")
            except Exception as e:
                app.logger.error(f"Error al generar cobro histórico para residente {id_residente}, mes {mes_pagado_str}: {str(e)}")
        
        # Avanzar al siguiente mes
        if fecha_actual.month == 12:
            fecha_actual = datetime(fecha_actual.year + 1, 1, 1).date()
        else:
            fecha_actual = datetime(fecha_actual.year, fecha_actual.month + 1, 1).date()
    
    return cobros_generados


def build_residencia_filter(table_alias='', column_name='id_residencia'):
    """
    Construye la cláusula WHERE para filtrar por residencias de acceso.
    
    Args:
        table_alias: Alias de la tabla (ej: 'r.' o 'p.')
        column_name: Nombre de la columna (default: 'id_residencia')
        
    Returns:
        tuple: (sql_condition, params)
               - Si Administrador: (None, None) = sin filtro
               - Si usuario normal: ('WHERE ... IN (...)', [lista_ids])
    """
    if g.id_rol == ADMIN_ROLE_ID:
        return None, None
    
    if not hasattr(g, 'residencias_acceso') or not g.residencias_acceso:
        # Usuario sin residencias (no debería pasar, pero por seguridad)
        return 'WHERE FALSE', []  # WHERE FALSE = no resultados
    
    column = f"{table_alias}{column_name}" if table_alias else column_name
    placeholders = ','.join(['%s'] * len(g.residencias_acceso))
    return f"WHERE {column} IN ({placeholders})", g.residencias_acceso


def usuario_tiene_permiso(id_usuario, id_rol, nombre_permiso, cursor=None):
    """
    Verifica si un usuario tiene un permiso específico.
    
    Lógica de verificación (en orden):
    1. Si es Administrador (id_rol=2) → bypass total (siempre True)
    2. Verificar permisos individuales en usuario_permiso
    3. Si no tiene permisos individuales, heredar permisos del rol desde rol_permiso
    
    Esto permite:
    - Permisos individualizados que sobrescriben/complementan los del rol
    - Herencia automática de permisos del rol
    - Máxima flexibilidad en la asignación de permisos
    
    Args:
        id_usuario: ID del usuario
        id_rol: ID del rol del usuario
        nombre_permiso: Nombre del permiso a verificar (ej: 'leer:documento')
        cursor: Cursor de base de datos (opcional, si no se proporciona se crea uno nuevo)
    
    Returns:
        bool: True si el usuario tiene el permiso (individual o heredado), False en caso contrario
    """
    # Administrador tiene bypass total
    if id_rol == ADMIN_ROLE_ID:
        return True
    
    usar_cursor_externo = cursor is not None
    if not usar_cursor_externo:
        conn = get_db_connection()
        cursor = conn.cursor()
    
    try:
        # Paso 1: Verificar permisos individuales en usuario_permiso
        cursor.execute("""
            SELECT 1 FROM usuario_permiso
            WHERE id_usuario = %s AND nombre_permiso = %s
        """, (id_usuario, nombre_permiso))
        
        if cursor.fetchone():
            return True  # Tiene permiso individual
        
        # Paso 2: Si no tiene permiso individual, verificar permisos heredados del rol
        cursor.execute("""
            SELECT 1 FROM rol_permiso
            WHERE id_rol = %s AND nombre_permiso = %s
        """, (id_rol, nombre_permiso))
        
        tiene_permiso_rol = cursor.fetchone() is not None
        
        if not usar_cursor_externo:
            cursor.close()
            conn.close()
        
        return tiene_permiso_rol
            
    except Exception as e:
        app.logger.error(f"Error al verificar permiso: {str(e)}")
        if not usar_cursor_externo:
            cursor.close()
            conn.close()
        return False


def permiso_requerido(nombre_permiso):
    """
    Decorador que valida permisos granulares para endpoints.
    
    IMPORTANTE: Solo verifica permisos individuales del usuario (tabla usuario_permiso).
    El rol no influye en la autorización, solo sirve para pre-seleccionar permisos en la UI.
    
    Lógica:
    1. Valida JWT (ya hecho en before_request)
    2. Si es Administrador (id_rol = 2): BYPASS TOTAL (retorna True inmediatamente)
    3. Si NO es Administrador: Verifica SOLO en usuario_permiso (no en rol_permiso)
    4. Adjunta lista de residencias a g.residencias_acceso (ya hecho en before_request)
    
    Args:
        nombre_permiso: String del permiso (ej: "editar:tratamiento", "leer:residente")
        
    Returns:
        Decorador de función Flask
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # 1. Validación JWT ya hecha en before_request
            # g.id_usuario y g.id_rol ya están disponibles
            
            # 2. BYPASS para Administrador
            if g.id_rol == ADMIN_ROLE_ID:
                return f(*args, **kwargs)
            
            # 3. Verificar permiso (rol + permisos individuales del usuario)
            tiene_permiso = usuario_tiene_permiso(g.id_usuario, g.id_rol, nombre_permiso)
            
            if not tiene_permiso:
                log_security_event('acceso_denegado', g.id_usuario, {
                    'permiso_requerido': nombre_permiso,
                    'endpoint': request.path
                })
                return jsonify({
                    'error': 'No tienes permisos para realizar esta acción',
                    'permiso_requerido': nombre_permiso
                }), 403
            
            # 4. Continuar con la ejecución del endpoint
            return f(*args, **kwargs)
            
        return decorated_function
    return decorator


@app.before_request
def before_request():
    """
    Middleware que aplica autenticación a todas las rutas excepto las públicas.
    Valida el token JWT, carga residencias del usuario y valida cambio de contraseña.
    """
    # Rutas públicas que no requieren autenticación
    public_paths = ['/api/v1/login', '/health', '/']
    # Rutas que requieren autenticación pero permiten cambio de contraseña
    # Incluye rutas para permitir que el usuario actualice su propia información (incluyendo contraseña)
    rutas_cambio_clave = ['/api/v1/usuario/cambio-clave', '/api/v1/usuarios/me']
    # Rutas que permiten actualización del propio usuario (incluyendo cambio de contraseña inicial)
    rutas_actualizacion_propia = ['/api/v1/usuarios']
    
    # Excluir archivos estáticos, favicon y rutas de Chrome DevTools
    if (request.path in public_paths or 
        request.path.startswith('/static/') or 
        request.path == '/favicon.ico' or
        request.path == '/favicon.svg' or
        request.path.startswith('/.well-known/')):
        return None
    
    # Obtener token del header Authorization
    auth_header = request.headers.get('Authorization')
    if not auth_header:
        app.logger.warning(f"Petición sin token a {request.path}")
        return jsonify({'error': 'Token de autenticación requerido'}), 401
    
    # Verificar formato Bearer
    try:
        token = auth_header.split(' ')[1]  # "Bearer <token>"
        if not token:
            app.logger.warning(f"Token vacío en petición a {request.path}")
            return jsonify({'error': 'Formato de token inválido. Use: Bearer <token>'}), 401
    except IndexError:
        app.logger.warning(f"Formato de token inválido en petición a {request.path}: {auth_header[:20]}...")
        return jsonify({'error': 'Formato de token inválido. Use: Bearer <token>'}), 401
    
    # Validar y decodificar token
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=['HS256'])
        # Almacenar información del usuario en g para uso en las rutas
        g.id_usuario = payload.get('id_usuario')
        g.id_rol = payload.get('id_rol')
        
        app.logger.debug(f"Token válido para usuario {g.id_usuario}, rol {g.id_rol}, ruta {request.path}")
        
        # Validar que los campos requeridos estén presentes (YA NO incluye id_residencia)
        if not all([g.id_usuario, g.id_rol]):
            app.logger.warning(f"Token inválido: faltan campos. id_usuario={g.id_usuario}, id_rol={g.id_rol}")
            return jsonify({'error': 'Token inválido: faltan campos requeridos'}), 401
            
    except jwt.ExpiredSignatureError:
        app.logger.warning(f"Token expirado en petición a {request.path}")
        return jsonify({'error': 'Token expirado'}), 401
    except jwt.InvalidTokenError as e:
        app.logger.warning(f"Token inválido en petición a {request.path}: {str(e)}")
        return jsonify({'error': f'Token inválido: {str(e)}'}), 401
    except Exception as e:
        app.logger.error(f"Error inesperado al decodificar token: {str(e)}")
        return jsonify({'error': 'Error al validar token'}), 401
    
    # Cargar residencias del usuario desde usuario_residencia
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Si es Administrador, establecer lista vacía (bypass total)
        app.logger.debug(f"Usuario {g.id_usuario} con rol {g.id_rol}, ADMIN_ROLE_ID={ADMIN_ROLE_ID}")
        if g.id_rol == ADMIN_ROLE_ID:
            app.logger.debug(f"Usuario {g.id_usuario} es Administrador, estableciendo residencias_acceso = []")
            g.residencias_acceso = []  # Lista vacía = acceso total a todas las residencias
        else:
            # Verificar si la tabla usuario_residencia existe
            try:
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables 
                        WHERE table_schema = 'public'
                        AND table_name = 'usuario_residencia'
                    )
                """)
                resultado = cursor.fetchone()
                tabla_existe = resultado[0] if resultado else False
                
                if tabla_existe:
                    # Cargar residencias desde usuario_residencia
                    cursor.execute("""
                        SELECT ur.id_residencia 
                        FROM usuario_residencia ur
                        JOIN residencia r ON ur.id_residencia = r.id_residencia
                        WHERE ur.id_usuario = %s AND r.activa = TRUE
                    """, (g.id_usuario,))
                    
                    g.residencias_acceso = [row[0] for row in cursor.fetchall()]
                else:
                    # Modo legacy: obtener id_residencia directamente de usuario
                    cursor.execute("""
                        SELECT id_residencia 
                        FROM usuario 
                        WHERE id_usuario = %s
                    """, (g.id_usuario,))
                    resultado = cursor.fetchone()
                    g.residencias_acceso = [resultado[0]] if resultado and resultado[0] else []
                
                # Validar que el usuario tenga al menos una residencia asignada
                if not g.residencias_acceso:
                    return jsonify({
                        'error': 'Usuario sin residencias asignadas. Contacte al administrador.'
                    }), 403
            except Exception as e:
                app.logger.error(f"Error al cargar residencias del usuario: {str(e)}")
                # En caso de error, intentar modo legacy
                try:
                    cursor.execute("""
                        SELECT id_residencia 
                        FROM usuario 
                        WHERE id_usuario = %s
                    """, (g.id_usuario,))
                    resultado = cursor.fetchone()
                    g.residencias_acceso = [resultado[0]] if resultado and resultado[0] else []
                except Exception as e2:
                    app.logger.error(f"Error en modo legacy: {str(e2)}")
                    g.residencias_acceso = []
        
        # Validar cambio de contraseña obligatorio (excepto para rutas permitidas)
        # Permitir también actualización del propio usuario (para cambiar contraseña inicial)
        # Solo permitir si es PUT a su propia cuenta (para cambiar contraseña en primer login)
        es_actualizacion_propia = (request.path.startswith('/api/v1/usuarios/') and 
                                   request.method == 'PUT' and 
                                   request.path.split('/')[-1].isdigit() and
                                   int(request.path.split('/')[-1]) == g.id_usuario)
        if request.path not in rutas_cambio_clave and not es_actualizacion_propia:
            try:
                cursor.execute(
                    "SELECT requiere_cambio_clave FROM usuario WHERE id_usuario = %s",
                    (g.id_usuario,)
                )
                usuario = cursor.fetchone()
                if usuario and usuario[0]:  # Si requiere_cambio_clave = TRUE
                    return jsonify({
                        'error': 'Debes cambiar tu contraseña antes de continuar',
                        'requiere_cambio_clave': True
                    }), 403
            except Exception as e:
                app.logger.error(f"Error al validar cambio de contraseña: {str(e)}")
                # Si hay error, permitir continuar para evitar bloqueos
                
    except Exception as e:
        app.logger.error(f"Error en middleware before_request: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        # Si hay un error crítico en el middleware, devolver error 500
        return jsonify({'error': 'Error interno del servidor al validar autenticación'}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
    
    return None


@app.route('/')
def index():
    """Sirve la página principal del frontend."""
    return send_from_directory('static', 'index.html')


@app.route('/favicon.ico')
def favicon():
    """Maneja la petición del favicon."""
    return send_from_directory('static', 'favicon.svg', mimetype='image/svg+xml')


@app.route('/favicon.svg')
def favicon_svg():
    """Maneja la petición del favicon SVG."""
    return send_from_directory('static', 'favicon.svg', mimetype='image/svg+xml')


@app.route('/health', methods=['GET'])
def health_check():
    """
    Endpoint de health check para verificar el estado del servicio.
    """
    return jsonify({
        'status': 'ok',
        'service': 'Violetas Backend API',
        'timestamp': datetime.utcnow().isoformat()
    }), 200


@app.route('/api/v1/login', methods=['POST'])
def login():
    """
    Endpoint de autenticación.
    
    Recibe:
    {
        "email": "usuario@ejemplo.com",
        "password": "contraseña"
    }
    
    Retorna:
    {
        "token": "jwt_token_here"
    }
    
    IMPORTANTE: Solo acepta POST con JSON en el cuerpo. NO acepta credenciales en URL.
    """
    try:
        # SEGURIDAD: Rechazar explícitamente credenciales en query parameters
        if request.args.get('email') or request.args.get('password'):
            app.logger.warning(f"⚠️ INTENTO DE LOGIN CON CREDENCIALES EN URL desde IP: {request.remote_addr}")
            log_security_event('login_credenciales_url', None, {
                'ip': request.remote_addr,
                'user_agent': request.headers.get('User-Agent', 'Unknown')
            })
            return jsonify({
                'error': 'Las credenciales no pueden enviarse en la URL por razones de seguridad. Use POST con JSON en el cuerpo.'
            }), 400
        
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        email = data.get('email')
        password = data.get('password')
        
        if not email or not password:
            return jsonify({'error': 'Email y contraseña son requeridos'}), 400
        
        # Rate limiting: verificar intentos de login
        ip_address = request.remote_addr
        if not check_rate_limit(ip_address, max_attempts=5, window_minutes=1):
            log_security_event('login_rate_limit_excedido', None, {'email': email, 'ip': ip_address})
            return jsonify({
                'error': 'Demasiados intentos de login. Intenta nuevamente en 1 minuto.'
            }), 429
        
        # Conectar a la base de datos
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
        except Exception as e:
            app.logger.error(f"Error al conectar a la base de datos en login: {str(e)}")
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        
        try:
            app.logger.info(f"Intento de login para email: {email}")
            
            # Verificar si la columna requiere_cambio_clave existe
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_schema = 'public' 
                AND table_name = 'usuario'
                AND column_name = 'requiere_cambio_clave'
            """)
            tiene_requiere_cambio = cursor.fetchone() is not None
            
            # Construir query según columnas disponibles
            if tiene_requiere_cambio:
                query = "SELECT id_usuario, email, password_hash, id_rol, requiere_cambio_clave FROM usuario WHERE email = %s"
            else:
                query = "SELECT id_usuario, email, password_hash, id_rol FROM usuario WHERE email = %s"
            
            cursor.execute(query, (email,))
            usuario = cursor.fetchone()
            
            if not usuario:
                app.logger.warning(f"Usuario no encontrado: {email}")
                log_security_event('login_fallido', None, {'email': email, 'razon': 'usuario_no_encontrado'})
                return jsonify({'error': 'Credenciales inválidas'}), 401
            
            # Desempaquetar valores de forma segura
            id_usuario = usuario[0]
            email_db = usuario[1]
            password_hash = usuario[2]
            id_rol = usuario[3]
            requiere_cambio_clave = usuario[4] if tiene_requiere_cambio and len(usuario) > 4 else False
            
            app.logger.info(f"Usuario encontrado: id={id_usuario}, rol={id_rol}, requiere_cambio_clave={requiere_cambio_clave}")
            
            # Verificar contraseña
            try:
                password_valida = check_password_hash(password_hash, password)
                app.logger.debug(f"Verificación de contraseña: {password_valida}")
            except Exception as e:
                app.logger.error(f"Error al verificar contraseña: {str(e)}")
                password_valida = False
            
            if not password_valida:
                app.logger.warning(f"Contraseña incorrecta para usuario: {email}")
                log_security_event('login_fallido', id_usuario, {'email': email, 'razon': 'contraseña_incorrecta'})
                return jsonify({'error': 'Credenciales inválidas'}), 401
            
            # Login exitoso - generar token JWT (SIN id_residencia)
            payload = {
                'id_usuario': id_usuario,
                'id_rol': id_rol,
                'exp': datetime.utcnow() + timedelta(hours=24)
            }
            
            token = jwt.encode(payload, get_jwt_secret(), algorithm='HS256')
            
            log_security_event('login_exitoso', id_usuario, {'email': email, 'requiere_cambio_clave': requiere_cambio_clave})
            
            response = {
                'token': token
            }
            
            # Si requiere cambio de contraseña, agregar flag
            if requiere_cambio_clave:
                response['requiere_cambio_clave'] = True
                response['mensaje'] = 'Debes cambiar tu contraseña antes de continuar'
            
            return jsonify(response), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except ValueError as e:
        app.logger.error(f"Error de valor en login: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
    except Exception as e:
        app.logger.error(f"Error en login: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error interno del servidor: {str(e)}'}), 500


@app.route('/api/v1/usuario/cambio-clave', methods=['POST'])
def cambiar_clave():
    """
    Endpoint para cambiar la contraseña del usuario autenticado.
    Requiere la contraseña anterior para verificación.
    
    Request:
    {
        "password_actual": "contraseña_actual",
        "password_nuevo": "nueva_contraseña_segura"
    }
    
    Response:
    {
        "mensaje": "Contraseña actualizada exitosamente"
    }
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        password_actual = data.get('password_actual')
        password_nuevo = data.get('password_nuevo')
        
        if not password_actual or not password_nuevo:
            return jsonify({'error': 'Contraseña actual y nueva contraseña son requeridas'}), 400
        
        # Validar fuerza de la nueva contraseña
        is_valid, error_msg = validate_password_strength(password_nuevo)
        if not is_valid:
            return jsonify({'error': error_msg}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Obtener usuario actual
            cursor.execute(
                "SELECT id_usuario, password_hash FROM usuario WHERE id_usuario = %s",
                (g.id_usuario,)
            )
            usuario = cursor.fetchone()
            
            if not usuario:
                return jsonify({'error': 'Usuario no encontrado'}), 404
            
            id_usuario, password_hash_actual = usuario
            
            # Verificar contraseña actual
            if not check_password_hash(password_hash_actual, password_actual):
                log_security_event('cambio_clave_fallido', id_usuario, {'razon': 'contraseña_actual_incorrecta'})
                return jsonify({'error': 'Contraseña actual incorrecta'}), 401
            
            # Hashear nueva contraseña (usar pbkdf2:sha256 para compatibilidad con Python 3.9.6)
            password_hash_nuevo = generate_password_hash(password_nuevo, method='pbkdf2:sha256')
            
            # Actualizar contraseña y marcar que ya no requiere cambio
            cursor.execute("""
                UPDATE usuario
                SET password_hash = %s,
                    requiere_cambio_clave = FALSE
                WHERE id_usuario = %s
                RETURNING id_usuario
            """, (password_hash_nuevo, id_usuario))
            
            conn.commit()
            
            log_security_event('cambio_clave_exitoso', id_usuario)
            
            return jsonify({
                'mensaje': 'Contraseña actualizada exitosamente'
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al cambiar contraseña: {str(e)}")
            return jsonify({'error': 'Error al actualizar contraseña'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/usuarios', methods=['POST'])
@permiso_requerido('crear:usuario')  # Nota: Administrador tiene bypass automático
def crear_usuario():
    """
    Endpoint para crear nuevos usuarios.
    SOLO accesible por Administrador (tiene bypass en el decorador).
    
    Request:
    {
        "email": "usuario@ejemplo.com",
        "password": "contraseña_temporal",
        "id_rol": 2,
        "nombre": "Juan",
        "apellido": "Pérez",
        "residencias": [1, 2]  # Lista de id_residencia
    }
    """
    try:
        # Verificación adicional: solo Administrador puede crear usuarios
        if g.id_rol != ADMIN_ROLE_ID:
            return jsonify({'error': 'Solo administradores pueden crear usuarios'}), 403
        
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        email = data.get('email')
        password = data.get('password')
        id_rol = data.get('id_rol')
        residencias = data.get('residencias', [])
        permisos = data.get('permisos', [])  # Array de nombres de permisos personalizados
        
        # Validar campos requeridos con mensajes específicos
        if not email or not email.strip():
            return jsonify({'error': 'El email es requerido'}), 400
        
        if not password or not password.strip():
            return jsonify({'error': 'La contraseña es requerida'}), 400
        
        if not id_rol:
            return jsonify({'error': 'El rol es requerido'}), 400
        
        try:
            id_rol = int(id_rol)
        except (ValueError, TypeError):
            return jsonify({'error': 'El id_rol debe ser un número válido'}), 400
        
        if not residencias or len(residencias) == 0:
            return jsonify({'error': 'Debe asignar al menos una residencia'}), 400
        
        # Solo Administradores pueden crear otros Administradores
        if id_rol == ADMIN_ROLE_ID and g.id_rol != ADMIN_ROLE_ID:
            return jsonify({
                'error': 'Solo los Administradores pueden crear otros Administradores'
            }), 403
        
        # Validar formato de email
        is_valid_email, error_msg = validate_email(email)
        if not is_valid_email:
            return jsonify({'error': error_msg or 'Email inválido'}), 400
        
        # Validar fuerza de contraseña
        is_valid, error_msg = validate_password_strength(password)
        if not is_valid:
            return jsonify({'error': error_msg}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el email no exista
            cursor.execute("SELECT id_usuario FROM usuario WHERE email = %s", (email,))
            if cursor.fetchone():
                return jsonify({'error': 'El email ya está registrado'}), 409
            
            # Verificar que el rol existe
            cursor.execute("SELECT id_rol FROM rol WHERE id_rol = %s AND activo = TRUE", (id_rol,))
            if not cursor.fetchone():
                return jsonify({'error': 'Rol no encontrado o inactivo'}), 404
            
            # Verificar que todas las residencias existen y están activas
            placeholders = ','.join(['%s'] * len(residencias))
            cursor.execute(f"""
                SELECT id_residencia FROM residencia 
                WHERE id_residencia IN ({placeholders}) AND activa = TRUE
            """, tuple(residencias))
            
            residencias_validas = [row[0] for row in cursor.fetchall()]
            
            if len(residencias_validas) == 0:
                return jsonify({
                    'error': 'Debe asignar al menos una residencia activa. Ninguna de las residencias seleccionadas está activa.'
                }), 400
            
            if len(residencias_validas) != len(residencias):
                residencias_invalidas = [r for r in residencias if r not in residencias_validas]
                return jsonify({
                    'error': f'Una o más residencias no existen o están inactivas: {residencias_invalidas}. Solo se pueden asignar residencias activas.',
                    'residencias_invalidas': residencias_invalidas
                }), 400
            
            # Hashear contraseña (usar pbkdf2:sha256 para compatibilidad con Python 3.9.6)
            password_hash = generate_password_hash(password, method='pbkdf2:sha256')
            
            # Crear usuario
            cursor.execute("""
                INSERT INTO usuario (email, password_hash, id_rol, nombre, apellido, requiere_cambio_clave)
                VALUES (%s, %s, %s, %s, %s, TRUE)
                RETURNING id_usuario
            """, (
                email,
                password_hash,
                id_rol,
                data.get('nombre'),
                data.get('apellido')
            ))
            
            id_usuario = cursor.fetchone()[0]
            
            # Asignar residencias (solo activas)
            for id_residencia in residencias_validas:
                cursor.execute("""
                    INSERT INTO usuario_residencia (id_usuario, id_residencia)
                    VALUES (%s, %s)
                    ON CONFLICT DO NOTHING
                """, (id_usuario, id_residencia))
            
            # Verificar que el usuario tiene al menos una residencia activa después de la creación
            cursor.execute("""
                SELECT COUNT(*) 
                FROM usuario_residencia ur
                JOIN residencia r ON ur.id_residencia = r.id_residencia
                WHERE ur.id_usuario = %s AND r.activa = TRUE
            """, (id_usuario,))
            total_residencias_activas = cursor.fetchone()[0]
            
            if total_residencias_activas == 0:
                conn.rollback()
                return jsonify({
                    'error': 'Error: El usuario debe tener al menos una residencia activa. No se pudo completar la creación.'
                }), 400
            
            # Crear tabla usuario_permiso si no existe
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS usuario_permiso (
                    id_usuario INTEGER NOT NULL,
                    nombre_permiso VARCHAR(255) NOT NULL,
                    PRIMARY KEY (id_usuario, nombre_permiso),
                    FOREIGN KEY (id_usuario) REFERENCES usuario(id_usuario) ON DELETE CASCADE
                )
            """)
            
            # Asignar permisos personalizados si se proporcionaron
            if permisos and len(permisos) > 0:
                # Verificar que los permisos existen antes de insertarlos
                for nombre_permiso in permisos:
                    # Verificar que el permiso existe
                    cursor.execute("""
                        SELECT COUNT(*) FROM permiso WHERE nombre_permiso = %s AND activo = TRUE
                    """, (nombre_permiso,))
                    permiso_existe = cursor.fetchone()[0] > 0
                    
                    if not permiso_existe:
                        app.logger.warning(f"Permiso '{nombre_permiso}' no existe en la tabla permiso. Se omitirá.")
                        continue
                    
                    cursor.execute("""
                        INSERT INTO usuario_permiso (id_usuario, nombre_permiso)
                        VALUES (%s, %s)
                        ON CONFLICT DO NOTHING
                    """, (id_usuario, nombre_permiso))
            
            conn.commit()
            
            log_security_event('usuario_creado', g.id_usuario, {
                'usuario_creado_id': id_usuario,
                'email': email,
                'id_rol': id_rol,
                'residencias': residencias_validas
            })
            
            return jsonify({
                'id_usuario': id_usuario,
                'mensaje': 'Usuario creado exitosamente. Debe cambiar su contraseña en el primer login.',
                'email': email,
                'requiere_cambio_clave': True
            }), 201
            
        except Exception as e:
            conn.rollback()
            import traceback
            error_trace = traceback.format_exc()
            app.logger.error(f"Error al crear usuario: {str(e)}\n{error_trace}")
            # En desarrollo, incluir más detalles del error
            if app.config.get('DEBUG'):
                return jsonify({
                    'error': 'Error al crear usuario',
                    'detalle': str(e),
                    'traceback': error_trace.split('\n')[-5:]  # Últimas 5 líneas del traceback
                }), 500
            return jsonify({'error': 'Error al crear usuario. Por favor, verifica los datos e intenta nuevamente.'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        app.logger.error(f"Error en crear_usuario: {str(e)}\n{error_trace}")
        # En desarrollo, incluir más detalles del error
        if app.config.get('DEBUG'):
            return jsonify({
                'error': 'Error interno del servidor',
                'detalle': str(e),
                'traceback': error_trace.split('\n')[-5:]  # Últimas 5 líneas del traceback
            }), 500
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============================================================================
# ENDPOINTS DE RESIDENTES
# ============================================================================

@app.route('/api/v1/residencias/<int:id_residencia>/habitaciones-ocupadas', methods=['GET'])
def obtener_habitaciones_ocupadas(id_residencia):
    """Obtiene las habitaciones ocupadas de una residencia (solo residentes activos)."""
    try:
        # Verificar que el usuario tenga acceso a esta residencia
        is_valid, error_response = validate_residencia_access(id_residencia)
        if not is_valid:
            return error_response
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Consulta simplificada sin expresiones regulares complejas
            cursor.execute("""
                SELECT DISTINCT habitacion
                FROM residente
                WHERE id_residencia = %s
                  AND activo = TRUE
                  AND habitacion IS NOT NULL
                  AND habitacion != ''
                ORDER BY habitacion
            """, (id_residencia,))
            
            habitaciones_ocupadas = [str(row[0]) for row in cursor.fetchall()]
            
            # Ordenar numéricamente en Python si es posible
            def sort_key(h):
                try:
                    return (0, int(h))  # Numeros primero
                except ValueError:
                    return (1, h)  # Texto después
            
            habitaciones_ocupadas.sort(key=sort_key)
            
            return jsonify({
                'habitaciones_ocupadas': habitaciones_ocupadas,
                'total': len(habitaciones_ocupadas)
            }), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        import traceback
        error_msg = str(e)
        error_trace = traceback.format_exc()
        app.logger.error(f"Error al obtener habitaciones ocupadas: {error_msg}\n{error_trace}")
        return jsonify({
            'error': 'Error al obtener habitaciones ocupadas',
            'details': error_msg
        }), 500


@app.route('/api/v1/residentes', methods=['GET'])
def listar_residentes():
    """
    Lista todos los residentes ordenados por residencia y habitación.
    Muestra todas las residencias ordenadas.
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar qué columnas opcionales existen
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'residente' 
                  AND column_name IN ('grupo_sanguineo', 'alergias', 'diagnosticos', 'restricciones_dieteticas',
                                      'nivel_dependencia', 'movilidad', 'medico_referencia', 'telefono_medico',
                                      'motivo_baja', 'fecha_baja')
            """)
            columnas_existentes = {row[0] for row in cursor.fetchall()}
            
            # Construir la consulta dinámicamente
            campos_opcionales = []
            if 'grupo_sanguineo' in columnas_existentes:
                campos_opcionales.append('r.grupo_sanguineo')
            else:
                campos_opcionales.append('NULL as grupo_sanguineo')
            
            if 'alergias' in columnas_existentes:
                campos_opcionales.append('r.alergias')
            else:
                campos_opcionales.append('NULL as alergias')
            
            if 'diagnosticos' in columnas_existentes:
                campos_opcionales.append('r.diagnosticos')
            else:
                campos_opcionales.append('NULL as diagnosticos')
            
            if 'restricciones_dieteticas' in columnas_existentes:
                campos_opcionales.append('r.restricciones_dieteticas')
            else:
                campos_opcionales.append('NULL as restricciones_dieteticas')
            
            if 'nivel_dependencia' in columnas_existentes:
                campos_opcionales.append('r.nivel_dependencia')
            else:
                campos_opcionales.append('NULL as nivel_dependencia')
            
            if 'movilidad' in columnas_existentes:
                campos_opcionales.append('r.movilidad')
            else:
                campos_opcionales.append('NULL as movilidad')
            
            if 'medico_referencia' in columnas_existentes:
                campos_opcionales.append('r.medico_referencia')
            else:
                campos_opcionales.append('NULL as medico_referencia')
            
            if 'telefono_medico' in columnas_existentes:
                campos_opcionales.append('r.telefono_medico')
            else:
                campos_opcionales.append('NULL as telefono_medico')
            
            if 'motivo_baja' in columnas_existentes:
                campos_opcionales.append('r.motivo_baja')
            else:
                campos_opcionales.append('NULL as motivo_baja')
            
            if 'fecha_baja' in columnas_existentes:
                campos_opcionales.append('r.fecha_baja')
            else:
                campos_opcionales.append('NULL as fecha_baja')
            
            campos_opcionales_str = ', ' + ', '.join(campos_opcionales)
            
            # Construir query base
            query = f"""
                SELECT r.id_residente, r.id_residencia, r.nombre, r.apellido, r.documento_identidad, 
                       r.fecha_nacimiento, r.telefono, r.direccion, r.contacto_emergencia,
                       r.telefono_emergencia, r.activo, r.fecha_ingreso, r.habitacion,
                       r.costo_habitacion, r.servicios_extra, r.medicaciones, r.peculiaridades, 
                       r.metodo_pago_preferido, r.fecha_creacion
                       {campos_opcionales_str},
                       res.nombre as nombre_residencia
                FROM residente r
                JOIN residencia res ON r.id_residencia = res.id_residencia
            """
            
            # Filtrar por residencias según acceso del usuario
            params = []
            if g.id_rol == ADMIN_ROLE_ID:
                # Administrador: sin filtro
                pass
            else:
                # Usuario normal: filtrar por lista de residencias
                if not g.residencias_acceso:
                    return jsonify({'residentes': [], 'total': 0}), 200
                
                placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                query += f" WHERE r.id_residencia IN ({placeholders})"
                params = g.residencias_acceso
            
            query += """
                ORDER BY r.id_residencia, 
                         CASE 
                             WHEN r.habitacion ~ '^[0-9]+$' THEN r.habitacion::INTEGER
                             ELSE 999999
                         END,
                         r.habitacion
            """
            
            cursor.execute(query, tuple(params) if params else None)
            
            residentes = cursor.fetchall()
            
            # Índices base (campos que siempre existen)
            idx_base = 18  # Después de fecha_creacion
            
            resultado = []
            for res in residentes:
                idx = idx_base
                resultado.append({
                    'id_residente': res[0],
                    'id_residencia': res[1],
                    'nombre': res[2],
                    'apellido': res[3],
                    'documento_identidad': res[4],
                    'fecha_nacimiento': str(res[5]) if res[5] else None,
                    'telefono': res[6],
                    'direccion': res[7],
                    'contacto_emergencia': res[8],
                    'telefono_emergencia': res[9],
                    'activo': res[10],
                    'fecha_ingreso': str(res[11]) if res[11] else None,
                    'habitacion': res[12],
                    'costo_habitacion': float(res[13]) if res[13] else None,
                    'servicios_extra': res[14],
                    'medicaciones': res[15],
                    'peculiaridades': res[16],
                    'metodo_pago_preferido': res[17],
                    'fecha_creacion': res[18].isoformat() if res[18] else None,
                    'grupo_sanguineo': res[idx] if len(res) > idx else None,
                    'alergias': res[idx + 1] if len(res) > idx + 1 else None,
                    'diagnosticos': res[idx + 2] if len(res) > idx + 2 else None,
                    'restricciones_dieteticas': res[idx + 3] if len(res) > idx + 3 else None,
                    'nivel_dependencia': res[idx + 4] if len(res) > idx + 4 else None,
                    'movilidad': res[idx + 5] if len(res) > idx + 5 else None,
                    'medico_referencia': res[idx + 6] if len(res) > idx + 6 else None,
                    'telefono_medico': res[idx + 7] if len(res) > idx + 7 else None,
                    'motivo_baja': res[idx + 8] if len(res) > idx + 8 else None,
                    'fecha_baja': str(res[idx + 9]) if len(res) > idx + 9 and res[idx + 9] else None,
                    'nombre_residencia': res[idx + 10] if len(res) > idx + 10 else None
                })
            
            return jsonify({'residentes': resultado, 'total': len(resultado)}), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al listar residentes: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error al obtener residentes: {str(e)}'}), 500


@app.route('/api/v1/residentes/<int:id_residente>', methods=['GET'])
def obtener_residente(id_residente):
    """Obtiene un residente específico. Permite obtener residentes de cualquier residencia para poder editarlos."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar qué columnas opcionales existen
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'residente' 
                  AND column_name IN ('grupo_sanguineo', 'alergias', 'diagnosticos', 'restricciones_dieteticas',
                                      'nivel_dependencia', 'movilidad', 'medico_referencia', 'telefono_medico',
                                      'motivo_baja', 'fecha_baja')
            """)
            columnas_existentes = {row[0] for row in cursor.fetchall()}
            
            # Construir la consulta dinámicamente
            campos_opcionales = []
            if 'grupo_sanguineo' in columnas_existentes:
                campos_opcionales.append('r.grupo_sanguineo')
            else:
                campos_opcionales.append('NULL as grupo_sanguineo')
            
            if 'alergias' in columnas_existentes:
                campos_opcionales.append('r.alergias')
            else:
                campos_opcionales.append('NULL as alergias')
            
            if 'diagnosticos' in columnas_existentes:
                campos_opcionales.append('r.diagnosticos')
            else:
                campos_opcionales.append('NULL as diagnosticos')
            
            if 'restricciones_dieteticas' in columnas_existentes:
                campos_opcionales.append('r.restricciones_dieteticas')
            else:
                campos_opcionales.append('NULL as restricciones_dieteticas')
            
            if 'nivel_dependencia' in columnas_existentes:
                campos_opcionales.append('r.nivel_dependencia')
            else:
                campos_opcionales.append('NULL as nivel_dependencia')
            
            if 'movilidad' in columnas_existentes:
                campos_opcionales.append('r.movilidad')
            else:
                campos_opcionales.append('NULL as movilidad')
            
            if 'medico_referencia' in columnas_existentes:
                campos_opcionales.append('r.medico_referencia')
            else:
                campos_opcionales.append('NULL as medico_referencia')
            
            if 'telefono_medico' in columnas_existentes:
                campos_opcionales.append('r.telefono_medico')
            else:
                campos_opcionales.append('NULL as telefono_medico')
            
            if 'motivo_baja' in columnas_existentes:
                campos_opcionales.append('r.motivo_baja')
            else:
                campos_opcionales.append('NULL as motivo_baja')
            
            if 'fecha_baja' in columnas_existentes:
                campos_opcionales.append('r.fecha_baja')
            else:
                campos_opcionales.append('NULL as fecha_baja')
            
            campos_opcionales_str = ', ' + ', '.join(campos_opcionales)
            
            query = f"""
                SELECT r.id_residente, r.id_residencia, r.nombre, r.apellido, r.documento_identidad,
                       r.fecha_nacimiento, r.telefono, r.direccion, r.contacto_emergencia,
                       r.telefono_emergencia, r.activo, r.fecha_ingreso, r.habitacion,
                       r.costo_habitacion, r.servicios_extra, r.medicaciones, r.peculiaridades, 
                       r.metodo_pago_preferido, r.fecha_creacion
                       {campos_opcionales_str}
                FROM residente r
                WHERE r.id_residente = %s
            """
            
            cursor.execute(query, (id_residente,))
            
            res = cursor.fetchone()
            
            if not res:
                return jsonify({'error': 'Residente no encontrado'}), 404
            
            # Validar acceso a la residencia
            is_valid, error_response = validate_residencia_access(res[1])  # res[1] es id_residencia
            if not is_valid:
                return error_response
            
            # Índices base (campos que siempre existen)
            idx = 19  # Después de fecha_creacion
            
            resultado = {
                'id_residente': res[0],
                'id_residencia': res[1],
                'nombre': res[2],
                'apellido': res[3],
                'documento_identidad': res[4],
                'fecha_nacimiento': str(res[5]) if res[5] else None,
                'telefono': res[6],
                'direccion': res[7],
                'contacto_emergencia': res[8],
                'telefono_emergencia': res[9],
                'activo': res[10],
                'fecha_ingreso': str(res[11]) if res[11] else None,
                'habitacion': res[12],
                'costo_habitacion': float(res[13]) if res[13] else None,
                'servicios_extra': res[14],
                'medicaciones': res[15],
                'peculiaridades': res[16],
                'metodo_pago_preferido': res[17],
                'fecha_creacion': res[18].isoformat() if res[18] else None,
                'grupo_sanguineo': res[idx] if len(res) > idx else None,
                'alergias': res[idx + 1] if len(res) > idx + 1 else None,
                'diagnosticos': res[idx + 2] if len(res) > idx + 2 else None,
                'restricciones_dieteticas': res[idx + 3] if len(res) > idx + 3 else None,
                'nivel_dependencia': res[idx + 4] if len(res) > idx + 4 else None,
                'movilidad': res[idx + 5] if len(res) > idx + 5 else None,
                'medico_referencia': res[idx + 6] if len(res) > idx + 6 else None,
                'telefono_medico': res[idx + 7] if len(res) > idx + 7 else None,
                'motivo_baja': res[idx + 8] if len(res) > idx + 8 else None,
                'fecha_baja': str(res[idx + 9]) if len(res) > idx + 9 and res[idx + 9] else None
            }
            
            return jsonify(resultado), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al obtener residente: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error al obtener residente: {str(e)}'}), 500


@app.route('/api/v1/residentes', methods=['POST'])
def crear_residente():
    """
    Crea un nuevo residente. Permite elegir la residencia (Violetas 1 o Violetas 2).
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        # Validar datos con el módulo de validación
        is_valid, errors = validate_residente_data(data, is_update=False)
        if not is_valid:
            return jsonify({'error': 'Errores de validación', 'detalles': errors}), 400
        
        nombre = data.get('nombre')
        apellido = data.get('apellido')
        id_residencia = data.get('id_residencia')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que la residencia existe
            cursor.execute("SELECT id_residencia FROM residencia WHERE id_residencia = %s", (id_residencia,))
            if not cursor.fetchone():
                return jsonify({'error': 'Residencia no encontrada'}), 404
            
            cursor.execute("""
                INSERT INTO residente (id_residencia, nombre, apellido, documento_identidad,
                                     fecha_nacimiento, telefono, direccion, contacto_emergencia,
                                     telefono_emergencia, activo, fecha_ingreso, habitacion,
                                     costo_habitacion, servicios_extra, medicaciones, peculiaridades,
                                     metodo_pago_preferido, grupo_sanguineo, alergias, diagnosticos,
                                     restricciones_dieteticas, nivel_dependencia, movilidad,
                                     medico_referencia, telefono_medico)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id_residente, fecha_creacion
            """, (
                id_residencia,
                nombre,
                apellido,
                data.get('documento_identidad'),
                data.get('fecha_nacimiento'),
                data.get('telefono'),
                data.get('direccion'),
                data.get('contacto_emergencia'),
                data.get('telefono_emergencia'),
                data.get('activo', True),
                data.get('fecha_ingreso'),
                data.get('habitacion'),
                data.get('costo_habitacion'),
                data.get('servicios_extra'),
                data.get('medicaciones'),
                data.get('peculiaridades'),
                data.get('metodo_pago_preferido'),
                data.get('grupo_sanguineo'),
                data.get('alergias'),
                data.get('diagnosticos'),
                data.get('restricciones_dieteticas'),
                data.get('nivel_dependencia'),
                data.get('movilidad'),
                data.get('medico_referencia'),
                data.get('telefono_medico')
            ))
            
            resultado = cursor.fetchone()
            id_residente = resultado[0]
            
            # Generar automáticamente cobro del mes siguiente si tiene costo_habitacion
            costo_habitacion = data.get('costo_habitacion')
            if costo_habitacion and costo_habitacion > 0:
                try:
                    # Calcular mes siguiente
                    hoy = datetime.now()
                    if hoy.month == 12:
                        siguiente_mes = datetime(hoy.year + 1, 1, 1)
                    else:
                        siguiente_mes = datetime(hoy.year, hoy.month + 1, 1)
                    
                    mes_siguiente_str = siguiente_mes.strftime('%Y-%m')
                    fecha_prevista = siguiente_mes.date()  # Día 1 del mes siguiente
                    
                    meses_espanol = {
                        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
                        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
                        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
                    }
                    nombre_mes = meses_espanol.get(siguiente_mes.month, 'mes')
                    # Formato: "Diciembre 25", "Enero 26" (solo mes y año corto, sin "Pago")
                    año_corto = str(siguiente_mes.year)[-2:]  # Últimos 2 dígitos
                    concepto = f"{nombre_mes.capitalize()} {año_corto}"
                    metodo_pago = data.get('metodo_pago_preferido') or 'transferencia'
                    
                    # Verificar si ya existe un cobro mensual duplicado (pendiente o completado)
                    existe_duplicado, id_pago_duplicado = verificar_cobro_mensual_duplicado(
                        cursor, id_residente, id_residencia, mes_siguiente_str, concepto
                    )
                    
                    if not existe_duplicado:
                        # Crear el cobro previsto para el mes siguiente
                        cursor.execute("""
                            INSERT INTO pago_residente (
                                id_residente, id_residencia, monto, fecha_pago, fecha_prevista,
                                mes_pagado, concepto, metodo_pago, estado, es_cobro_previsto
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            id_residente,
                            id_residencia,
                            costo_habitacion,
                            None,  # fecha_pago es NULL para cobros previstos
                            fecha_prevista,
                            mes_siguiente_str,
                            concepto,
                            metodo_pago,
                            'pendiente',
                            True
                        ))
                        app.logger.info(f"Cobro previsto generado automáticamente para nuevo residente {nombre} {apellido} (ID: {id_residente}): €{costo_habitacion}, mes: {mes_siguiente_str}")
                except Exception as e:
                    app.logger.error(f"Error al generar cobro previsto automático para nuevo residente {id_residente}: {str(e)}")
                    # No fallar la creación del residente si falla la generación del cobro
            
            # Generar cobros históricos completados desde fecha_ingreso hasta mes anterior al actual
            fecha_ingreso = data.get('fecha_ingreso')
            if fecha_ingreso and costo_habitacion and costo_habitacion > 0:
                try:
                    cobros_historicos = generar_cobros_historicos_completados(
                        cursor, id_residente, id_residencia, fecha_ingreso, costo_habitacion, metodo_pago
                    )
                    if cobros_historicos > 0:
                        app.logger.info(f"Generados {cobros_historicos} cobros históricos completados para nuevo residente {nombre} {apellido} (ID: {id_residente})")
                except Exception as e:
                    app.logger.error(f"Error al generar cobros históricos para nuevo residente {id_residente}: {str(e)}")
                    # No fallar la creación del residente si falla la generación de históricos
            
            conn.commit()
            
            return jsonify({
                'id_residente': id_residente,
                'mensaje': 'Residente creado exitosamente'
            }), 201
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al crear residente: {str(e)}")
            return jsonify({'error': 'Error al crear residente'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/residentes/<int:id_residente>/baja', methods=['POST'])
def dar_baja_residente(id_residente):
    """Da de baja a un residente con motivo y fecha."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        motivo_baja = data.get('motivo_baja')
        if not motivo_baja:
            return jsonify({'error': 'El motivo de baja es requerido'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el residente existe, está activo y pertenece a la residencia del usuario
            cursor.execute("""
                SELECT id_residente, activo, id_residencia FROM residente
                WHERE id_residente = %s
            """, (id_residente,))
            
            residente = cursor.fetchone()
            if not residente:
                return jsonify({'error': 'Residente no encontrado'}), 404
            
            # Verificar que el usuario tiene acceso a la residencia del residente
            is_valid, error_response = validate_residencia_access(residente[2])
            if not is_valid:
                return error_response
            
            if not residente[1]:  # Si ya está inactivo
                return jsonify({'error': 'El residente ya está dado de baja'}), 400
            
            # Verificar si las columnas de baja existen
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'residente' 
                  AND column_name IN ('motivo_baja', 'fecha_baja')
            """)
            columnas_baja = {row[0] for row in cursor.fetchall()}
            
            # Actualizar el residente: activo = False, motivo_baja, fecha_baja = hoy
            from datetime import date
            fecha_baja = date.today()
            
            # PASO 1: Guardar los pagos pendientes que están después de la fecha de baja
            # Buscar pagos pendientes con fecha_pago o fecha_prevista posterior a fecha_baja
            # Verificar si existe la columna fecha_prevista
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'pago_residente' 
                  AND column_name = 'fecha_prevista'
            """)
            tiene_fecha_prevista = cursor.fetchone() is not None
            
            if tiene_fecha_prevista:
                cursor.execute("""
                    SELECT id_pago, id_residente, id_residencia, monto, fecha_pago, mes_pagado, 
                           concepto, metodo_pago, estado, fecha_creacion, fecha_prevista
                    FROM pago_residente
                    WHERE id_residente = %s
                      AND estado IN ('pendiente', 'previsto')
                      AND (fecha_pago > %s OR fecha_prevista > %s)
                """, (id_residente, fecha_baja, fecha_baja))
            else:
                cursor.execute("""
                    SELECT id_pago, id_residente, id_residencia, monto, fecha_pago, mes_pagado, 
                           concepto, metodo_pago, estado, fecha_creacion
                    FROM pago_residente
                    WHERE id_residente = %s
                      AND estado IN ('pendiente', 'previsto')
                      AND fecha_pago > %s
                """, (id_residente, fecha_baja))
            
            pagos_a_eliminar = cursor.fetchall()
            
            # Guardar los pagos en la tabla auxiliar antes de eliminarlos
            if pagos_a_eliminar:
                try:
                    # Verificar si la tabla pago_residente_baja existe
                    cursor.execute("""
                        SELECT EXISTS (
                            SELECT FROM information_schema.tables 
                            WHERE table_schema = 'public'
                            AND table_name = 'pago_residente_baja'
                        )
                    """)
                    tabla_existe = cursor.fetchone()[0]
                    
                    if tabla_existe:
                        for pago in pagos_a_eliminar:
                            # Verificar si la tabla auxiliar tiene fecha_prevista
                            cursor.execute("""
                                SELECT column_name 
                                FROM information_schema.columns 
                                WHERE table_name = 'pago_residente_baja' 
                                  AND column_name = 'fecha_prevista'
                            """)
                            tabla_tiene_fecha_prevista = cursor.fetchone() is not None
                            
                            if tabla_tiene_fecha_prevista and tiene_fecha_prevista and len(pago) > 10:
                                cursor.execute("""
                                    INSERT INTO pago_residente_baja 
                                    (id_pago, id_residente, id_residencia, monto, fecha_pago, mes_pagado,
                                     concepto, metodo_pago, estado, fecha_creacion, fecha_baja_residente, fecha_prevista)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                """, (pago[0], pago[1], pago[2], pago[3], pago[4], pago[5], 
                                      pago[6], pago[7], pago[8], pago[9], fecha_baja, pago[10] if len(pago) > 10 else None))
                            else:
                                cursor.execute("""
                                    INSERT INTO pago_residente_baja 
                                    (id_pago, id_residente, id_residencia, monto, fecha_pago, mes_pagado,
                                     concepto, metodo_pago, estado, fecha_creacion, fecha_baja_residente)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                """, (pago[0], pago[1], pago[2], pago[3], pago[4], pago[5], 
                                      pago[6], pago[7], pago[8], pago[9], fecha_baja))
                        
                        # Eliminar los pagos pendientes posteriores a la fecha de baja
                        if tiene_fecha_prevista:
                            cursor.execute("""
                                DELETE FROM pago_residente
                                WHERE id_residente = %s
                                  AND estado IN ('pendiente', 'previsto')
                                  AND (fecha_pago > %s OR fecha_prevista > %s)
                            """, (id_residente, fecha_baja, fecha_baja))
                        else:
                            cursor.execute("""
                                DELETE FROM pago_residente
                                WHERE id_residente = %s
                                  AND estado IN ('pendiente', 'previsto')
                                  AND fecha_pago > %s
                            """, (id_residente, fecha_baja))
                        
                        app.logger.info(f"Eliminados {len(pagos_a_eliminar)} pagos pendientes del residente {id_residente} por baja")
                except Exception as e:
                    app.logger.warning(f"Error al eliminar pagos pendientes (tabla puede no existir): {str(e)}")
                    # Continuar con la baja aunque falle la eliminación de pagos
            
            # Construir la consulta dinámicamente según las columnas que existan
            # Al dar de baja, NO se libera la habitación (se mantiene para referencia histórica)
            # El sistema de habitaciones ocupadas filtra por activo = TRUE, así que no hay conflicto
            updates = ['activo = FALSE']
            valores = []
            
            if 'motivo_baja' in columnas_baja:
                updates.append('motivo_baja = %s')
                valores.append(motivo_baja)
            
            if 'fecha_baja' in columnas_baja:
                updates.append('fecha_baja = %s')
                valores.append(fecha_baja)
            
            valores.append(id_residente)
            
            query = f"""
                UPDATE residente
                SET {', '.join(updates)}
                WHERE id_residente = %s
                RETURNING id_residente
            """
            
            cursor.execute(query, tuple(valores))
            conn.commit()
            
            return jsonify({
                'mensaje': 'Residente dado de baja exitosamente',
                'id_residente': id_residente,
                'motivo_baja': motivo_baja if 'motivo_baja' in columnas_baja else None,
                'fecha_baja': str(fecha_baja) if 'fecha_baja' in columnas_baja else None
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al dar de baja al residente: {str(e)}")
            return jsonify({'error': 'Error al dar de baja al residente'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/residentes/<int:id_residente>/alta', methods=['POST'])
def dar_alta_residente(id_residente):
    """Reactiva un residente que estaba de baja (baja por error)."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el residente existe y está inactivo
            cursor.execute("""
                SELECT id_residente, activo, id_residencia FROM residente
                WHERE id_residente = %s
            """, (id_residente,))
            
            residente = cursor.fetchone()
            if not residente:
                return jsonify({'error': 'Residente no encontrado'}), 404
            
            # Verificar que el usuario tiene acceso a la residencia del residente
            is_valid, error_response = validate_residencia_access(residente[2])
            if not is_valid:
                return error_response
            
            if residente[1]:  # Si ya está activo
                return jsonify({'error': 'El residente ya está activo'}), 400
            
            # Verificar si las columnas de baja existen
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'residente' 
                  AND column_name IN ('motivo_baja', 'fecha_baja')
            """)
            columnas_baja = {row[0] for row in cursor.fetchall()}
            
            # PASO 1: Restaurar los pagos que fueron eliminados por la baja
            try:
                # Verificar si la tabla pago_residente_baja existe
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables 
                        WHERE table_schema = 'public'
                        AND table_name = 'pago_residente_baja'
                    )
                """)
                tabla_existe = cursor.fetchone()[0]
                
                if tabla_existe:
                    # Buscar pagos eliminados que aún no han sido restaurados
                    # Verificar si la tabla tiene fecha_prevista
                    cursor.execute("""
                        SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_name = 'pago_residente_baja' 
                          AND column_name = 'fecha_prevista'
                    """)
                    tabla_tiene_fecha_prevista = cursor.fetchone() is not None
                    
                    if tabla_tiene_fecha_prevista:
                        cursor.execute("""
                            SELECT id_pago, id_residente, id_residencia, monto, fecha_pago, mes_pagado,
                                   concepto, metodo_pago, estado, fecha_creacion, fecha_prevista
                            FROM pago_residente_baja
                            WHERE id_residente = %s
                              AND fecha_restauracion IS NULL
                        """, (id_residente,))
                    else:
                        cursor.execute("""
                            SELECT id_pago, id_residente, id_residencia, monto, fecha_pago, mes_pagado,
                                   concepto, metodo_pago, estado, fecha_creacion
                            FROM pago_residente_baja
                            WHERE id_residente = %s
                              AND fecha_restauracion IS NULL
                        """, (id_residente,))
                    
                    pagos_a_restaurar = cursor.fetchall()
                    
                    if pagos_a_restaurar:
                        from datetime import datetime
                        fecha_restauracion = datetime.now()
                        
                        for pago in pagos_a_restaurar:
                            # Verificar si el pago original aún existe (por si acaso)
                            cursor.execute("""
                                SELECT id_pago FROM pago_residente WHERE id_pago = %s
                            """, (pago[0],))
                            
                            if not cursor.fetchone():
                                # El pago fue eliminado, restaurarlo
                                # Verificar si existe fecha_prevista en la tabla pago_residente
                                cursor.execute("""
                                    SELECT column_name 
                                    FROM information_schema.columns 
                                    WHERE table_name = 'pago_residente' 
                                      AND column_name = 'fecha_prevista'
                                """)
                                tiene_fecha_prevista_tabla = cursor.fetchone() is not None
                                
                                if tiene_fecha_prevista_tabla and tabla_tiene_fecha_prevista and len(pago) > 10:
                                    cursor.execute("""
                                        INSERT INTO pago_residente 
                                        (id_residente, id_residencia, monto, fecha_pago, fecha_prevista, mes_pagado,
                                         concepto, metodo_pago, estado, fecha_creacion)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                        RETURNING id_pago
                                    """, (pago[1], pago[2], pago[3], pago[4], pago[10] if len(pago) > 10 else None, 
                                          pago[5], pago[6], pago[7], pago[8], pago[9]))
                                else:
                                    cursor.execute("""
                                        INSERT INTO pago_residente 
                                        (id_residente, id_residencia, monto, fecha_pago, mes_pagado,
                                         concepto, metodo_pago, estado, fecha_creacion)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                                        RETURNING id_pago
                                    """, (pago[1], pago[2], pago[3], pago[4], pago[5], 
                                          pago[6], pago[7], pago[8], pago[9]))
                                
                                nuevo_id_pago = cursor.fetchone()[0]
                                
                                # Actualizar el registro en pago_residente_baja con la fecha de restauración
                                cursor.execute("""
                                    UPDATE pago_residente_baja
                                    SET fecha_restauracion = %s, id_pago = %s
                                    WHERE id_registro = (
                                        SELECT id_registro FROM pago_residente_baja
                                        WHERE id_residente = %s
                                          AND fecha_restauracion IS NULL
                                          AND id_pago = %s
                                        LIMIT 1
                                    )
                                """, (fecha_restauracion, nuevo_id_pago, id_residente, pago[0]))
                            else:
                                # El pago ya existe, solo marcar como restaurado
                                cursor.execute("""
                                    UPDATE pago_residente_baja
                                    SET fecha_restauracion = %s
                                    WHERE id_residente = %s
                                      AND id_pago = %s
                                      AND fecha_restauracion IS NULL
                                """, (fecha_restauracion, id_residente, pago[0]))
                        
                        app.logger.info(f"Restaurados {len(pagos_a_restaurar)} pagos del residente {id_residente} al dar de alta")
            except Exception as e:
                app.logger.warning(f"Error al restaurar pagos (tabla puede no existir): {str(e)}")
                # Continuar con el alta aunque falle la restauración de pagos
            
            # Reactivar el residente: activo = True, limpiar motivo_baja y fecha_baja
            updates = ['activo = TRUE']
            valores = []
            
            if 'motivo_baja' in columnas_baja:
                updates.append('motivo_baja = NULL')
            
            if 'fecha_baja' in columnas_baja:
                updates.append('fecha_baja = NULL')
            
            valores.append(id_residente)
            
            query = f"""
                UPDATE residente
                SET {', '.join(updates)}
                WHERE id_residente = %s
                RETURNING id_residente
            """
            
            cursor.execute(query, tuple(valores))
            conn.commit()
            
            return jsonify({
                'mensaje': 'Residente reactivado exitosamente',
                'id_residente': id_residente
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al reactivar al residente: {str(e)}")
            import traceback
            app.logger.error(traceback.format_exc())
            return jsonify({'error': 'Error al reactivar al residente', 'details': str(e)}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/residentes/<int:id_residente>', methods=['DELETE'])
def eliminar_residente(id_residente):
    """Elimina completamente un residente de la base de datos."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el residente existe
            cursor.execute("""
                SELECT id_residente, id_residencia FROM residente
                WHERE id_residente = %s
            """, (id_residente,))
            
            residente = cursor.fetchone()
            if not residente:
                return jsonify({'error': 'Residente no encontrado'}), 404
            
            # Verificar que el usuario tiene acceso a la residencia del residente
            is_valid, error_response = validate_residencia_access(residente[1])
            if not is_valid:
                return error_response
            
            # Verificar que está dado de baja antes de eliminar
            cursor.execute("""
                SELECT activo FROM residente
                WHERE id_residente = %s
            """, (id_residente,))
            
            activo = cursor.fetchone()[0]
            if activo:
                return jsonify({'error': 'No se puede eliminar un residente activo. Primero debe darse de baja.'}), 400
            
            # Eliminar documentos asociados (si existe la tabla)
            try:
                cursor.execute("""
                    DELETE FROM documento_residente
                    WHERE id_residente = %s
                """, (id_residente,))
            except Exception as e:
                app.logger.warning(f"Error al eliminar documentos del residente {id_residente}: {str(e)}")
            
            # Eliminar pagos asociados antes de eliminar el residente
            # Esto es necesario porque hay una foreign key constraint
            try:
                cursor.execute("""
                    DELETE FROM pago_residente
                    WHERE id_residente = %s
                """, (id_residente,))
                app.logger.info(f"Pagos eliminados para residente {id_residente}: {cursor.rowcount}")
            except Exception as e:
                app.logger.warning(f"Error al eliminar pagos del residente {id_residente}: {str(e)}")
                # Si falla, intentar continuar de todas formas
            
            # Eliminar el residente
            cursor.execute("""
                DELETE FROM residente
                WHERE id_residente = %s
                RETURNING id_residente
            """, (id_residente,))
            
            conn.commit()
            
            return jsonify({
                'mensaje': 'Residente eliminado completamente',
                'id_residente': id_residente
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al eliminar al residente: {str(e)}")
            import traceback
            app.logger.error(traceback.format_exc())
            return jsonify({'error': 'Error al eliminar al residente', 'details': str(e)}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/residentes/<int:id_residente>', methods=['PUT'])
def actualizar_residente(id_residente):
    """Actualiza un residente (solo de la residencia del usuario)."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        # Validar datos con el módulo de validación
        is_valid, errors = validate_residente_data(data, is_update=True)
        if not is_valid:
            return jsonify({'error': 'Errores de validación', 'detalles': errors}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Si se está cambiando la residencia, validar que sea 1 o 2
            nueva_residencia = data.get('id_residencia')
            if nueva_residencia is not None:
                valid, error = validate_residencia_id(nueva_residencia)
                if not valid:
                    return jsonify({'error': error}), 400
                # Verificar que la residencia existe
                cursor.execute("SELECT id_residencia FROM residencia WHERE id_residencia = %s", (nueva_residencia,))
                if not cursor.fetchone():
                    return jsonify({'error': 'Residencia no encontrada'}), 404
            
            # Verificar que el residente existe (sin filtrar por residencia para permitir cambio)
            cursor.execute("""
                SELECT id_residente, id_residencia FROM residente
                WHERE id_residente = %s
            """, (id_residente,))
            
            residente_actual = cursor.fetchone()
            if not residente_actual:
                return jsonify({'error': 'Residente no encontrado'}), 404
            
            # Validar acceso a la residencia actual del residente
            is_valid, error_response = validate_residencia_access(residente_actual[1])  # residente_actual[1] es id_residencia
            if not is_valid:
                return error_response
            
            # La residencia actual del residente (para el WHERE)
            id_residencia_actual = residente_actual[1]
            
            # Verificar qué columnas opcionales existen en la base de datos
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'residente' 
                  AND column_name IN ('grupo_sanguineo', 'alergias', 'diagnosticos', 'restricciones_dieteticas',
                                      'nivel_dependencia', 'movilidad', 'medico_referencia', 'telefono_medico',
                                      'motivo_baja', 'fecha_baja')
            """)
            columnas_existentes = {row[0] for row in cursor.fetchall()}
            
            # Campos base que siempre existen
            campos_base = [
                'id_residencia', 'nombre', 'apellido', 'documento_identidad', 'fecha_nacimiento',
                'telefono', 'direccion', 'contacto_emergencia', 'telefono_emergencia',
                'fecha_ingreso', 'habitacion', 'costo_habitacion',
                'servicios_extra', 'medicaciones', 'peculiaridades', 'metodo_pago_preferido'
            ]
            
            # Campos opcionales que pueden o no existir
            campos_opcionales = [
                'grupo_sanguineo', 'alergias', 'diagnosticos', 'restricciones_dieteticas',
                'nivel_dependencia', 'movilidad', 'medico_referencia', 'telefono_medico'
            ]
            
            # Construir lista completa de campos actualizables (solo los que existen)
            campos_actualizables = campos_base.copy()
            for campo in campos_opcionales:
                if campo in columnas_existentes:
                    campos_actualizables.append(campo)
            
            updates = []
            valores = []
            
            for campo in campos_actualizables:
                if campo in data:
                    updates.append(f"{campo} = %s")
                    valores.append(data[campo])
            
            if not updates:
                return jsonify({'error': 'No hay campos para actualizar'}), 400
            
            # Usar la residencia ACTUAL para el WHERE (antes del cambio)
            valores.extend([id_residente, id_residencia_actual])
            
            query = f"""
                UPDATE residente
                SET {', '.join(updates)}
                WHERE id_residente = %s AND id_residencia = %s
                RETURNING id_residente
            """
            
            # Obtener el costo_habitacion y fecha_ingreso ANTES de actualizar para comparar
            cursor.execute("""
                SELECT costo_habitacion, fecha_ingreso FROM residente
                WHERE id_residente = %s
            """, (id_residente,))
            datos_anteriores_result = cursor.fetchone()
            costo_anterior = float(datos_anteriores_result[0]) if datos_anteriores_result and datos_anteriores_result[0] else None
            fecha_ingreso_anterior = datos_anteriores_result[1] if datos_anteriores_result and datos_anteriores_result[1] else None
            
            cursor.execute(query, valores)
            
            # Si se actualizó costo_habitacion, actualizar cobros pendientes futuros
            costo_habitacion_nuevo = data.get('costo_habitacion')
            id_residencia_final = data.get('id_residencia', id_residencia_actual)
            
            # Verificar si cambió el costo_habitacion
            if costo_habitacion_nuevo is not None and costo_anterior is not None:
                costo_nuevo_float = float(costo_habitacion_nuevo)
                if abs(costo_nuevo_float - costo_anterior) > 0.01:  # Si cambió significativamente
                    # Actualizar todos los cobros pendientes (futuros) del residente con concepto de mes
                    cursor.execute("""
                        UPDATE pago_residente
                        SET monto = %s
                        WHERE id_residente = %s
                          AND id_residencia = %s
                          AND estado = 'pendiente'
                          AND (concepto ILIKE 'enero %%' OR concepto ILIKE 'febrero %%' OR concepto ILIKE 'marzo %%' 
                               OR concepto ILIKE 'abril %%' OR concepto ILIKE 'mayo %%' OR concepto ILIKE 'junio %%'
                               OR concepto ILIKE 'julio %%' OR concepto ILIKE 'agosto %%' OR concepto ILIKE 'septiembre %%'
                               OR concepto ILIKE 'octubre %%' OR concepto ILIKE 'noviembre %%' OR concepto ILIKE 'diciembre %%'
                               OR concepto ILIKE 'Pago %%')
                    """, (costo_nuevo_float, id_residente, id_residencia_final))
                    cobros_actualizados = cursor.rowcount
                    if cobros_actualizados > 0:
                        app.logger.info(f"Actualizados {cobros_actualizados} cobros pendientes del residente {id_residente} con nuevo costo: €{costo_anterior} → €{costo_nuevo_float}")
            
            # Obtener el costo_habitacion actualizado o el existente
            cursor.execute("""
                SELECT costo_habitacion, metodo_pago_preferido FROM residente
                WHERE id_residente = %s
            """, (id_residente,))
            residente_actualizado = cursor.fetchone()
            costo_actual = float(residente_actualizado[0]) if residente_actualizado and residente_actualizado[0] else None
            metodo_pago_actual = residente_actualizado[1] if residente_actualizado else None
            
            # Generar cobro del mes siguiente si:
            # 1. Tiene costo_habitacion > 0
            # 2. No tiene ya un cobro previsto para el mes siguiente
            if costo_actual and costo_actual > 0:
                try:
                    # Calcular mes siguiente
                    hoy = datetime.now()
                    if hoy.month == 12:
                        siguiente_mes = datetime(hoy.year + 1, 1, 1)
                    else:
                        siguiente_mes = datetime(hoy.year, hoy.month + 1, 1)
                    
                    mes_siguiente_str = siguiente_mes.strftime('%Y-%m')
                    fecha_prevista = siguiente_mes.date()  # Día 1 del mes siguiente
                    
                    # Verificar si ya existe un cobro mensual duplicado (pendiente o completado)
                    existe_duplicado, id_pago_duplicado = verificar_cobro_mensual_duplicado(
                        cursor, id_residente, id_residencia_final, mes_siguiente_str, concepto
                    )
                    
                    if not existe_duplicado:
                        # Crear el cobro previsto para el mes siguiente
                        meses_espanol = {
                            1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
                            5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
                            9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
                        }
                        nombre_mes = meses_espanol.get(siguiente_mes.month, 'mes')
                        # Formato: "Diciembre 25", "Enero 26" (solo mes y año corto, sin "Pago")
                        año_corto = str(siguiente_mes.year)[-2:]  # Últimos 2 dígitos
                        concepto = f"{nombre_mes.capitalize()} {año_corto}"
                        metodo_pago = metodo_pago_actual or 'transferencia'
                        
                        cursor.execute("""
                            INSERT INTO pago_residente (
                                id_residente, id_residencia, monto, fecha_pago, fecha_prevista,
                                mes_pagado, concepto, metodo_pago, estado, es_cobro_previsto
                            )
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            id_residente,
                            id_residencia_final,
                            costo_actual,
                            None,  # fecha_pago es NULL para cobros previstos
                            fecha_prevista,
                            mes_siguiente_str,
                            concepto,
                            metodo_pago,
                            'pendiente',
                            True
                        ))
                        app.logger.info(f"Cobro previsto generado automáticamente para residente actualizado (ID: {id_residente}): €{costo_actual}, mes: {mes_siguiente_str}")
                except Exception as e:
                    app.logger.error(f"Error al generar cobro previsto automático para residente actualizado {id_residente}: {str(e)}")
                    # No fallar la actualización del residente si falla la generación del cobro
            
            # Si se actualizó fecha_ingreso, eliminar cobros completados anteriores a la nueva fecha
            fecha_ingreso_nueva = data.get('fecha_ingreso')
            if 'fecha_ingreso' in data and fecha_ingreso_nueva:
                try:
                    # Convertir fecha_ingreso_nueva a date si es string
                    if isinstance(fecha_ingreso_nueva, str):
                        fecha_ingreso_nueva_date = datetime.strptime(fecha_ingreso_nueva, '%Y-%m-%d').date()
                    else:
                        fecha_ingreso_nueva_date = fecha_ingreso_nueva
                    
                    # Calcular mes de ingreso (primer día del mes)
                    mes_ingreso_nuevo = datetime(fecha_ingreso_nueva_date.year, fecha_ingreso_nueva_date.month, 1).date()
                    
                    # Eliminar TODOS los cobros completados que tengan fecha_pago anterior al mes de ingreso nuevo
                    # Esto asegura que si se cambió la fecha de ingreso por error, se eliminen los cobros incorrectos
                    cursor.execute("""
                        DELETE FROM pago_residente
                        WHERE id_residente = %s
                          AND estado = 'cobrado'
                          AND fecha_pago IS NOT NULL
                          AND fecha_pago < %s
                    """, (id_residente, mes_ingreso_nuevo))
                    cobros_eliminados = cursor.rowcount
                    if cobros_eliminados > 0:
                        app.logger.info(f"Eliminados {cobros_eliminados} cobros completados anteriores a la nueva fecha de ingreso ({mes_ingreso_nuevo}) para residente {id_residente}")
                        
                except Exception as e:
                    app.logger.error(f"Error al eliminar cobros anteriores a nueva fecha de ingreso para residente {id_residente}: {str(e)}")
                    # Continuar aunque falle la eliminación
            
            # Generar cobros históricos completados si se actualizó fecha_ingreso o si no existen históricos previos
            if fecha_ingreso_nueva and costo_actual and costo_actual > 0:
                try:
                    # Verificar si ya existen cobros históricos completados para este residente
                    cursor.execute("""
                        SELECT COUNT(*) FROM pago_residente
                        WHERE id_residente = %s
                          AND estado = 'cobrado'
                          AND es_cobro_previsto = FALSE
                    """, (id_residente,))
                    cobros_historicos_existentes = cursor.fetchone()[0]
                    
                    # Si no hay cobros históricos o se actualizó fecha_ingreso, generar históricos
                    if cobros_historicos_existentes == 0 or 'fecha_ingreso' in data:
                        cobros_historicos = generar_cobros_historicos_completados(
                            cursor, id_residente, id_residencia_final, fecha_ingreso_nueva, costo_actual, metodo_pago_actual
                        )
                        if cobros_historicos > 0:
                            app.logger.info(f"Generados {cobros_historicos} cobros históricos completados para residente actualizado (ID: {id_residente})")
                except Exception as e:
                    app.logger.error(f"Error al generar cobros históricos para residente actualizado {id_residente}: {str(e)}")
                    # No fallar la actualización del residente si falla la generación de históricos
            
            conn.commit()
            
            return jsonify({'mensaje': 'Residente actualizado exitosamente'}), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al actualizar residente: {str(e)}")
            return jsonify({'error': 'Error al actualizar residente'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============================================================================
# ENDPOINTS DE FACTURACIÓN - COBROS PREVISTOS DE RESIDENTES
# ============================================================================

@app.route('/api/v1/facturacion/cobros', methods=['GET'])
@permiso_requerido('leer:cobro')
def listar_cobros():
    """
    Lista los cobros del período cercano (Facturación):
    - Todos los cobros pendientes
    - El último cobro completado de cada residente
    - Cobros completados del mes actual y mes anterior
    - Todo lo demás va a Históricos
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Calcular fechas de referencia
            hoy = datetime.now()
            mes_actual_inicio = datetime(hoy.year, hoy.month, 1).date()
            mes_anterior_inicio = (mes_actual_inicio - timedelta(days=1)).replace(day=1)
            
            # Construir filtro de residencias
            where_clause, params = build_residencia_filter('p.', 'id_residencia')
            
            # NUEVA LÓGICA: Filtrar solo período cercano
            # 1. Todos los cobros pendientes
            # 2. Último cobro completado de cada residente
            # 3. Cobros completados del mes actual y mes anterior
            
            # Guardar la query para poder reutilizarla después de generar cobros pendientes
            query = None
            params_query = None
            
            if where_clause:
                # Usuario normal: filtrar por residencias asignadas
                # where_clause ya incluye "WHERE", así que lo usamos directamente
                # Usar subquery para DISTINCT ON ya que no puede estar directamente en UNION ALL
                query = f"""
                    WITH cobros_pendientes AS (
                        -- Todos los cobros pendientes
                    SELECT p.id_pago, p.id_residente, r.nombre || ' ' || r.apellido as residente,
                           p.monto, p.fecha_pago, p.fecha_prevista, p.mes_pagado, p.concepto,
                           p.metodo_pago, p.estado, p.es_cobro_previsto, p.observaciones, p.fecha_creacion,
                               res.id_residencia, res.nombre as nombre_residencia,
                               1 as orden_prioridad
                    FROM pago_residente p
                    JOIN residente r ON p.id_residente = r.id_residente
                    JOIN residencia res ON p.id_residencia = res.id_residencia
                        {where_clause}
                          AND p.estado = 'pendiente'
                    ),
                    ultimos_cobros_completados AS (
                        -- Último cobro completado de cada residente
                        SELECT DISTINCT ON (p.id_residente)
                               p.id_pago, p.id_residente, r.nombre || ' ' || r.apellido as residente,
                               p.monto, p.fecha_pago, p.fecha_prevista, p.mes_pagado, p.concepto,
                               p.metodo_pago, p.estado, p.es_cobro_previsto, p.observaciones, p.fecha_creacion,
                               res.id_residencia, res.nombre as nombre_residencia,
                               2 as orden_prioridad
                        FROM pago_residente p
                        JOIN residente r ON p.id_residente = r.id_residente
                        JOIN residencia res ON p.id_residencia = res.id_residencia
                        {where_clause}
                          AND p.estado = 'cobrado'
                          AND p.fecha_pago IS NOT NULL
                        ORDER BY p.id_residente, p.fecha_pago DESC, p.fecha_creacion DESC
                    ),
                    cobros_mes_actual_anterior AS (
                        -- Cobros completados del mes actual y mes anterior (excluyendo los que ya están en ultimos_cobros_completados)
                        SELECT p.id_pago, p.id_residente, r.nombre || ' ' || r.apellido as residente,
                               p.monto, p.fecha_pago, p.fecha_prevista, p.mes_pagado, p.concepto,
                               p.metodo_pago, p.estado, p.es_cobro_previsto, p.observaciones, p.fecha_creacion,
                               res.id_residencia, res.nombre as nombre_residencia,
                               3 as orden_prioridad
                        FROM pago_residente p
                        JOIN residente r ON p.id_residente = r.id_residente
                        JOIN residencia res ON p.id_residencia = res.id_residencia
                        {where_clause}
                          AND p.estado = 'cobrado'
                          AND p.fecha_pago IS NOT NULL
                          AND p.fecha_pago >= %s
                          AND NOT EXISTS (
                              -- Excluir cobros que ya están en ultimos_cobros_completados
                              SELECT 1 FROM pago_residente p2
                              JOIN residente r2 ON p2.id_residente = r2.id_residente
                              WHERE p2.id_residente = p.id_residente
                                AND p2.estado = 'cobrado'
                                AND p2.fecha_pago IS NOT NULL
                                AND (p2.fecha_pago > p.fecha_pago 
                                     OR (p2.fecha_pago = p.fecha_pago AND p2.fecha_creacion > p.fecha_creacion))
                          )
                    ),
                    cobros_periodo_cercano AS (
                        SELECT * FROM cobros_pendientes
                        UNION
                        SELECT * FROM ultimos_cobros_completados
                        UNION
                        SELECT * FROM cobros_mes_actual_anterior
                    )
                    SELECT DISTINCT ON (id_pago) * FROM cobros_periodo_cercano
                    ORDER BY id_pago, orden_prioridad,
                             CASE 
                                 WHEN fecha_prevista IS NOT NULL THEN fecha_prevista
                                 WHEN fecha_pago IS NOT NULL THEN fecha_pago
                                 ELSE '9999-12-31'::date
                             END ASC,
                             fecha_creacion DESC
                """
                # Agregar fecha del mes anterior a los params (necesitamos params para cada CTE)
                params_extended = list(params) * 3 + [mes_anterior_inicio]
                params_query = params_extended  # Guardar params para reutilizar
                cursor.execute(query, params_extended)
            else:
                # Administrador: sin filtro (acceso total)
                query = """
                    WITH cobros_pendientes AS (
                        -- Todos los cobros pendientes
                    SELECT p.id_pago, p.id_residente, r.nombre || ' ' || r.apellido as residente,
                           p.monto, p.fecha_pago, p.fecha_prevista, p.mes_pagado, p.concepto,
                           p.metodo_pago, p.estado, p.es_cobro_previsto, p.observaciones, p.fecha_creacion,
                               res.id_residencia, res.nombre as nombre_residencia,
                               1 as orden_prioridad
                    FROM pago_residente p
                    JOIN residente r ON p.id_residente = r.id_residente
                    JOIN residencia res ON p.id_residencia = res.id_residencia
                        WHERE p.estado = 'pendiente'
                    ),
                    ultimos_cobros_completados AS (
                        -- Último cobro completado de cada residente
                        SELECT DISTINCT ON (p.id_residente)
                               p.id_pago, p.id_residente, r.nombre || ' ' || r.apellido as residente,
                               p.monto, p.fecha_pago, p.fecha_prevista, p.mes_pagado, p.concepto,
                               p.metodo_pago, p.estado, p.es_cobro_previsto, p.observaciones, p.fecha_creacion,
                               res.id_residencia, res.nombre as nombre_residencia,
                               2 as orden_prioridad
                        FROM pago_residente p
                        JOIN residente r ON p.id_residente = r.id_residente
                        JOIN residencia res ON p.id_residencia = res.id_residencia
                        WHERE p.estado = 'cobrado'
                          AND p.fecha_pago IS NOT NULL
                        ORDER BY p.id_residente, p.fecha_pago DESC, p.fecha_creacion DESC
                    ),
                    cobros_mes_actual_anterior AS (
                        -- Cobros completados del mes actual y mes anterior
                        SELECT p.id_pago, p.id_residente, r.nombre || ' ' || r.apellido as residente,
                               p.monto, p.fecha_pago, p.fecha_prevista, p.mes_pagado, p.concepto,
                               p.metodo_pago, p.estado, p.es_cobro_previsto, p.observaciones, p.fecha_creacion,
                               res.id_residencia, res.nombre as nombre_residencia,
                               3 as orden_prioridad
                        FROM pago_residente p
                        JOIN residente r ON p.id_residente = r.id_residente
                        JOIN residencia res ON p.id_residencia = res.id_residencia
                        WHERE p.estado = 'cobrado'
                          AND p.fecha_pago IS NOT NULL
                          AND p.fecha_pago >= %s
                    ),
                    cobros_periodo_cercano AS (
                        SELECT * FROM cobros_pendientes
                        UNION
                        SELECT * FROM ultimos_cobros_completados
                        UNION
                        SELECT * FROM cobros_mes_actual_anterior
                    )
                    SELECT DISTINCT ON (id_pago) * FROM cobros_periodo_cercano
                    ORDER BY id_pago, orden_prioridad,
                             CASE 
                                 WHEN fecha_prevista IS NOT NULL THEN fecha_prevista
                                 WHEN fecha_pago IS NOT NULL THEN fecha_pago
                                 ELSE '9999-12-31'::date
                             END ASC,
                             fecha_creacion DESC
                """
                params_extended = [mes_anterior_inicio]
                params_query = params_extended  # Guardar params para reutilizar
                cursor.execute(query, params_extended)
            
            cobros = cursor.fetchall()
            
            # NUEVA LÓGICA: Generar automáticamente cobros pendientes para el mes siguiente
            # si un residente tiene cobros completados pero no tiene cobro pendiente
            # Calcular mes siguiente
            hoy = datetime.now()
            if hoy.month == 12:
                siguiente_mes = datetime(hoy.year + 1, 1, 1)
            else:
                siguiente_mes = datetime(hoy.year, hoy.month + 1, 1)
            mes_siguiente_str = siguiente_mes.strftime('%Y-%m')
            fecha_prevista = siguiente_mes.date()
            
            meses_espanol = {
                1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
                5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
                9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
            }
            nombre_mes = meses_espanol.get(siguiente_mes.month, 'mes')
            # Formato: "Diciembre 25", "Enero 26" (solo mes y año corto, sin "Pago")
            año_corto = str(siguiente_mes.year)[-2:]  # Últimos 2 dígitos
            concepto_siguiente = f"{nombre_mes.capitalize()} {año_corto}"
            
            # Identificar residentes con cobros completados pero sin cobro pendiente para el mes siguiente
            # Obtener residentes activos con costo_habitacion que tienen cobros completados
            if where_clause and params:
                # Usuario normal: filtrar por residencias asignadas
                # Construir filtro para residente (donde clause tiene formato "WHERE p.id_residencia IN (...)")
                # Necesitamos extraer los IDs de residencia de params
                    placeholders = ','.join(['%s'] * len(params))
                    query_residentes = f"""
                        SELECT DISTINCT r.id_residente, r.id_residencia, r.costo_habitacion, 
                               r.metodo_pago_preferido, r.nombre, r.apellido
                        FROM residente r
                        JOIN pago_residente p ON r.id_residente = p.id_residente
                        WHERE r.activo = TRUE
                          AND r.costo_habitacion IS NOT NULL
                          AND r.costo_habitacion > 0
                          AND p.estado = 'cobrado'
                          AND p.fecha_pago IS NOT NULL
                          AND r.id_residencia IN ({placeholders})
                          AND NOT EXISTS (
                              SELECT 1 FROM pago_residente p2
                              WHERE p2.id_residente = r.id_residente
                                AND p2.id_residencia = r.id_residencia
                                AND p2.mes_pagado = %s
                                AND (p2.concepto ILIKE 'enero %%' OR p2.concepto ILIKE 'febrero %%' OR p2.concepto ILIKE 'marzo %%' 
                                     OR p2.concepto ILIKE 'abril %%' OR p2.concepto ILIKE 'mayo %%' OR p2.concepto ILIKE 'junio %%'
                                     OR p2.concepto ILIKE 'julio %%' OR p2.concepto ILIKE 'agosto %%' OR p2.concepto ILIKE 'septiembre %%'
                                     OR p2.concepto ILIKE 'octubre %%' OR p2.concepto ILIKE 'noviembre %%' OR p2.concepto ILIKE 'diciembre %%'
                                     OR p2.concepto ILIKE 'Pago %%')
                          )
                    """
                    cursor.execute(query_residentes, list(params) + [mes_siguiente_str])
                    residentes_sin_cobro_pendiente = cursor.fetchall()
            elif where_clause and not params:
                # Usuario sin residencias activas asignadas
                residentes_sin_cobro_pendiente = []
            else:
                # Administrador: sin filtro
                query_residentes = """
                    SELECT DISTINCT r.id_residente, r.id_residencia, r.costo_habitacion, 
                           r.metodo_pago_preferido, r.nombre, r.apellido
                    FROM residente r
                    JOIN pago_residente p ON r.id_residente = p.id_residente
                    WHERE r.activo = TRUE
                      AND r.costo_habitacion IS NOT NULL
                      AND r.costo_habitacion > 0
                      AND p.estado = 'cobrado'
                      AND p.fecha_pago IS NOT NULL
                          AND NOT EXISTS (
                              SELECT 1 FROM pago_residente p2
                              WHERE p2.id_residente = r.id_residente
                                AND p2.id_residencia = r.id_residencia
                                AND p2.mes_pagado = %s
                                AND (p2.concepto ILIKE 'enero %%' OR p2.concepto ILIKE 'febrero %%' OR p2.concepto ILIKE 'marzo %%' 
                                     OR p2.concepto ILIKE 'abril %%' OR p2.concepto ILIKE 'mayo %%' OR p2.concepto ILIKE 'junio %%'
                                     OR p2.concepto ILIKE 'julio %%' OR p2.concepto ILIKE 'agosto %%' OR p2.concepto ILIKE 'septiembre %%'
                                     OR p2.concepto ILIKE 'octubre %%' OR p2.concepto ILIKE 'noviembre %%' OR p2.concepto ILIKE 'diciembre %%'
                                     OR p2.concepto ILIKE 'Pago %%')
                      )
                """
                cursor.execute(query_residentes, [mes_siguiente_str])
                residentes_sin_cobro_pendiente = cursor.fetchall()
            
            # Inicializar variable si no se ha hecho (para evitar errores)
            if 'residentes_sin_cobro_pendiente' not in locals():
                residentes_sin_cobro_pendiente = []
            
            # Generar cobros pendientes faltantes
            cobros_generados = 0
            for res in residentes_sin_cobro_pendiente:
                id_residente = res[0]
                id_residencia = res[1]
                costo_habitacion = float(res[2])
                metodo_pago = res[3] or 'transferencia'
                nombre = res[4]
                apellido = res[5]
                
                # Verificar duplicados antes de insertar
                existe_duplicado, id_pago_duplicado = verificar_cobro_mensual_duplicado(
                    cursor, id_residente, id_residencia, mes_siguiente_str, concepto_siguiente
                )
                
                if existe_duplicado:
                    app.logger.warning(f"Se intentó generar cobro duplicado para residente {id_residente}, mes {mes_siguiente_str}. Ya existe cobro con ID {id_pago_duplicado}")
                    continue
                
                try:
                    cursor.execute("""
                        INSERT INTO pago_residente (
                            id_residente, id_residencia, monto, fecha_pago, fecha_prevista,
                            mes_pagado, concepto, metodo_pago, estado, es_cobro_previsto
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id_pago
                    """, (
                        id_residente,
                        id_residencia,
                        costo_habitacion,
                        None,  # fecha_pago es NULL para cobros previstos
                        fecha_prevista,
                        mes_siguiente_str,
                        concepto_siguiente,
                        metodo_pago,
                        'pendiente',
                        True
                    ))
                    cobros_generados += 1
                    app.logger.info(f"Cobro pendiente generado automáticamente para {nombre} {apellido} (ID: {id_residente}): €{costo_habitacion}, mes: {mes_siguiente_str}")
                except Exception as e:
                    app.logger.error(f"Error al generar cobro pendiente automático para residente {id_residente}: {str(e)}")
            
            # Si se generaron cobros, hacer commit y volver a consultar para incluirlos
            if cobros_generados > 0:
                conn.commit()
                app.logger.info(f"Se generaron {cobros_generados} cobros pendientes automáticamente. Reconsultando lista de cobros...")
                # Volver a ejecutar la query original para incluir los nuevos cobros pendientes
                if query and params_query:
                    cursor.execute(query, params_query)
                    cobros_nuevos = cursor.fetchall()
                    # Combinar cobros originales con nuevos, eliminando duplicados por id_pago
                    cobros_dict = {cobro[0]: cobro for cobro in cobros}  # id_pago como clave
                    for cobro in cobros_nuevos:
                        if cobro[0] not in cobros_dict:  # Solo agregar si no existe
                            cobros_dict[cobro[0]] = cobro
                    cobros = list(cobros_dict.values())
            
            # Agrupar por residencia
            cobros_por_residencia = {}
            for cobro in cobros:
                idx_residencia = 13
                idx_nombre_residencia = 14
                
                id_residencia = cobro[idx_residencia]
                nombre_residencia = cobro[idx_nombre_residencia] if len(cobro) > idx_nombre_residencia else f"Residencia {id_residencia}"
                
                if id_residencia not in cobros_por_residencia:
                    cobros_por_residencia[id_residencia] = {
                        'id_residencia': id_residencia,
                        'nombre_residencia': nombre_residencia,
                        'cobros': []
                    }
                
                cobros_por_residencia[id_residencia]['cobros'].append({
                    'id_pago': cobro[0],
                    'id_residente': cobro[1],
                    'residente': cobro[2],
                    'monto': float(cobro[3]),
                    'fecha_pago': str(cobro[4]) if cobro[4] else None,
                    'fecha_prevista': str(cobro[5]) if cobro[5] else None,
                    'mes_pagado': cobro[6],
                    'concepto': cobro[7],
                    'metodo_pago': cobro[8],
                    'estado': cobro[9],
                    'es_cobro_previsto': cobro[10],
                    'observaciones': cobro[11],
                    'fecha_creacion': cobro[12].isoformat() if cobro[12] else None,
                    'nombre_residencia': nombre_residencia
                })
            
            # Convertir a lista ordenada por id_residencia
            resultado_agrupado = list(cobros_por_residencia.values())
            resultado_agrupado.sort(key=lambda x: x['id_residencia'])
            
            return jsonify({
                'cobros': [c for grupo in resultado_agrupado for c in grupo['cobros']],  # Lista plana para compatibilidad
                'cobros_agrupados': resultado_agrupado,  # Estructura agrupada
                'total': len(cobros)
            }), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        app.logger.error(f"Error al listar cobros: {str(e)}\n{error_trace}")
        # En desarrollo, incluir más detalles del error
        if app.config.get('DEBUG'):
            return jsonify({
                'error': 'Error al obtener cobros',
                'detalle': str(e),
                'traceback': error_trace.split('\n')[-5:]  # Últimas 5 líneas del traceback
            }), 500
        return jsonify({'error': 'Error al obtener cobros'}), 500


@app.route('/api/v1/facturacion/cobros', methods=['POST'])
def crear_cobro():
    """Crea un cobro previsto o registra un cobro realizado."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        # Validar datos con el módulo de validación
        is_valid, errors = validate_cobro_data(data, is_update=False)
        if not is_valid:
            return jsonify({'error': 'Errores de validación', 'detalles': errors}), 400
        
        id_residente = data.get('id_residente')
        monto = data.get('monto')
        fecha_prevista = data.get('fecha_prevista')
        fecha_pago = data.get('fecha_pago')
        
        # Determinar automáticamente si es cobro previsto: si NO tiene fecha_pago, es previsto
        es_cobro_previsto = data.get('es_cobro_previsto')
        if es_cobro_previsto is None:
            es_cobro_previsto = not fecha_pago or fecha_pago == ''
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el residente existe y obtener su residencia
            cursor.execute("""
                SELECT id_residente, id_residencia FROM residente
                WHERE id_residente = %s
            """, (id_residente,))
            
            residente_data = cursor.fetchone()
            if not residente_data:
                return jsonify({'error': 'Residente no encontrado'}), 404
            
            residente_id_residencia = residente_data[1]
            
            # Verificar que el usuario tiene acceso a la residencia del residente
            is_valid, error_response = validate_residencia_access(residente_id_residencia)
            if not is_valid:
                return error_response
            
            # Usar la residencia del residente (no la del usuario) para el cobro
            id_residencia_cobro = residente_id_residencia
            
            # Determinar estado: si tiene fecha_pago, es cobrado; si no, es pendiente
            estado_final = data.get('estado')
            if not estado_final:
                if fecha_pago:
                    estado_final = 'cobrado'
                else:
                    estado_final = 'pendiente'
            
            # Convertir "Pago mensual habitación" a formato "Pago [Mes] [Año]" basándose en fecha_prevista o mes_pagado
            concepto = data.get('concepto', '')
            mes_pagado = data.get('mes_pagado')
            fecha_prevista = data.get('fecha_prevista')
            
            if concepto and concepto.strip() == 'Pago mensual habitación':
                meses_espanol = {
                    1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
                    5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
                    9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
                }
                
                mes = None
                año = None
                
                if mes_pagado:
                    # mes_pagado está en formato 'YYYY-MM'
                    partes = mes_pagado.split('-')
                    if len(partes) == 2:
                        año = int(partes[0])
                        mes = int(partes[1])
                elif fecha_prevista:
                    # fecha_prevista puede ser string ISO o date
                    if isinstance(fecha_prevista, str):
                        try:
                            fecha = datetime.strptime(fecha_prevista, '%Y-%m-%d')
                            año = fecha.year
                            mes = fecha.month
                        except:
                            pass
                    else:
                        año = fecha_prevista.year
                        mes = fecha_prevista.month
                
                if mes and año:
                    nombre_mes = meses_espanol.get(mes, 'mes')
                    # Formato: "Diciembre 25", "Enero 26" (solo mes y año corto, sin "Pago")
                    año_corto = str(año)[-2:]  # Últimos 2 dígitos
                    concepto = f"{nombre_mes.capitalize()} {año_corto}"
            
            # Prevenir duplicados: Si el concepto es un mes (pago mensual de habitación), verificar que no exista otro cobro
            # con el mismo id_residente, mes_pagado y concepto de mes (tanto pendientes como completados)
            # Si no hay mes_pagado pero sí hay concepto de mes, intentar extraer el mes del concepto
            mes_pagado_para_validacion = mes_pagado
            if not mes_pagado_para_validacion and concepto:
                # Intentar extraer mes y año del concepto (ej: "Diciembre 25" -> mes_pagado = "2025-12")
                meses_espanol = {
                    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
                    'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
                    'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
                }
                concepto_lower = concepto.lower().strip()
                for mes_nombre, mes_num in meses_espanol.items():
                    if concepto_lower.startswith(mes_nombre):
                        # Extraer año (puede ser 2 o 4 dígitos)
                        partes = concepto_lower.split()
                        if len(partes) >= 2:
                            año_str = partes[1]
                            try:
                                año = int(año_str)
                                if len(año_str) == 2:
                                    año = 2000 + año  # Convertir "25" a "2025"
                                mes_pagado_para_validacion = f"{año}-{mes_num:02d}"
                                break
                            except ValueError:
                                pass
            
            existe_duplicado, id_pago_duplicado = verificar_cobro_mensual_duplicado(
                cursor, id_residente, id_residencia_cobro, mes_pagado_para_validacion, concepto
            )
            
            if existe_duplicado:
                return jsonify({
                    'error': f'Ya existe un cobro de habitación para este residente con el concepto "{concepto}". No se pueden crear cobros duplicados con concepto de mes (tanto pendientes como completados).'
                }), 400
            
            cursor.execute("""
                INSERT INTO pago_residente (id_residente, id_residencia, monto, fecha_pago, fecha_prevista,
                                          mes_pagado, concepto, metodo_pago, estado, es_cobro_previsto, observaciones)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id_pago
            """, (
                id_residente,
                id_residencia_cobro,
                monto,
                fecha_pago,  # Permitir fecha_pago siempre que se proporcione
                fecha_prevista,  # Permitir fecha_prevista siempre que se proporcione
                mes_pagado,  # Usar mes_pagado (puede ser None)
                concepto,  # Usar concepto actualizado (puede ser "Diciembre 25" en lugar de "Pago mensual habitación")
                data.get('metodo_pago'),
                estado_final,
                es_cobro_previsto,
                data.get('observaciones')
            ))
            
            id_pago = cursor.fetchone()[0]
            conn.commit()
            
            return jsonify({
                'id_pago': id_pago,
                'mensaje': 'Cobro previsto registrado exitosamente' if es_cobro_previsto else 'Cobro registrado exitosamente'
            }), 201
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al crear cobro: {str(e)}")
            return jsonify({'error': 'Error al crear cobro'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/facturacion/cobros/regenerar-historicos', methods=['POST'])
@permiso_requerido('crear:cobro')
def regenerar_cobros_historicos_endpoint():
    """
    Regenera cobros históricos completados para todos los residentes activos.
    Útil para corregir residentes que fueron creados recientemente pero tienen fecha_ingreso anterior.
    Solo accesible por Administrador o Administrador.
    """
    try:
        # Solo Administrador puede ejecutar esto
        if g.id_rol != ADMIN_ROLE_ID:
            return jsonify({'error': 'Solo administradores pueden regenerar cobros históricos'}), 403
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Obtener todos los residentes activos con fecha_ingreso y costo_habitacion
            cursor.execute("""
                SELECT id_residente, nombre, apellido, fecha_ingreso, costo_habitacion, 
                       metodo_pago_preferido, id_residencia
                FROM residente
                WHERE activo = TRUE
                  AND fecha_ingreso IS NOT NULL
                  AND costo_habitacion IS NOT NULL
                  AND costo_habitacion > 0
                ORDER BY id_residente
            """)
            
            residentes = cursor.fetchall()
            
            total_cobros_generados = 0
            residentes_procesados = 0
            detalles = []
            
            for residente in residentes:
                id_residente, nombre, apellido, fecha_ingreso, costo_habitacion, metodo_pago, id_residencia = residente
                
                # Verificar cuántos cobros históricos completados tiene actualmente
                cursor.execute("""
                    SELECT COUNT(*) FROM pago_residente
                    WHERE id_residente = %s
                      AND estado = 'cobrado'
                      AND es_cobro_previsto = FALSE
                """, (id_residente,))
                cobros_existentes = cursor.fetchone()[0]
                
                # Calcular cuántos cobros debería tener
                if isinstance(fecha_ingreso, str):
                    fecha_ingreso_date = datetime.strptime(fecha_ingreso, '%Y-%m-%d').date()
                else:
                    fecha_ingreso_date = fecha_ingreso
                
                hoy = datetime.now().date()
                mes_actual = datetime(hoy.year, hoy.month, 1).date()
                mes_ingreso = datetime(fecha_ingreso_date.year, fecha_ingreso_date.month, 1).date()
                
                meses_esperados = 0
                if mes_ingreso <= mes_actual:
                    fecha_temp = mes_ingreso
                    while fecha_temp <= mes_actual:
                        meses_esperados += 1
                        if fecha_temp.month == 12:
                            fecha_temp = datetime(fecha_temp.year + 1, 1, 1).date()
                        else:
                            fecha_temp = datetime(fecha_temp.year, fecha_temp.month + 1, 1).date()
                
                # Si faltan cobros, regenerar
                if cobros_existentes < meses_esperados:
                    try:
                        cobros_generados = generar_cobros_historicos_completados(
                            cursor, id_residente, id_residencia, fecha_ingreso_date, 
                            costo_habitacion, metodo_pago or 'transferencia'
                        )
                        conn.commit()
                        
                        if cobros_generados > 0:
                            total_cobros_generados += cobros_generados
                            detalles.append({
                                'residente': f"{nombre} {apellido}",
                                'id_residente': id_residente,
                                'cobros_generados': cobros_generados,
                                'cobros_existentes': cobros_existentes,
                                'meses_esperados': meses_esperados
                            })
                        
                        residentes_procesados += 1
                    except Exception as e:
                        conn.rollback()
                        app.logger.error(f"Error al generar cobros para {nombre} {apellido}: {str(e)}")
                        detalles.append({
                            'residente': f"{nombre} {apellido}",
                            'id_residente': id_residente,
                            'error': str(e)
                        })
            
            return jsonify({
                'mensaje': 'Regeneración de cobros históricos completada',
                'residentes_procesados': residentes_procesados,
                'total_cobros_generados': total_cobros_generados,
                'detalles': detalles
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al regenerar cobros históricos: {str(e)}")
            return jsonify({'error': f'Error al regenerar cobros históricos: {str(e)}'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/facturacion/cobros/generar-previstos', methods=['POST'])
@permiso_requerido('crear:cobro')
def generar_cobros_previstos():
    """
    Genera automáticamente cobros previstos para el MES SIGUIENTE para todos los residentes activos
    que tengan costo_habitacion definido.
    
    NUEVA LÓGICA:
    - Solo genera el cobro del mes siguiente (no histórico)
    - Se cobra por adelantado el mes
    - Fecha prevista: día 1 del mes siguiente
    - Si el residente ya tiene un cobro (completado o previsto) para el mes siguiente, no se genera
    
    Para generar cobros históricos (desde fecha_ingreso), usar el endpoint generar-historicos.
    """
    try:
        data = request.get_json() or {}
        mes_referencia = data.get('mes')  # Formato: 'YYYY-MM', opcional
        año_referencia = data.get('año')  # Opcional
        
        # Si no se especifica mes, usar el mes ACTUAL (no el siguiente)
        if mes_referencia:
            try:
                fecha_base = datetime.strptime(f"{mes_referencia}-01", "%Y-%m-%d")
            except:
                return jsonify({'error': 'Formato de mes inválido. Use YYYY-MM'}), 400
        else:
            # Mes ACTUAL por defecto (cambiar lógica: generar para mes actual si no hay cobro completado)
            hoy = datetime.now()
            fecha_base = datetime(hoy.year, hoy.month, 1)
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Calcular mes siguiente
            hoy = datetime.now()
            if hoy.month == 12:
                siguiente_mes = datetime(hoy.year + 1, 1, 1)
            else:
                siguiente_mes = datetime(hoy.year, hoy.month + 1, 1)
            mes_siguiente = siguiente_mes.strftime('%Y-%m')
            
            # Limpiar TODOS los cobros previstos pendientes antes de regenerar
            # Si es Administrador, limpia de TODAS las residencias. Si no, solo de las asignadas.
            if g.id_rol == ADMIN_ROLE_ID:
                cursor.execute("""
                    DELETE FROM pago_residente
                    WHERE es_cobro_previsto = TRUE
                      AND estado = 'pendiente'
                """)
            else:
                # Usuario normal: limpiar solo de sus residencias asignadas
                if not g.residencias_acceso:
                    return jsonify({'error': 'Usuario sin residencias asignadas'}), 403
                
                placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                cursor.execute(f"""
                    DELETE FROM pago_residente
                    WHERE id_residencia IN ({placeholders})
                      AND es_cobro_previsto = TRUE
                      AND estado = 'pendiente'
                """, tuple(g.residencias_acceso))
            
            cobros_eliminados = cursor.rowcount
            
            # Calcular mes siguiente (una sola vez)
            hoy = datetime.now()
            mes_actual = hoy.month
            año_actual = hoy.year
            
            if mes_actual == 12:
                siguiente_mes = datetime(año_actual + 1, 1, 1)
            else:
                siguiente_mes = datetime(año_actual, mes_actual + 1, 1)
            
            mes_siguiente_str = siguiente_mes.strftime('%Y-%m')
            fecha_prevista = siguiente_mes.date()  # Día 1 del mes siguiente
            
            # Obtener residentes activos con costo_habitacion
            # Si es Administrador, obtiene de TODAS las residencias. Si no, solo de las asignadas.
            if g.id_rol == ADMIN_ROLE_ID:
                cursor.execute("""
                    SELECT id_residente, nombre, apellido, costo_habitacion, metodo_pago_preferido, fecha_ingreso, id_residencia
                    FROM residente
                    WHERE activo = TRUE 
                      AND costo_habitacion IS NOT NULL 
                      AND costo_habitacion > 0
                      AND fecha_ingreso IS NOT NULL
                """)
                cursor.execute("SELECT COUNT(*) FROM residente WHERE activo = TRUE")
                total_residentes_activos = cursor.fetchone()[0]
                app.logger.info(f"Generando cobros previstos GLOBAL (Administrador) para mes siguiente: {mes_siguiente_str}")
            else:
                # Usuario normal: filtrar por residencias asignadas
                if not g.residencias_acceso:
                    return jsonify({'error': 'Usuario sin residencias asignadas'}), 403
                
                placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                cursor.execute(f"""
                    SELECT id_residente, nombre, apellido, costo_habitacion, metodo_pago_preferido, fecha_ingreso, id_residencia
                FROM residente
                    WHERE id_residencia IN ({placeholders})
                  AND activo = TRUE 
                  AND costo_habitacion IS NOT NULL 
                  AND costo_habitacion > 0
                      AND fecha_ingreso IS NOT NULL
                """, tuple(g.residencias_acceso))
                
                cursor.execute(f"SELECT COUNT(*) FROM residente WHERE id_residencia IN ({placeholders}) AND activo = TRUE", tuple(g.residencias_acceso))
                total_residentes_activos = cursor.fetchone()[0]
                app.logger.info(f"Generando cobros previstos para residencias: {g.residencias_acceso}, mes siguiente: {mes_siguiente_str}")
            
            residentes = cursor.fetchall()

            app.logger.info(f"Total residentes activos en alcance: {total_residentes_activos}")
            app.logger.info(f"Residentes candidatos encontrados (con costo y fecha_ingreso): {len(residentes)}")
            app.logger.info(f"Mes siguiente a generar: {mes_siguiente_str}")
            
            if not residentes:
                return jsonify({
                    'mensaje': 'No hay residentes activos con costo_habitacion definido que deban tener cobros previstos',
                    'cobros_generados': 0,
                    'cobros_eliminados': cobros_eliminados,
                    'mes_referencia': mes_siguiente_str,
                    'total_residentes_activos': total_residentes_activos,
                    'residentes_candidatos': 0
                }), 200
            
            cobros_generados = 0
            cobros_duplicados = 0
            cobros_ya_existentes = 0
            errores = []
            residentes_procesados = []
            
            meses_espanol = {
                1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
                5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
                9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
            }
            
            # NUEVA LÓGICA: Solo generar cobro para el mes siguiente
            for residente in residentes:
                id_residente = residente[0]
                nombre = residente[1]
                apellido = residente[2]
                costo_habitacion = float(residente[3])
                metodo_pago = residente[4] or 'transferencia'  # Por defecto transferencia
                fecha_ingreso = residente[5]  # fecha_ingreso del residente
                residencia_del_residente = residente[6] # ID de la residencia del residente actual
                
                if not fecha_ingreso:
                    app.logger.warning(f"Residente {nombre} {apellido} (ID: {id_residente}) no tiene fecha_ingreso, saltando")
                    continue
                
                # Generar concepto con el nombre del mes (formato nuevo: "Mes Año")
                nombre_mes = meses_espanol.get(siguiente_mes.month, 'mes')
                año_corto = str(siguiente_mes.year)[-2:]  # Últimos 2 dígitos
                concepto = f"{nombre_mes.capitalize()} {año_corto}"
                
                # Verificar si ya existe un cobro mensual duplicado (pendiente o completado)
                existe_duplicado, id_pago_duplicado = verificar_cobro_mensual_duplicado(
                    cursor, id_residente, residencia_del_residente, mes_siguiente_str, concepto
                )
                
                if existe_duplicado:
                    # Ya existe un cobro para el mes siguiente
                    cursor.execute("""
                        SELECT estado FROM pago_residente WHERE id_pago = %s
                    """, (id_pago_duplicado,))
                    estado_result = cursor.fetchone()
                    estado_duplicado = estado_result[0] if estado_result else None
                    
                    if estado_duplicado == 'cobrado':
                        cobros_ya_existentes += 1
                        app.logger.debug(f"Residente {nombre} {apellido} (ID: {id_residente}) ya tiene cobro completado para {mes_siguiente_str}")
                    else:
                        cobros_duplicados += 1
                        app.logger.debug(f"Residente {nombre} {apellido} (ID: {id_residente}) ya tiene cobro previsto para {mes_siguiente_str}")
                    continue
                
                # Crear el cobro previsto para el mes siguiente
                try:
                    cursor.execute("""
                        INSERT INTO pago_residente (
                            id_residente, id_residencia, monto, fecha_pago, fecha_prevista,
                            mes_pagado, concepto, metodo_pago, estado, es_cobro_previsto
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id_pago
                    """, (
                        id_residente,
                        residencia_del_residente,
                        costo_habitacion,
                        None,  # fecha_pago es NULL para cobros previstos
                        fecha_prevista,  # Día 1 del mes siguiente
                        mes_siguiente_str,
                        concepto,
                        metodo_pago,
                        'pendiente',
                        True
                    ))
                    
                    cobros_generados += 1
                    app.logger.info(f"Cobro previsto generado para {nombre} {apellido} (ID: {id_residente}): €{costo_habitacion}, mes: {mes_siguiente_str}")
                    residentes_procesados.append(f"{nombre} {apellido} (ID: {id_residente})")
                    
                except Exception as e:
                    errores.append(f"Error al crear cobro para {nombre} {apellido} mes {mes_siguiente_str}: {str(e)}")
                    app.logger.error(f"Error al crear cobro previsto para residente {id_residente} mes {mes_siguiente_str}: {str(e)}")
            
            conn.commit()
            
            resultado = {
                'mensaje': f'Cobros previstos generados exitosamente para el mes siguiente ({mes_siguiente_str})',
                'cobros_generados': cobros_generados,
                'cobros_eliminados': cobros_eliminados,
                'cobros_duplicados': cobros_duplicados,
                'cobros_ya_existentes': cobros_ya_existentes,
                'mes_actual': mes_actual.strftime('%Y-%m'),
                'total_residentes_procesados': len(residentes),
                'total_residentes_candidatos': len(residentes),
                'residentes_procesados': residentes_procesados
            }
            
            app.logger.info(f"Resumen: {cobros_generados} cobros generados, {cobros_duplicados} duplicados (previstos), {cobros_ya_existentes} ya existentes (cobrados), {len(residentes)} residentes procesados")
            
            if errores:
                resultado['errores'] = errores
            
            return jsonify(resultado), 201
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al generar cobros previstos: {str(e)}")
            return jsonify({'error': f'Error al generar cobros previstos: {str(e)}'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/facturacion/cobros/estadisticas', methods=['GET'])
def estadisticas_cobros():
    """Obtiene estadísticas mensuales de cobros (histórico y estimaciones)."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Si es Admin (rol 1), obtiene de TODAS las residencias agrupado. Si no, solo de la suya.
            if g.id_rol == ADMIN_ROLE_ID:
                # Obtener cobros históricos (cobrados) agrupados por mes y residencia
                cursor.execute("""
                    SELECT 
                        CASE 
                            WHEN p.metodo_pago ILIKE 'remesa' 
                                 AND EXTRACT(DAY FROM p.fecha_pago) = 30 
                                 AND p.mes_pagado IS NOT NULL
                            THEN p.mes_pagado
                            ELSE TO_CHAR(p.fecha_pago, 'YYYY-MM')
                        END as mes,
                        res.nombre as nombre_residencia,
                        res.id_residencia,
                        SUM(p.monto) as total_cobrado,
                        COUNT(*) as cantidad
                    FROM pago_residente p
                    JOIN residencia res ON p.id_residencia = res.id_residencia
                    WHERE p.estado = 'cobrado'
                      AND p.fecha_pago IS NOT NULL
                    GROUP BY 
                        CASE 
                            WHEN p.metodo_pago ILIKE 'remesa' 
                                 AND EXTRACT(DAY FROM p.fecha_pago) = 30 
                                 AND p.mes_pagado IS NOT NULL
                            THEN p.mes_pagado
                            ELSE TO_CHAR(p.fecha_pago, 'YYYY-MM')
                        END,
                        res.id_residencia, res.nombre
                    ORDER BY mes DESC
                    LIMIT 48
                """)
                historico = cursor.fetchall()
                
                # Obtener estimaciones futuras agrupados por mes y residencia
                cursor.execute("""
                    SELECT 
                        CASE 
                            WHEN p.metodo_pago ILIKE 'remesa' AND p.mes_pagado IS NOT NULL
                            THEN p.mes_pagado
                            ELSE TO_CHAR(p.fecha_prevista, 'YYYY-MM')
                        END as mes,
                        res.nombre as nombre_residencia,
                        res.id_residencia,
                        SUM(p.monto) as total_previsto,
                        COUNT(*) as cantidad
                    FROM pago_residente p
                    JOIN residencia res ON p.id_residencia = res.id_residencia
                    WHERE p.es_cobro_previsto = TRUE
                      AND p.estado = 'pendiente'
                      AND p.fecha_prevista IS NOT NULL
                    GROUP BY 
                        CASE 
                            WHEN p.metodo_pago ILIKE 'remesa' AND p.mes_pagado IS NOT NULL
                            THEN p.mes_pagado
                            ELSE TO_CHAR(p.fecha_prevista, 'YYYY-MM')
                        END,
                        res.id_residencia, res.nombre
                    ORDER BY mes ASC
                    LIMIT 24
                """)
                estimaciones = cursor.fetchall()

                historico_data = []
                for row in historico:
                    historico_data.append({
                        'mes': row[0],
                        'nombre_residencia': row[1],
                        'id_residencia': row[2],
                        'total': float(row[3]),
                        'cantidad': row[4]
                    })

                estimaciones_data = []
                for row in estimaciones:
                    estimaciones_data.append({
                        'mes': row[0],
                        'nombre_residencia': row[1],
                        'id_residencia': row[2],
                        'total': float(row[3]),
                        'cantidad': row[4]
                    })

            else:
                # Usuario normal: filtrar por residencias asignadas
                if not g.residencias_acceso:
                    return jsonify({'historico': [], 'estimaciones': []}), 200
                
                placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                
                # Obtener cobros históricos (cobrados) agrupados por mes
                cursor.execute(f"""
                    SELECT 
                        CASE 
                            WHEN metodo_pago ILIKE 'remesa' 
                                 AND EXTRACT(DAY FROM fecha_pago) = 30 
                                 AND mes_pagado IS NOT NULL
                            THEN mes_pagado
                            ELSE TO_CHAR(fecha_pago, 'YYYY-MM')
                        END as mes,
                        SUM(monto) as total_cobrado,
                        COUNT(*) as cantidad
                    FROM pago_residente
                    WHERE id_residencia IN ({placeholders})
                      AND estado = 'cobrado'
                      AND fecha_pago IS NOT NULL
                    GROUP BY 
                        CASE 
                            WHEN metodo_pago ILIKE 'remesa' 
                                 AND EXTRACT(DAY FROM fecha_pago) = 30 
                                 AND mes_pagado IS NOT NULL
                            THEN mes_pagado
                            ELSE TO_CHAR(fecha_pago, 'YYYY-MM')
                        END
                    ORDER BY mes DESC
                    LIMIT 12
                """, tuple(g.residencias_acceso))
                
                historico = cursor.fetchall()
                
                # Obtener estimaciones futuras
                cursor.execute(f"""
                    SELECT 
                        CASE 
                            WHEN metodo_pago ILIKE 'remesa' AND mes_pagado IS NOT NULL
                            THEN mes_pagado
                            ELSE TO_CHAR(fecha_prevista, 'YYYY-MM')
                        END as mes,
                        SUM(monto) as total_previsto,
                        COUNT(*) as cantidad
                    FROM pago_residente
                    WHERE id_residencia IN ({placeholders})
                      AND es_cobro_previsto = TRUE
                      AND estado = 'pendiente'
                      AND fecha_prevista IS NOT NULL
                    GROUP BY 
                        CASE 
                            WHEN metodo_pago ILIKE 'remesa' AND mes_pagado IS NOT NULL
                            THEN mes_pagado
                            ELSE TO_CHAR(fecha_prevista, 'YYYY-MM')
                        END
                    ORDER BY mes ASC
                    LIMIT 6
                """, tuple(g.residencias_acceso))
                
                estimaciones = cursor.fetchall()
                
                historico_data = []
                for row in historico:
                    historico_data.append({
                        'mes': row[0],
                        'total': float(row[1]),
                        'cantidad': row[2]
                    })
                
                estimaciones_data = []
                for row in estimaciones:
                    estimaciones_data.append({
                        'mes': row[0],
                        'total': float(row[1]),
                        'cantidad': row[2]
                    })
            
            return jsonify({
                'historico': historico_data,
                'estimaciones': estimaciones_data
            }), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al obtener estadísticas: {str(e)}")
        return jsonify({'error': 'Error al obtener estadísticas'}), 500


@app.route('/api/v1/facturacion/cobros/ultimos-completados', methods=['GET'])
@permiso_requerido('leer:cobro')
def ultimos_cobros_completados():
    """
    Obtiene el último pago mensual y el último pago extra completado de cada residente.
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar qué columnas existen en la tabla pago_residente
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'pago_residente'
            """)
            columnas_existentes = {row[0] for row in cursor.fetchall()}
            
            # Construir SELECT dinámicamente según columnas existentes
            columnas_base = [
                'p.id_pago', 'p.id_residente', 
                "r.nombre || ' ' || r.apellido as residente",
                'p.monto', 'p.fecha_pago', 'p.fecha_prevista', 'p.mes_pagado', 'p.concepto',
                'p.metodo_pago', 'p.estado', 'p.es_cobro_previsto', 'p.fecha_creacion'
            ]
            
            if 'observaciones' in columnas_existentes:
                columnas_base.insert(-1, 'p.observaciones')
            else:
                columnas_base.insert(-1, 'NULL as observaciones')
            
            select_clause = ', '.join(columnas_base)
            
            # Obtener último pago mensual de cada residente
            # Consideramos "mensual" los que tienen concepto que empieza con "Pago" seguido de un mes
            if g.id_rol == ADMIN_ROLE_ID:
                # Administrador: sin filtro
                query_mensual = f"""
                SELECT DISTINCT ON (p.id_residente)
                    {select_clause}
                FROM pago_residente p
                JOIN residente r ON p.id_residente = r.id_residente
                    WHERE p.estado = 'cobrado'
                  AND p.fecha_pago IS NOT NULL
                      AND (p.concepto ILIKE 'enero %%' OR p.concepto ILIKE 'febrero %%' OR p.concepto ILIKE 'marzo %%' 
                           OR p.concepto ILIKE 'abril %%' OR p.concepto ILIKE 'mayo %%' OR p.concepto ILIKE 'junio %%'
                           OR p.concepto ILIKE 'julio %%' OR p.concepto ILIKE 'agosto %%' OR p.concepto ILIKE 'septiembre %%'
                           OR p.concepto ILIKE 'octubre %%' OR p.concepto ILIKE 'noviembre %%' OR p.concepto ILIKE 'diciembre %%'
                           OR p.concepto ILIKE 'Pago %%' OR p.concepto ILIKE 'Pago mensual%%')
                ORDER BY p.id_residente, p.fecha_pago DESC, p.fecha_creacion DESC
            """
                cursor.execute(query_mensual)
                ultimos_mensuales = cursor.fetchall()
                
                query_extra = f"""
                    SELECT DISTINCT ON (p.id_residente)
                        {select_clause}
                    FROM pago_residente p
                    JOIN residente r ON p.id_residente = r.id_residente
                    WHERE p.estado = 'cobrado'
                      AND p.fecha_pago IS NOT NULL
                      AND (p.concepto IS NULL OR (p.concepto NOT ILIKE 'enero %%' AND p.concepto NOT ILIKE 'febrero %%' 
                           AND p.concepto NOT ILIKE 'marzo %%' AND p.concepto NOT ILIKE 'abril %%' 
                           AND p.concepto NOT ILIKE 'mayo %%' AND p.concepto NOT ILIKE 'junio %%'
                           AND p.concepto NOT ILIKE 'julio %%' AND p.concepto NOT ILIKE 'agosto %%' 
                           AND p.concepto NOT ILIKE 'septiembre %%' AND p.concepto NOT ILIKE 'octubre %%' 
                           AND p.concepto NOT ILIKE 'noviembre %%' AND p.concepto NOT ILIKE 'diciembre %%'
                           AND p.concepto NOT ILIKE 'Pago %%' AND p.concepto NOT ILIKE 'Pago mensual%%'))
                    ORDER BY p.id_residente, p.fecha_pago DESC, p.fecha_creacion DESC
                """
                cursor.execute(query_extra)
            else:
                # Usuario normal: filtrar por lista de residencias
                if not g.residencias_acceso:
                    return jsonify({'cobros': [], 'total': 0}), 200
                
                placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                
                query_mensual = f"""
                    SELECT DISTINCT ON (p.id_residente)
                        {select_clause}
                    FROM pago_residente p
                    JOIN residente r ON p.id_residente = r.id_residente
                    WHERE p.id_residencia IN ({placeholders})
                      AND p.estado = 'cobrado'
                      AND p.fecha_pago IS NOT NULL
                      AND (p.concepto ILIKE 'enero %%' OR p.concepto ILIKE 'febrero %%' OR p.concepto ILIKE 'marzo %%' 
                           OR p.concepto ILIKE 'abril %%' OR p.concepto ILIKE 'mayo %%' OR p.concepto ILIKE 'junio %%'
                           OR p.concepto ILIKE 'julio %%' OR p.concepto ILIKE 'agosto %%' OR p.concepto ILIKE 'septiembre %%'
                           OR p.concepto ILIKE 'octubre %%' OR p.concepto ILIKE 'noviembre %%' OR p.concepto ILIKE 'diciembre %%'
                           OR p.concepto ILIKE 'Pago %%' OR p.concepto ILIKE 'Pago mensual%%')
                    ORDER BY p.id_residente, p.fecha_pago DESC, p.fecha_creacion DESC
                """
                cursor.execute(query_mensual, tuple(g.residencias_acceso))
                ultimos_mensuales = cursor.fetchall()
                
                query_extra = f"""
                    SELECT DISTINCT ON (p.id_residente)
                        {select_clause}
                    FROM pago_residente p
                    JOIN residente r ON p.id_residente = r.id_residente
                    WHERE p.id_residencia IN ({placeholders})
                      AND p.estado = 'cobrado'
                      AND p.fecha_pago IS NOT NULL
                      AND (p.concepto IS NULL OR (p.concepto NOT ILIKE 'enero %%' AND p.concepto NOT ILIKE 'febrero %%' 
                           AND p.concepto NOT ILIKE 'marzo %%' AND p.concepto NOT ILIKE 'abril %%' 
                           AND p.concepto NOT ILIKE 'mayo %%' AND p.concepto NOT ILIKE 'junio %%'
                           AND p.concepto NOT ILIKE 'julio %%' AND p.concepto NOT ILIKE 'agosto %%' 
                           AND p.concepto NOT ILIKE 'septiembre %%' AND p.concepto NOT ILIKE 'octubre %%' 
                           AND p.concepto NOT ILIKE 'noviembre %%' AND p.concepto NOT ILIKE 'diciembre %%'
                           AND p.concepto NOT ILIKE 'Pago %%' AND p.concepto NOT ILIKE 'Pago mensual%%'))
                    ORDER BY p.id_residente, p.fecha_pago DESC, p.fecha_creacion DESC
                """
                cursor.execute(query_extra, tuple(g.residencias_acceso))
            
            ultimos_extras = cursor.fetchall()
            
            # Formatear resultados
            resultado = []
            
            # Determinar índices basados en si observaciones existe
            # Orden: id_pago(0), id_residente(1), residente(2), monto(3), fecha_pago(4), 
            # fecha_prevista(5), mes_pagado(6), concepto(7), metodo_pago(8), estado(9), 
            # es_cobro_previsto(10), observaciones(11 si existe), fecha_creacion(último)
            tiene_observaciones = 'observaciones' in columnas_existentes
            idx_fecha_creacion = 12 if tiene_observaciones else 11
            
            for cobro in ultimos_mensuales:
                try:
                    if len(cobro) < 11:
                        app.logger.warning(f"Cobro mensual con menos columnas de las esperadas: {len(cobro)}")
                        continue
                    
                    resultado.append({
                        'id_pago': cobro[0],
                        'id_residente': cobro[1],
                        'residente': cobro[2],
                        'monto': float(cobro[3]) if cobro[3] is not None else 0.0,
                        'fecha_pago': str(cobro[4]) if cobro[4] else None,
                        'fecha_prevista': str(cobro[5]) if cobro[5] else None,
                        'mes_pagado': cobro[6],
                        'concepto': cobro[7],
                        'metodo_pago': cobro[8],
                        'estado': cobro[9],
                        'es_cobro_previsto': cobro[10] if len(cobro) > 10 else False,
                        'observaciones': cobro[11] if tiene_observaciones and len(cobro) > 11 else None,
                        'fecha_creacion': cobro[idx_fecha_creacion].isoformat() if len(cobro) > idx_fecha_creacion and cobro[idx_fecha_creacion] else None,
                        'tipo': 'mensual'
                    })
                except (IndexError, TypeError, AttributeError) as e:
                    app.logger.error(f"Error al formatear cobro mensual: {str(e)}, tupla len: {len(cobro) if cobro else 0}")
                    continue
            
            for cobro in ultimos_extras:
                try:
                    resultado.append({
                        'id_pago': cobro[0],
                        'id_residente': cobro[1],
                        'residente': cobro[2],
                        'monto': float(cobro[3]) if cobro[3] is not None else 0.0,
                        'fecha_pago': str(cobro[4]) if cobro[4] else None,
                        'fecha_prevista': str(cobro[5]) if cobro[5] else None,
                        'mes_pagado': cobro[6],
                        'concepto': cobro[7],
                        'metodo_pago': cobro[8],
                        'estado': cobro[9],
                        'es_cobro_previsto': cobro[10] if len(cobro) > 10 else False,
                        'observaciones': cobro[11] if tiene_observaciones and len(cobro) > 11 else None,
                        'fecha_creacion': cobro[idx_fecha_creacion].isoformat() if len(cobro) > idx_fecha_creacion and cobro[idx_fecha_creacion] else None,
                        'tipo': 'extra'
                    })
                except (IndexError, TypeError, AttributeError) as e:
                    app.logger.error(f"Error al formatear cobro extra: {str(e)}, tupla: {cobro}, len: {len(cobro) if cobro else 0}")
                    continue
            
            return jsonify({'cobros': resultado, 'total': len(resultado)}), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        import traceback
        app.logger.error(f"Error al obtener últimos cobros completados: {str(e)}")
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error al obtener últimos cobros completados: {str(e)}'}), 500


@app.route('/api/v1/facturacion/cobros/<int:id_pago>', methods=['GET'])
@permiso_requerido('leer:cobro')
def obtener_cobro(id_pago):
    """Obtiene un cobro específico por su ID."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Primero verificar que el cobro existe y pertenece a la residencia
            cursor.execute("""
                SELECT id_pago, id_residente, id_residencia
                FROM pago_residente
                WHERE id_pago = %s
            """, (id_pago,))
            
            cobro_basico = cursor.fetchone()
            
            if not cobro_basico:
                app.logger.warning(f"Cobro {id_pago} no encontrado")
                return jsonify({'error': 'Cobro no encontrado'}), 404
            
            # Verificar que el usuario tiene acceso a la residencia del cobro
            is_valid, error_response = validate_residencia_access(cobro_basico[2])
            if not is_valid:
                return error_response
            
            # Ahora obtener la información completa con JOIN
            cursor.execute("""
                SELECT p.id_pago, p.id_residente, p.id_residencia,
                       COALESCE(r.nombre || ' ' || r.apellido, 'Residente no encontrado') as residente,
                       p.monto, p.fecha_pago, p.fecha_prevista, p.mes_pagado, p.concepto,
                       p.metodo_pago, p.estado, p.es_cobro_previsto, p.observaciones, p.fecha_creacion
                FROM pago_residente p
                LEFT JOIN residente r ON p.id_residente = r.id_residente
                WHERE p.id_pago = %s
            """, (id_pago,))
            
            cobro = cursor.fetchone()
            
            if not cobro:
                return jsonify({'error': 'Error al obtener información del cobro'}), 500
            
            resultado = {
                'id_pago': cobro[0],
                'id_residente': cobro[1],
                'id_residencia': cobro[2],
                'residente': cobro[3],  # COALESCE(r.nombre || ' ' || r.apellido, ...)
                'monto': float(cobro[4]),  # p.monto
                'fecha_pago': str(cobro[5]) if cobro[5] else None,  # p.fecha_pago
                'fecha_prevista': str(cobro[6]) if cobro[6] else None,  # p.fecha_prevista
                'mes_pagado': cobro[7],  # p.mes_pagado
                'concepto': cobro[8],  # p.concepto
                'metodo_pago': cobro[9],  # p.metodo_pago
                'estado': cobro[10],  # p.estado
                'es_cobro_previsto': cobro[11],  # p.es_cobro_previsto
                'observaciones': cobro[12],  # p.observaciones
                'fecha_creacion': cobro[13].isoformat() if cobro[13] else None  # p.fecha_creacion
            }
            
            return jsonify(resultado), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al obtener cobro {id_pago}: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': 'Error al obtener cobro', 'details': str(e)}), 500


@app.route('/api/v1/facturacion/cobros/normalizar-conceptos', methods=['POST'])
@permiso_requerido('editar:cobro')
def normalizar_conceptos_cobros():
    """
    Normaliza los conceptos de cobros existentes:
    - Convierte "Pago mensual habitación" o "Pago [Mes] [Año]" a formato "[Mes] [Año corto]"
    - Ejemplos: "Pago Diciembre 2025" → "Diciembre 25", "Pago mensual habitación" → "Diciembre 25"
    - Basándose en mes_pagado o fecha_prevista
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            meses_espanol = {
                1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
                5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
                9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
            }
            
            # Obtener cobros con conceptos que necesitan normalización
            # Buscar conceptos que sean "Pago mensual habitación" o "Pago [mes] [año completo]" 
            # y convertirlos a formato "[Mes] [año corto]" (ej: "Diciembre 25")
            cursor.execute("""
                SELECT id_pago, concepto, mes_pagado, fecha_prevista, fecha_pago
                FROM pago_residente
                WHERE concepto ILIKE 'Pago mensual%%' 
                   OR concepto ILIKE 'mensual habitación%%'
                   OR concepto = 'Pago mensual habitación'
                   OR (concepto ILIKE 'Pago %%' AND concepto NOT SIMILAR TO 'Pago (enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre) [0-9]{2}')
                   OR (concepto ILIKE 'Pago %%' AND concepto SIMILAR TO 'Pago (enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre) [0-9]{4}')
            """)
            
            cobros_a_normalizar = cursor.fetchall()
            actualizados = 0
            errores = []
            
            for cobro in cobros_a_normalizar:
                id_pago = cobro[0]
                concepto_actual = cobro[1]
                mes_pagado = cobro[2]
                fecha_prevista = cobro[3]
                fecha_pago = cobro[4]
                
                # Determinar mes y año
                mes = None
                año = None
                
                if mes_pagado:
                    # mes_pagado está en formato 'YYYY-MM'
                    partes = mes_pagado.split('-')
                    if len(partes) == 2:
                        año = int(partes[0])
                        mes = int(partes[1])
                elif fecha_prevista:
                    # fecha_prevista es date
                    if isinstance(fecha_prevista, str):
                        try:
                            fecha = datetime.strptime(fecha_prevista, '%Y-%m-%d')
                            año = fecha.year
                            mes = fecha.month
                        except:
                            pass
                    else:
                        año = fecha_prevista.year
                        mes = fecha_prevista.month
                elif fecha_pago:
                    # fecha_pago es date
                    if isinstance(fecha_pago, str):
                        try:
                            fecha = datetime.strptime(fecha_pago, '%Y-%m-%d')
                            año = fecha.year
                            mes = fecha.month
                        except:
                            pass
                    else:
                        año = fecha_pago.year
                        mes = fecha_pago.month
                
                if mes and año:
                    nombre_mes = meses_espanol.get(mes, 'mes')
                    # Formato: "Diciembre 25", "Enero 26" (solo mes y año corto, sin "Pago")
                    año_corto = str(año)[-2:]  # Últimos 2 dígitos
                    concepto_nuevo = f"{nombre_mes.capitalize()} {año_corto}"
                    
                    try:
                        cursor.execute("""
                            UPDATE pago_residente
                            SET concepto = %s
                            WHERE id_pago = %s
                        """, (concepto_nuevo, id_pago))
                        actualizados += 1
                    except Exception as e:
                        errores.append(f"Error al actualizar cobro {id_pago}: {str(e)}")
                else:
                    errores.append(f"Cobro {id_pago}: No se pudo determinar mes/año")
            
            conn.commit()
            
            return jsonify({
                'mensaje': f'Conceptos normalizados exitosamente',
                'actualizados': actualizados,
                'total_revisados': len(cobros_a_normalizar),
                'errores': errores if errores else None
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al normalizar conceptos: {str(e)}")
            import traceback
            app.logger.error(traceback.format_exc())
            return jsonify({'error': 'Error al normalizar conceptos', 'details': str(e)}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/facturacion/cobros/<int:id_pago>', methods=['PUT'])
@permiso_requerido('editar:cobro')
def actualizar_cobro(id_pago):
    """Actualiza un cobro (cambiar estado, marcar como cobrado, etc.)."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el cobro existe
            # Verificar que el cobro existe y obtener su residencia
            cursor.execute("""
                SELECT id_pago, id_residencia FROM pago_residente
                    WHERE id_pago = %s
                """, (id_pago,))
            
            cobro_existente = cursor.fetchone()
            if not cobro_existente:
                return jsonify({'error': 'Cobro no encontrado'}), 404
            
            # Verificar acceso a la residencia del cobro
            is_valid, error_response = validate_residencia_access(cobro_existente[1])
            if not is_valid:
                return error_response
            
            # Validar datos si se proporcionan
            if 'estado' in data:
                valid, error = validate_estado(data.get('estado'))
                if not valid:
                    return jsonify({'error': error}), 400
            
            if 'monto' in data and data.get('monto') is not None:
                valid, error = validate_monto(data.get('monto'), 'Monto', required=False)
                if not valid:
                    return jsonify({'error': error}), 400
            
            if 'metodo_pago' in data and data.get('metodo_pago'):
                valid, error = validate_metodo_pago(data.get('metodo_pago'), required=False)
                if not valid:
                    return jsonify({'error': error}), 400
            
            # Campos actualizables
            campos_actualizables = [
                'estado', 'fecha_pago', 'fecha_prevista', 'monto', 'concepto',
                'metodo_pago', 'mes_pagado', 'es_cobro_previsto', 'observaciones'
            ]
            
            updates = []
            valores = []
            
            for campo in campos_actualizables:
                if campo in data:
                    updates.append(f"{campo} = %s")
                    valores.append(data[campo])
            
            if not updates:
                return jsonify({'error': 'No hay campos para actualizar'}), 400
            
            # Actualizar el cobro (ya validamos acceso arriba)
                valores.append(id_pago)
                query = f"""
                    UPDATE pago_residente
                    SET {', '.join(updates)}
                    WHERE id_pago = %s
                    RETURNING id_pago
                """
            
            cursor.execute(query, valores)
            conn.commit()
            
            return jsonify({'mensaje': 'Cobro actualizado exitosamente'}), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al actualizar cobro: {str(e)}")
            return jsonify({'error': 'Error al actualizar cobro'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/facturacion/cobros/<int:id_pago>', methods=['DELETE'])
@permiso_requerido('eliminar:cobro')
def eliminar_cobro(id_pago):
    """Elimina un cobro completamente. Solo permite eliminar cobros pendientes/previstos."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el cobro existe y obtener su estado
            cursor.execute("""
                SELECT id_pago, id_residencia, estado, es_cobro_previsto FROM pago_residente
                    WHERE id_pago = %s
                """, (id_pago,))
            
            cobro = cursor.fetchone()
            
            if not cobro:
                return jsonify({'error': 'Cobro no encontrado'}), 404
            
            id_pago_db, id_residencia, estado, es_cobro_previsto = cobro
            
            # Verificar acceso a la residencia del cobro
            is_valid, error_response = validate_residencia_access(id_residencia)
            if not is_valid:
                return error_response
            
            # Solo permitir eliminar cobros pendientes o previstos
            # No permitir eliminar cobros completados (cobrados)
            if estado == 'cobrado':
                return jsonify({
                    'error': 'No se pueden eliminar cobros completados. Solo se pueden eliminar cobros pendientes o previstos.'
                }), 400
            
            # Eliminar el cobro
                cursor.execute("""
                    DELETE FROM pago_residente
                    WHERE id_pago = %s
                    RETURNING id_pago
                """, (id_pago,))
            
            eliminado = cursor.fetchone()
            
            if not eliminado:
                return jsonify({'error': 'No se pudo eliminar el cobro'}), 500
            
            conn.commit()
            
            return jsonify({
                'mensaje': 'Cobro eliminado exitosamente',
                'id_pago': id_pago
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al eliminar cobro: {str(e)}")
            import traceback
            app.logger.error(traceback.format_exc())
            return jsonify({'error': 'Error al eliminar el cobro', 'details': str(e)}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/facturacion/proveedores', methods=['GET'])
@permiso_requerido('leer:pago_proveedor')
def listar_pagos_proveedores():
    """Lista los pagos a proveedores de la residencia."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Construir query con filtros por residencias de acceso
            query = """
                SELECT p.id_pago, p.proveedor, p.concepto, p.monto, p.fecha_pago, 
                       p.fecha_prevista, p.metodo_pago, p.estado, p.numero_factura,
                       p.observaciones, p.fecha_creacion, res.id_residencia, res.nombre as nombre_residencia
                FROM pago_proveedor p
                JOIN residencia res ON p.id_residencia = res.id_residencia
                WHERE 1=1
            """
            params = []
            
            # Filtrar por residencias de acceso (excepto Administrador)
            if g.id_rol != ADMIN_ROLE_ID:
                if g.residencias_acceso:
                    placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                    query += f" AND p.id_residencia IN ({placeholders})"
                    params.extend(g.residencias_acceso)
                else:
                    # Usuario sin residencias
                    return jsonify({'pagos': [], 'total': 0}), 200
            
            query += " ORDER BY COALESCE(p.fecha_pago, p.fecha_prevista, p.fecha_creacion::date) DESC, p.fecha_creacion DESC"
            
            cursor.execute(query, params)
            pagos = cursor.fetchall()
            
            resultado = []
            for pago in pagos:
                resultado.append({
                    'id_pago': pago[0],
                    'proveedor': pago[1],
                    'concepto': pago[2],
                    'monto': float(pago[3]) if pago[3] else 0,
                    'fecha_pago': str(pago[4]) if pago[4] else None,
                    'fecha_prevista': str(pago[5]) if pago[5] else None,
                    'metodo_pago': pago[6],
                    'estado': pago[7],
                    'numero_factura': pago[8],
                    'observaciones': pago[9],
                    'fecha_creacion': pago[10].isoformat() if pago[10] else None,
                    'id_residencia': pago[11],
                    'nombre_residencia': pago[12]
                })
            
            return jsonify({'pagos': resultado, 'total': len(resultado)}), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al listar pagos a proveedores: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error al obtener pagos: {str(e)}'}), 500


@app.route('/api/v1/facturacion/proveedores', methods=['POST'])
@permiso_requerido('crear:pago_proveedor')
def crear_pago_proveedor():
    """Crea un pago a proveedor o una estimación basada en historial."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        proveedor = data.get('proveedor')
        concepto = data.get('concepto')
        monto = data.get('monto')
        fecha_pago = data.get('fecha_pago')
        fecha_prevista = data.get('fecha_prevista')
        
        # Validaciones básicas
        valid, error = validate_text(proveedor, 'Proveedor', min_length=2, max_length=255, required=True)
        if not valid:
            return jsonify({'error': error}), 400
        
        valid, error = validate_text(concepto, 'Concepto', min_length=3, max_length=500, required=True)
        if not valid:
            return jsonify({'error': error}), 400
        
        if monto is not None:
            valid, error = validate_monto(monto, 'Monto', required=False)
            if not valid:
                return jsonify({'error': error}), 400
        
        # Calcular estado automáticamente basado en las fechas
        # Si tiene fecha_pago → pagado, si tiene fecha_prevista → pendiente
        if fecha_pago:
            estado = 'pagado'
        elif fecha_prevista:
            estado = 'pendiente'
        else:
            estado = data.get('estado', 'pendiente')  # Por defecto pendiente si no hay fechas
        
        if not monto or monto <= 0:
            return jsonify({'error': 'monto es requerido'}), 400
        
        # Validar que se proporcione id_residencia
        id_residencia = data.get('id_residencia')
        if not id_residencia:
            return jsonify({'error': 'id_residencia es requerido'}), 400
        
        # Verificar acceso a la residencia
        is_valid, error_response = validate_residencia_access(id_residencia)
        if not is_valid:
            return error_response
            
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar si existen las columnas opcionales
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'pago_proveedor' 
                  AND column_name IN ('factura_blob_path', 'id_receiver')
            """)
            columnas_existentes = {row[0] for row in cursor.fetchall()}
            tiene_columna_factura = 'factura_blob_path' in columnas_existentes
            tiene_columna_receiver = 'id_receiver' in columnas_existentes
            
            # Obtener id_receiver si se proporciona
            id_receiver = data.get('id_receiver')
            
            # Si se proporciona un nombre de receiver pero no id_receiver, buscar o crear
            receiver_nombre = data.get('receiver')
            if receiver_nombre and not id_receiver:
                # Buscar receiver existente por nombre
                cursor.execute("""
                    SELECT id_receiver FROM receiver 
                    WHERE UPPER(TRIM(nombre)) = UPPER(TRIM(%s)) AND activo = TRUE
                    LIMIT 1
                """, (receiver_nombre,))
                receiver_existente = cursor.fetchone()
                
                if receiver_existente:
                    id_receiver = receiver_existente[0]
                else:
                    # Crear nuevo receiver
                    cursor.execute("""
                        INSERT INTO receiver (nombre, activo)
                        VALUES (%s, TRUE)
                        RETURNING id_receiver
                    """, (receiver_nombre,))
                    id_receiver = cursor.fetchone()[0]
                    
                    # Asociar receiver a la residencia si no existe la relación
                    cursor.execute("""
                        INSERT INTO residencia_receiver (id_residencia, id_receiver, activo)
                        VALUES (%s, %s, TRUE)
                        ON CONFLICT (id_residencia, id_receiver) DO UPDATE SET activo = TRUE
                    """, (id_residencia, id_receiver))
            
            # Construir query dinámicamente según columnas existentes
            columnas_insert = ['id_residencia', 'proveedor', 'concepto', 'monto', 'fecha_pago',
                             'fecha_prevista', 'metodo_pago', 'estado', 'numero_factura', 'observaciones']
            valores_insert = [id_residencia, proveedor, concepto, monto, fecha_pago,
                            fecha_prevista, data.get('metodo_pago'), estado,
                            data.get('numero_factura'), data.get('observaciones')]
            
            if tiene_columna_factura:
                columnas_insert.append('factura_blob_path')
                valores_insert.append(data.get('factura_blob_path'))
            
            if tiene_columna_receiver and id_receiver:
                columnas_insert.append('id_receiver')
                valores_insert.append(id_receiver)
            
            placeholders = ','.join(['%s'] * len(valores_insert))
            columnas_str = ','.join(columnas_insert)
            
            cursor.execute(f"""
                INSERT INTO pago_proveedor ({columnas_str})
                VALUES ({placeholders})
                RETURNING id_pago
            """, tuple(valores_insert))
            
            id_pago = cursor.fetchone()[0]
            
            # Si hay factura_blob_path, también guardar en documentación
            factura_blob_path = data.get('factura_blob_path')
            if factura_blob_path:
                try:
                    # Verificar si existe la tabla documento
                    cursor.execute("""
                        SELECT EXISTS (
                            SELECT FROM information_schema.tables 
                            WHERE table_schema = 'public' 
                            AND table_name = 'documento'
                        )
                    """)
                    tabla_documento_existe = cursor.fetchone()[0]
                    
                    if tabla_documento_existe:
                        # Intentar obtener id_proveedor del nombre del proveedor
                        id_proveedor = None
                        cursor.execute("""
                            SELECT id_proveedor FROM proveedor 
                            WHERE nombre = %s AND id_residencia = %s AND activo = TRUE
                            LIMIT 1
                        """, (proveedor, id_residencia))
                        
                        proveedor_result = cursor.fetchone()
                        if proveedor_result:
                            id_proveedor = proveedor_result[0]
                        else:
                            # Si el proveedor no existe, crear uno nuevo automáticamente
                            app.logger.info(f"Proveedor '{proveedor}' no encontrado, creando automáticamente")
                            cursor.execute("""
                                INSERT INTO proveedor (id_residencia, nombre, activo)
                                VALUES (%s, %s, TRUE)
                                RETURNING id_proveedor
                            """, (id_residencia, proveedor))
                            id_proveedor = cursor.fetchone()[0]
                            app.logger.info(f"Proveedor '{proveedor}' creado con ID {id_proveedor}")
                        
                        # Obtener nombre del archivo del blob_path
                        numero_factura = data.get('numero_factura')
                        nombre_archivo = factura_blob_path.split('/')[-1] if '/' in factura_blob_path else f"Factura_{numero_factura or id_pago}.pdf"
                        
                        # Crear documento en la tabla documento
                        cursor.execute("""
                            INSERT INTO documento (tipo_entidad, id_entidad, id_residencia, categoria_documento,
                                                  tipo_documento, nombre_archivo, descripcion, url_archivo,
                                                  tamaño_bytes, tipo_mime, id_usuario_subida, activo)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE)
                            RETURNING id_documento
            """, (
                            'proveedor',
                            id_proveedor,
                id_residencia,
                            'Pagos',  # Nueva categoría para pagos a proveedores
                            'Pago a proveedor',  # Tipo de documento específico
                            nombre_archivo,
                            f"Factura de pago: {proveedor} - {concepto}",
                            factura_blob_path,
                            None,  # tamaño_bytes - no lo tenemos aquí
                            'application/pdf',
                            g.id_usuario
                        ))
                        
                        id_documento = cursor.fetchone()[0]
                        app.logger.info(f"Documento de factura creado (ID: {id_documento}) para proveedor {id_proveedor}, pago {id_pago}")
                except Exception as doc_error:
                    # Si falla crear el documento, no fallar el pago
                    app.logger.warning(f"Error al crear documento de factura: {str(doc_error)}")
                    import traceback
                    app.logger.warning(traceback.format_exc())
            
            conn.commit()
            
            return jsonify({
                'id_pago': id_pago,
                'mensaje': 'Pago a proveedor registrado exitosamente'
            }), 201
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al crear pago a proveedor: {str(e)}")
            return jsonify({'error': 'Error al crear pago a proveedor'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/facturacion/proveedores/<int:id_pago>', methods=['GET'])
@permiso_requerido('leer:pago_proveedor')
def obtener_pago_proveedor(id_pago):
    """Obtiene un pago a proveedor por su ID."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar si existe la columna factura_blob_path
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'pago_proveedor' 
                  AND column_name = 'factura_blob_path'
            """)
            tiene_columna_factura = cursor.fetchone() is not None
            
            if tiene_columna_factura:
                cursor.execute("""
                    SELECT id_pago, id_residencia, proveedor, concepto, monto, fecha_pago, fecha_prevista,
                           metodo_pago, estado, numero_factura, observaciones, fecha_creacion, factura_blob_path
                    FROM pago_proveedor
                    WHERE id_pago = %s
                """, (id_pago,))
            else:
                cursor.execute("""
                    SELECT id_pago, id_residencia, proveedor, concepto, monto, fecha_pago, fecha_prevista,
                       metodo_pago, estado, numero_factura, observaciones, fecha_creacion
                FROM pago_proveedor
                    WHERE id_pago = %s
            """, (id_pago,))
            
            pago = cursor.fetchone()
            if not pago:
                return jsonify({'error': 'Pago no encontrado'}), 404
            
            # Verificar acceso a la residencia del pago
            is_valid, error_response = validate_residencia_access(pago[1])
            if not is_valid:
                return error_response
            
            # Construir respuesta
            respuesta = {
                'id_pago': pago[0],
                'id_residencia': pago[1],
                'proveedor': pago[2],
                'concepto': pago[3],
                'monto': float(pago[4]),
                'fecha_pago': str(pago[5]) if pago[5] else None,
                'fecha_prevista': str(pago[6]) if pago[6] else None,
                'metodo_pago': pago[7],
                'estado': pago[8],
                'numero_factura': pago[9],
                'observaciones': pago[10],
                'fecha_creacion': pago[11].isoformat() if pago[11] else None,
                'factura_blob_path': None,
                'factura_url': None
            }
            
            # Añadir factura_blob_path y URL si existe
            app.logger.info(f"DEBUG obtener_pago_proveedor: id_pago={id_pago}, tiene_columna_factura={tiene_columna_factura}, len(pago)={len(pago) if pago else 0}")
            if tiene_columna_factura and len(pago) > 12:
                factura_blob_path = pago[12]
                app.logger.info(f"DEBUG: factura_blob_path obtenido de BD: '{factura_blob_path}', tipo: {type(factura_blob_path)}, es None: {factura_blob_path is None}, es vacío: {factura_blob_path == '' if factura_blob_path else 'N/A'}")
                if factura_blob_path and factura_blob_path != '' and str(factura_blob_path).strip() != '':
                    respuesta['factura_blob_path'] = factura_blob_path
                    # Generar URL firmada para la factura
                    try:
                        pdf_url = get_document_url(factura_blob_path, expiration_minutes=60)
                        respuesta['factura_url'] = pdf_url
                        app.logger.info(f"URL de factura generada exitosamente: {pdf_url}")
                    except Exception as e:
                        app.logger.warning(f"Error al generar URL firmada para factura: {str(e)}")
                        respuesta['factura_url'] = None
                else:
                    app.logger.warning(f"Factura blob_path es None o vacío para pago {id_pago}. Valor: '{factura_blob_path}'")
                    respuesta['factura_blob_path'] = None
                    respuesta['factura_url'] = None
            else:
                app.logger.warning(f"No se puede obtener factura_blob_path: tiene_columna_factura={tiene_columna_factura}, len(pago)={len(pago) if pago else 0}")
            
            return jsonify(respuesta), 200
            
        except Exception as e:
            app.logger.error(f"Error al obtener pago a proveedor: {str(e)}")
            import traceback
            app.logger.error(traceback.format_exc())
            return jsonify({'error': 'Error al obtener pago a proveedor'}), 500
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()
    except Exception as e:
        app.logger.error(f"Error al obtener conexión a la base de datos: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': 'Error al obtener pago a proveedor'}), 500


@app.route('/api/v1/facturacion/proveedores/<int:id_pago>', methods=['DELETE'])
@permiso_requerido('eliminar:pago_proveedor')
def eliminar_pago_proveedor(id_pago):
    """Elimina completamente un pago a proveedor de la base de datos."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar si existe la columna factura_blob_path
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'pago_proveedor' 
                  AND column_name = 'factura_blob_path'
            """)
            tiene_columna_factura = cursor.fetchone() is not None
            
            # Verificar que el pago existe y obtener datos para validar acceso
            if tiene_columna_factura:
                cursor.execute("""
                    SELECT id_residencia, factura_blob_path
                    FROM pago_proveedor
                    WHERE id_pago = %s
                """, (id_pago,))
            else:
                cursor.execute("""
                    SELECT id_residencia
                    FROM pago_proveedor
                    WHERE id_pago = %s
                """, (id_pago,))
            
            pago = cursor.fetchone()
            if not pago:
                return jsonify({'error': 'Pago no encontrado'}), 404
            
            id_residencia = pago[0]
            factura_blob_path = pago[1] if tiene_columna_factura and len(pago) > 1 else None
            
            # Verificar acceso a la residencia del pago
            is_valid, error_response = validate_residencia_access(id_residencia)
            if not is_valid:
                return error_response
            
            # Eliminar factura de Cloud Storage si existe
            if factura_blob_path:
                try:
                    from storage_manager import delete_document
                    delete_document(factura_blob_path)
                    app.logger.info(f"Factura eliminada de Cloud Storage: {factura_blob_path}")
                except Exception as e:
                    app.logger.warning(f"Error al eliminar factura de Cloud Storage: {str(e)}")
                    # Continuar con la eliminación del registro aunque falle la eliminación del archivo
            
            # Eliminar el registro de la base de datos
            cursor.execute("""
                DELETE FROM pago_proveedor
                WHERE id_pago = %s
            """, (id_pago,))
            
            conn.commit()
            
            app.logger.info(f"Pago a proveedor {id_pago} eliminado exitosamente")
            
            return jsonify({
                'mensaje': 'Pago a proveedor eliminado exitosamente',
                'id_pago': id_pago
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al eliminar pago a proveedor: {str(e)}")
            import traceback
            app.logger.error(traceback.format_exc())
            return jsonify({'error': 'Error al eliminar pago a proveedor'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/facturacion/procesar-factura', methods=['POST'])
@permiso_requerido('crear:pago_proveedor')
def procesar_factura():
    """
    Procesa una factura PDF usando Google Document AI para extraer datos.
    Guarda el PDF en Cloud Storage y devuelve los datos extraídos.
    """
    try:
        if 'factura' not in request.files:
            return jsonify({'error': 'No se proporcionó archivo de factura'}), 400
        
        archivo = request.files['factura']
        if archivo.filename == '':
            return jsonify({'error': 'Nombre de archivo vacío'}), 400
        
        # Validar que sea PDF o imagen
        nombre_lower = archivo.filename.lower()
        extensiones_permitidas = ['.pdf', '.jpg', '.jpeg', '.png']
        tipos_mime_permitidos = ['application/pdf', 'image/jpeg', 'image/jpg', 'image/png']
        
        extension_valida = any(nombre_lower.endswith(ext) for ext in extensiones_permitidas)
        tipo_mime_valido = archivo.content_type in tipos_mime_permitidos
        
        if not extension_valida and not tipo_mime_valido:
            return jsonify({'error': 'El archivo debe ser un PDF o una imagen (JPG, PNG)'}), 400
        
        # Determinar tipo de archivo
        es_imagen = nombre_lower.endswith(('.jpg', '.jpeg', '.png')) or archivo.content_type.startswith('image/')
        es_pdf = nombre_lower.endswith('.pdf') or archivo.content_type == 'application/pdf'
        
        # Obtener id_residencia si se proporciona
        id_residencia = request.form.get('id_residencia', type=int)
        if not id_residencia:
            # Si no se proporciona, usar la primera residencia del usuario
            if g.id_rol == ADMIN_ROLE_ID:
                # Para Administrador, usar 1 por defecto
                id_residencia = 1
            else:
                if not g.residencias_acceso:
                    return jsonify({'error': 'No se pudo determinar la residencia'}), 400
                id_residencia = g.residencias_acceso[0]
        
        # Verificar acceso a la residencia
        is_valid, error_response = validate_residencia_access(id_residencia)
        if not is_valid:
            return error_response
        
        # Leer contenido del archivo
        file_content = archivo.read()
        nombre_archivo = archivo.filename
        
        # Inicializar variables para detección automática de residencia
        id_residencia_detectada = None
        residencias_usuario = None
        
        # Procesar con Google Document AI
        datos_extraidos = {}
        try:
            from google.cloud import documentai
            
            # Obtener configuración de Document AI
            # Valores por defecto basados en la configuración proporcionada
            project_id = os.getenv('GOOGLE_CLOUD_PROJECT_ID', '621063984498')
            location = os.getenv('DOCUMENT_AI_LOCATION', 'eu')  # Región europea por defecto
            # Usar Invoice Parser especializado por defecto
            processor_id = os.getenv('DOCUMENT_AI_PROCESSOR_ID', 'cdb6e7f4178248c4')
            
            if project_id and processor_id:
                # Inicializar cliente de Document AI con el endpoint correcto según la región
                from google.api_core.client_options import ClientOptions
                
                # Configurar endpoint según la región
                if location == 'eu':
                    api_endpoint = 'eu-documentai.googleapis.com:443'
                elif location == 'asia':
                    api_endpoint = 'asia-documentai.googleapis.com:443'
                else:  # us o por defecto
                    api_endpoint = 'us-documentai.googleapis.com:443'
                
                client_options = ClientOptions(api_endpoint=api_endpoint)
                client = documentai.DocumentProcessorServiceClient(client_options=client_options)
                name = f"projects/{project_id}/locations/{location}/processors/{processor_id}"
                
                # Determinar MIME type según el tipo de archivo
                if es_imagen:
                    # Determinar tipo MIME específico de la imagen
                    if nombre_lower.endswith('.png') or archivo.content_type == 'image/png':
                        mime_type = 'image/png'
                    elif nombre_lower.endswith(('.jpg', '.jpeg')) or archivo.content_type in ['image/jpeg', 'image/jpg']:
                        mime_type = 'image/jpeg'
                    else:
                        mime_type = 'image/jpeg'  # Por defecto JPEG
                else:
                    mime_type = 'application/pdf'
                
                # Configurar la solicitud
                raw_document = documentai.RawDocument(
                    content=file_content,
                    mime_type=mime_type
                )
                
                request_doc = documentai.ProcessRequest(
                    name=name,
                    raw_document=raw_document
                )
                
                # Procesar el documento
                result = client.process_document(request=request_doc)
                document = result.document
                
                # Extraer texto completo
                texto_completo = document.text
                app.logger.info(f"Document AI procesado. Texto extraído: {len(texto_completo)} caracteres")
                app.logger.debug(f"Primeros 500 caracteres del texto: {texto_completo[:500]}")
                
                import re
                
                # ============================================================================
                # NUEVA ESTRATEGIA: BÚSQUEDA INTELIGENTE EN TEXTO OCR
                # Ignoramos los campos estructurados de Invoice Parser (son ambiguos)
                # Buscamos directamente en el texto con conocimiento del dominio español
                # ============================================================================
                
                # ============================================================================
                # FUNCIONES DE BÚSQUEDA INTELIGENTE CON CONOCIMIENTO DEL DOMINIO
                # ============================================================================
                
                def buscar_en_texto_con_patrones(texto, patrones, nombre_campo="campo"):
                    """
                    Busca un valor numérico en el texto usando múltiples patrones.
                    Devuelve el primer match válido encontrado.
                    """
                    for patron in patrones:
                        matches = re.finditer(patron, texto, re.IGNORECASE | re.MULTILINE)
                        for match in matches:
                            try:
                                valor_str = match.group(1)
                                valor = parsear_monto_espanol(valor_str)
                                if valor and valor > 0:
                                    app.logger.info(f"✅ {nombre_campo} encontrado: {valor} (patrón: '{patron[:50]}...', texto: '{valor_str}')")
                                    return valor
                            except:
                                continue
                    return None
                
                # FUNCIÓN MEJORADA DE PARSING DE NÚMEROS
                def parsear_monto_espanol(valor_str):
                    """
                    Parsea un string de monto manejando formatos españoles Y casos de la IA.
                    
                    PROBLEMA: Document AI devuelve mention_text en formato mixto:
                    - '241,80' (español) → a veces llega como '241.8' (punto decimal)
                    - '2.418,50' (español con miles)
                    
                    Ejemplos:
                    - "241,80" → 241.80
                    - "241.8" → 241.8 (de la IA, punto decimal)
                    - "241.80" → 241.80 (de la IA, punto decimal)
                    - "2.418,50" → 2418.50 (español, miles + decimales)
                    - "15234" → 15234.0
                    """
                    if not valor_str or not isinstance(valor_str, str):
                        return None
                    
                    # Limpiar espacios, símbolos de moneda
                    cleaned = valor_str.replace('€', '').replace('EUR', '').replace(' ', '').strip()
                    
                    # Sin separadores → número simple
                    if ',' not in cleaned and '.' not in cleaned:
                        try:
                            return float(cleaned)
                        except:
                            return None
                    
                    # Tiene coma → formato español clásico
                    if ',' in cleaned:
                        # Quitar puntos (miles) y cambiar coma por punto (decimal)
                        cleaned = cleaned.replace('.', '').replace(',', '.')
                        try:
                            return float(cleaned)
                        except:
                            return None
                    
                    # Solo tiene punto, sin coma → ambiguo
                    # Puede ser:
                    # - Decimal de la IA: '241.8' → 241.8
                    # - Miles español: '2.418' → 2418.0
                    if '.' in cleaned:
                        partes = cleaned.split('.')
                        if len(partes) == 2:
                            parte_entera, parte_decimal = partes
                            # Si la parte decimal tiene ≤2 dígitos → probablemente es decimal
                            if len(parte_decimal) <= 2:
                                # Es decimal: '241.8' → 241.8
                                try:
                                    return float(cleaned)
                                except:
                                    return None
                            else:
                                # Es miles: '2.418' → 2418
                                cleaned = cleaned.replace('.', '')
                                try:
                                    return float(cleaned)
                                except:
                                    return None
                    
                    # Fallback
                    try:
                        return float(cleaned)
                    except:
                        return None
                
                # ============================================================================
                # EXTRACCIÓN DIRECTA DEL TEXTO OCR (NUEVA ESTRATEGIA)
                # ============================================================================
                
                app.logger.info("🎯 Iniciando extracción DIRECTA del texto OCR con patrones específicos...")
                
                # 1. TOTAL CON IVA (el más importante)
                app.logger.info("📊 Buscando TOTAL con IVA...")
                patrones_total = [
                    r'total\s+facturat?[:\s]+([\d.,]+)',
                    r'importe\s+total[:\s]+([\d.,]+)',
                    r'total[:\s]+([\d.,]+)\s*€',
                    r'total[:\s]+([\d.,]+)\s*\(',  # "241,80 (Euros"
                    r'importe[:\s]+([\d.,]+)\s*€',
                ]
                total_con_iva = buscar_en_texto_con_patrones(texto_completo, patrones_total, "TOTAL CON IVA")
                
                # 2. BASE IMPONIBLE
                app.logger.info("📊 Buscando BASE IMPONIBLE...")
                patrones_base = [
                    r'base\s+imp(?:onible)?[:\.\s]+([\d.,]+)',
                    r'neto\s+fact[:\.\s]+([\d.,]+)',
                    r'base[:\s]+([\d.,]+)\s*€',
                    r'total\s+neto[:\s]+([\d.,]+)',
                ]
                base_imponible = buscar_en_texto_con_patrones(texto_completo, patrones_base, "BASE IMPONIBLE")
                
                # 3. IMPORTE DE IVA (no porcentaje)
                app.logger.info("📊 Buscando IMPORTE DE IVA...")
                patrones_iva_importe = [
                    r'imp\.?\s*iva[:\s]+([\d.,]+)',
                    r'importe\s+iva[:\s]+([\d.,]+)',
                    r'cuota\s+iva[:\s]+([\d.,]+)',
                    r'iva[:\s]+([\d.,]+)\s*€',
                ]
                iva_importe = buscar_en_texto_con_patrones(texto_completo, patrones_iva_importe, "IMPORTE IVA")
                
                # 4. PORCENTAJE DE IVA
                app.logger.info("📊 Buscando PORCENTAJE DE IVA...")
                patrones_iva_porcentaje = [
                    r'%\s*iva[:\s]+([\d,]+)',
                    r'iva[:\s]+([\d,]+)\s*%',
                    r'tipo\s+iva[:\s]+([\d,]+)',
                ]
                iva_porcentaje_encontrado = None
                for patron in patrones_iva_porcentaje:
                    match = re.search(patron, texto_completo, re.IGNORECASE)
                    if match:
                        try:
                            valor_str = match.group(1).replace(',', '.')
                            valor = float(valor_str)
                            # Validar que sea un porcentaje válido en España
                            if valor in [4, 10, 21]:
                                iva_porcentaje_encontrado = int(valor)
                                app.logger.info(f"✅ PORCENTAJE IVA encontrado: {iva_porcentaje_encontrado}% (texto: '{valor_str}')")
                                break
                            elif 0 < valor < 1:  # Formato decimal: 0.21 → 21%
                                valor_convertido = int(valor * 100)
                                if valor_convertido in [4, 10, 21]:
                                    iva_porcentaje_encontrado = valor_convertido
                                    app.logger.info(f"✅ PORCENTAJE IVA encontrado y convertido: {iva_porcentaje_encontrado}% (texto: '{valor_str}')")
                                    break
                        except:
                            continue
                
                # Si no encontramos el porcentaje, calcularlo desde base e IVA
                if not iva_porcentaje_encontrado and base_imponible and iva_importe:
                    porcentaje_calculado = (iva_importe / base_imponible) * 100
                    # Redondear al porcentaje español más cercano
                    porcentajes_validos = [4, 10, 21]
                    iva_porcentaje_encontrado = min(porcentajes_validos, key=lambda x: abs(x - porcentaje_calculado))
                    app.logger.info(f"✅ PORCENTAJE IVA calculado: {iva_porcentaje_encontrado}% (desde importe/base: {porcentaje_calculado:.2f}%)")
                
                # 5. NÚMERO DE FACTURA
                app.logger.info("📊 Buscando NÚMERO DE FACTURA...")
                patrones_factura = [
                    r'factura\s+n[°º\.]?\s*[:\s]*([\w\-/]+)',
                    r'n[°º]\s*factura[:\s]*([\w\-/]+)',
                    r'fact(?:ura)?[:\s]*([\d\-/A-Z]+)',
                ]
                numero_factura = None
                for patron in patrones_factura:
                    match = re.search(patron, texto_completo, re.IGNORECASE)
                    if match:
                        numero_factura = match.group(1).strip()
                        app.logger.info(f"✅ NÚMERO FACTURA encontrado: {numero_factura}")
                        break
                
                # 6. FECHA
                app.logger.info("📊 Buscando FECHA...")
                patrones_fecha = [
                    r'fecha[:\s]+(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
                    r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
                ]
                fecha_factura = None
                for patron in patrones_fecha:
                    match = re.search(patron, texto_completo, re.IGNORECASE)
                    if match:
                        fecha_str = match.group(1)
                        try:
                            # Convertir a formato ISO
                            from datetime import datetime
                            for fmt in ['%d-%m-%Y', '%d/%m/%Y', '%d-%m-%y', '%d/%m/%y']:
                                try:
                                    fecha_obj = datetime.strptime(fecha_str, fmt)
                                    fecha_factura = fecha_obj.strftime('%Y-%m-%d')
                                    app.logger.info(f"✅ FECHA encontrada: {fecha_factura} (texto: '{fecha_str}')")
                                    break
                                except:
                                    continue
                            if fecha_factura:
                                break
                        except:
                            continue
                
                # 7. PROVEEDOR (nombre)
                app.logger.info("📊 Buscando PROVEEDOR...")
                # Buscar en las primeras líneas (usualmente está al principio)
                primeras_lineas = '\n'.join(texto_completo.split('\n')[:10])
                proveedor = None
                # Buscar líneas con texto (sin números ni símbolos)
                for linea in primeras_lineas.split('\n'):
                    linea_limpia = linea.strip()
                    if linea_limpia and len(linea_limpia) > 3:
                        # Si tiene principalmente letras y espacios
                        if sum(c.isalpha() or c.isspace() for c in linea_limpia) / len(linea_limpia) > 0.7:
                            proveedor = linea_limpia
                            app.logger.info(f"✅ PROVEEDOR encontrado: {proveedor}")
                            break
                
                # VALIDACIÓN DE COHERENCIA
                app.logger.info("🔍 Validando coherencia de datos extraídos...")
                if total_con_iva and base_imponible and iva_importe:
                    suma_calculada = round(base_imponible + iva_importe, 2)
                    diferencia = abs(suma_calculada - total_con_iva)
                    if diferencia <= 0.50:
                        app.logger.info(f"✅ COHERENCIA OK: Base ({base_imponible}) + IVA ({iva_importe}) = {suma_calculada} ≈ Total ({total_con_iva}). Dif: {diferencia}€")
                    else:
                        app.logger.warning(f"⚠️ INCOHERENCIA: Base ({base_imponible}) + IVA ({iva_importe}) = {suma_calculada} ≠ Total ({total_con_iva}). Dif: {diferencia}€")
                
                # Si no encontramos IVA pero tenemos total y base, calcularlo
                if not iva_importe and total_con_iva and base_imponible:
                    iva_importe = round(total_con_iva - base_imponible, 2)
                    app.logger.info(f"✅ IVA calculado (Total - Base): {iva_importe}€")
                
                # Construir datos extraídos con la nueva estrategia
                datos_extraidos = {}
                if numero_factura:
                    datos_extraidos['numero_factura'] = numero_factura
                if fecha_factura:
                    datos_extraidos['fecha_pago'] = fecha_factura
                if proveedor:
                    datos_extraidos['proveedor'] = proveedor
                if total_con_iva:
                    datos_extraidos['total_con_impuestos'] = round(total_con_iva, 2)
                    datos_extraidos['monto'] = round(total_con_iva, 2)
                if base_imponible:
                    datos_extraidos['total'] = round(base_imponible, 2)
                if iva_importe:
                    datos_extraidos['impuestos'] = round(iva_importe, 2)
                    datos_extraidos['iva'] = round(iva_importe, 2)
                if iva_porcentaje_encontrado:
                    datos_extraidos['iva_porcentaje'] = iva_porcentaje_encontrado
                
                app.logger.info(f"📊 Datos extraídos con nueva estrategia: {list(datos_extraidos.keys())}")
                
                # ============================================================================
                # FALLBACK: Si falla la búsqueda directa, usar Invoice Parser
                # ============================================================================
                if not datos_extraidos or len(datos_extraidos) < 3:
                    app.logger.warning("⚠️ Búsqueda directa no encontró suficientes datos, usando Invoice Parser como fallback...")
                
                # EXTRAER DATOS ESTRUCTURADOS DEL INVOICE PARSER (SOLO COMO FALLBACK)
                # El Invoice Parser proporciona entidades estructuradas con campos específicos
                if (not datos_extraidos or len(datos_extraidos) < 3) and hasattr(document, 'entities') and document.entities:
                    app.logger.info(f"Invoice Parser detectado. Extrayendo {len(document.entities)} entidades estructuradas...")
                    
                    # Log DETALLADO de todas las entidades disponibles para debugging
                    tipos_entidades = set()
                    entidades_detalladas = {}
                    entidades_con_montos = []
                    
                    for entity in document.entities:
                        tipos_entidades.add(entity.type_)
                        if entity.type_ not in entidades_detalladas:
                            entidades_detalladas[entity.type_] = []
                        
                        # Guardar información completa de la entidad para debugging
                        entidad_info = {
                            'type': entity.type_,
                            'confidence': getattr(entity, 'confidence', None),
                            'mention_text': getattr(entity, 'mention_text', None)
                        }
                        
                        # Si tiene normalized_value con money_value, guardar el valor
                        if hasattr(entity, 'normalized_value') and entity.normalized_value:
                            norm_val = entity.normalized_value
                            if hasattr(norm_val, 'money_value') and norm_val.money_value:
                                money = norm_val.money_value
                                unidades = float(money.units) if hasattr(money, 'units') else 0
                                nanos = float(money.nanos) / 1e9 if hasattr(money, 'nanos') else 0
                                valor_money = unidades + nanos
                                entidad_info['money_value'] = valor_money
                                entidad_info['money_currency'] = getattr(money, 'currency_code', None)
                                
                                # Guardar entidades con montos para debugging
                                entidades_con_montos.append({
                                    'type': entity.type_,
                                    'value': valor_money,
                                    'currency': getattr(money, 'currency_code', None),
                                    'mention_text': getattr(entity, 'mention_text', None)
                                })
                        
                        entidades_detalladas[entity.type_].append(entidad_info)
                    
                    app.logger.info(f"📋 Tipos de entidades encontradas por Invoice Parser ({len(tipos_entidades)} tipos): {sorted(tipos_entidades)}")
                    app.logger.info(f"💰 Entidades con montos encontradas: {entidades_con_montos}")
                    app.logger.debug(f"📋 Detalles completos de entidades: {entidades_detalladas}")
                    
                    # CARGAR MAPEO DE CAMPOS DESDE LA BASE DE DATOS
                    mapeo_campos = {}
                    try:
                        conn_mapeo = get_db_connection()
                        cursor_mapeo = conn_mapeo.cursor()
                        try:
                            # Verificar si existe la tabla de mapeo
                            cursor_mapeo.execute("""
                                SELECT EXISTS (
                                    SELECT FROM information_schema.tables 
                                    WHERE table_schema = 'public'
                                    AND table_name = 'mapeo_campos_ia'
                                )
                            """)
                            tabla_existe = cursor_mapeo.fetchone()[0]
                            
                            if tabla_existe:
                                cursor_mapeo.execute("""
                                    SELECT campo_sistema, campo_ia, tipos_alternativos
                                    FROM mapeo_campos_ia
                                    WHERE activo = TRUE
                                """)
                                for row in cursor_mapeo.fetchall():
                                    campo_sistema = row[0]
                                    campo_ia = row[1]
                                    tipos_alt_str = row[2] if row[2] else '[]'
                                    try:
                                        tipos_alternativos = json.loads(tipos_alt_str) if tipos_alt_str else []
                                    except:
                                        tipos_alternativos = []
                                    mapeo_campos[campo_sistema] = {
                                        'campo_ia': campo_ia,
                                        'tipos_alternativos': tipos_alternativos
                                    }
                                app.logger.info(f"✅ Mapeo de campos cargado: {len(mapeo_campos)} campos configurados")
                        finally:
                            cursor_mapeo.close()
                            conn_mapeo.close()
                    except Exception as e:
                        app.logger.warning(f"Error al cargar mapeo de campos (usando valores por defecto): {str(e)}")
                        mapeo_campos = {}  # Si falla, usar valores por defecto hardcodeados
                    
                    # Función auxiliar para obtener valor de entidad
                    def get_entity_value(entity_type, alternative_types=None):
                        tipos_buscar = [entity_type]
                        if alternative_types:
                            tipos_buscar.extend(alternative_types)
                        
                        for tipo_buscar in tipos_buscar:
                            for entity in document.entities:
                                if entity.type_ == tipo_buscar:
                                    # Prioridad 1: normalized_value (valores estructurados)
                                    if hasattr(entity, 'normalized_value') and entity.normalized_value:
                                        norm_val = entity.normalized_value
                                        
                                        # Money value (montos) - pero solo si es mayor que 0
                                        if hasattr(norm_val, 'money_value') and norm_val.money_value:
                                            money = norm_val.money_value
                                            unidades = float(money.units) if hasattr(money, 'units') else 0
                                            nanos = float(money.nanos) / 1e9 if hasattr(money, 'nanos') else 0
                                            valor_money = unidades + nanos
                                            # Solo usar money_value si es mayor que 0, sino usar mention_text
                                            if valor_money > 0:
                                                return valor_money
                                        
                                        # Date value (fechas)
                                        if hasattr(norm_val, 'date_value') and norm_val.date_value:
                                            date_val = norm_val.date_value
                                            year = date_val.year if hasattr(date_val, 'year') else None
                                            month = date_val.month if hasattr(date_val, 'month') else 1
                                            day = date_val.day if hasattr(date_val, 'day') else 1
                                            if year:
                                                return f"{year}-{month:02d}-{day:02d}"
                                        
                                        # Text value (texto normalizado)
                                        if hasattr(norm_val, 'text') and norm_val.text:
                                            return norm_val.text
                                    
                                    # Prioridad 2: mention_text (texto mencionado) - IMPORTANTE cuando money_value es 0.0
                                    if hasattr(entity, 'mention_text') and entity.mention_text:
                                        return entity.mention_text
                                    
                                    # Prioridad 3: text_anchor (extraer del texto completo)
                                    if hasattr(entity, 'text_anchor') and entity.text_anchor and entity.text_anchor.text_segments:
                                        for segment in entity.text_anchor.text_segments:
                                            start_index = segment.start_index if hasattr(segment, 'start_index') else 0
                                            end_index = segment.end_index if hasattr(segment, 'end_index') else len(texto_completo)
                                            if end_index > start_index:
                                                return texto_completo[start_index:end_index]
                        return None
                    
                    # Extraer campos del Invoice Parser usando el mapeo configurado
                    # Si no hay mapeo configurado, usar valores por defecto hardcodeados
                    
                    def extraer_campo_con_mapeo(campo_sistema, valores_por_defecto=None):
                        """Extrae un campo usando el mapeo configurado o valores por defecto."""
                        if campo_sistema in mapeo_campos and mapeo_campos[campo_sistema].get('activo', True):
                            config = mapeo_campos[campo_sistema]
                            campo_ia = config['campo_ia']
                            tipos_alt = config.get('tipos_alternativos', [])
                            valor = get_entity_value(campo_ia, tipos_alt)
                        elif valores_por_defecto:
                            # Usar valores por defecto si no hay mapeo configurado
                            campo_ia = valores_por_defecto[0] if valores_por_defecto else None
                            tipos_alt = valores_por_defecto[1:] if len(valores_por_defecto) > 1 else []
                            valor = get_entity_value(campo_ia, tipos_alt) if campo_ia else None
                        else:
                            valor = None
                        return valor
                    
                    # 1. Número de factura
                    invoice_id = extraer_campo_con_mapeo('numero_factura', ['invoice_id', 'invoice_number', 'invoice_id_number'])
                    if invoice_id:
                        datos_extraidos['numero_factura'] = str(invoice_id).strip()
                        app.logger.info(f"✅ Número de factura (Invoice Parser): {datos_extraidos['numero_factura']}")
                    
                    # 2. Fecha de factura
                    invoice_date = extraer_campo_con_mapeo('fecha_pago', ['invoice_date', 'invoice_date_invoice'])
                    if invoice_date:
                        datos_extraidos['fecha_pago'] = str(invoice_date).strip()
                        app.logger.info(f"✅ Fecha de factura (Invoice Parser): {datos_extraidos['fecha_pago']}")
                    
                    # 3. Fecha de vencimiento
                    due_date = extraer_campo_con_mapeo('fecha_vencimiento', ['due_date', 'invoice_date_due'])
                    if due_date:
                        datos_extraidos['fecha_vencimiento'] = str(due_date).strip()
                        app.logger.info(f"✅ Fecha de vencimiento (Invoice Parser): {datos_extraidos['fecha_vencimiento']}")
                    
                    # 4. Nombre del proveedor
                    supplier_name = extraer_campo_con_mapeo('proveedor', ['supplier_name', 'supplier', 'supplier_name_supplier_name'])
                    if supplier_name:
                        datos_extraidos['proveedor'] = str(supplier_name).strip()
                        app.logger.info(f"✅ Proveedor (Invoice Parser): {datos_extraidos['proveedor']}")
                    
                    # 5. Dirección del proveedor
                    supplier_address = extraer_campo_con_mapeo('proveedor_direccion', ['supplier_address', 'supplier_address_supplier_address'])
                    if supplier_address:
                        datos_extraidos['proveedor_direccion'] = str(supplier_address).strip()[:500]
                        app.logger.info(f"✅ Dirección proveedor (Invoice Parser): {datos_extraidos['proveedor_direccion']}")
                    
                    # 6. Email del proveedor
                    supplier_email = extraer_campo_con_mapeo('proveedor_email', ['supplier_email', 'supplier_email_supplier_email'])
                    if supplier_email:
                        datos_extraidos['proveedor_email'] = str(supplier_email).strip().lower()
                        app.logger.info(f"✅ Email proveedor (Invoice Parser): {datos_extraidos['proveedor_email']}")
                    
                    # 7. Teléfono del proveedor
                    supplier_phone = extraer_campo_con_mapeo('proveedor_telefono', ['supplier_phone', 'supplier_phone_supplier_phone'])
                    if supplier_phone:
                        datos_extraidos['proveedor_telefono'] = str(supplier_phone).strip().replace(' ', '').replace('-', '')
                        app.logger.info(f"✅ Teléfono proveedor (Invoice Parser): {datos_extraidos['proveedor_telefono']}")
                    
                    # 8. NIF/CIF del proveedor
                    supplier_tax_id = extraer_campo_con_mapeo('proveedor_nif', ['supplier_tax_id'])
                    if supplier_tax_id:
                        datos_extraidos['proveedor_nif'] = str(supplier_tax_id).strip()
                        app.logger.info(f"✅ NIF/CIF proveedor (Invoice Parser): {datos_extraidos['proveedor_nif']}")
                    
                    # 8.1. RECEIVER (Sociedad/Pagador): Extraer receiver_name
                    receiver_name = get_entity_value('receiver_name', ['receiver_name', 'receiver', 'customer_name', 'bill_to'])
                    if receiver_name:
                        datos_extraidos['receiver'] = str(receiver_name).strip()
                        app.logger.info(f"✅ Receiver/Sociedad (Invoice Parser): {datos_extraidos['receiver']}")
                    
                    # 8. TOTAL CON IMPUESTOS: Usar mapeo configurado o buscar con lógica especial
                    total_amount = extraer_campo_con_mapeo('monto', ['total_amount', 'invoice_amount', 'total_amount_due', 'amount_due', 'invoice_total'])
                    
                    # Si no encontramos con el mapeo, buscar cualquier entidad con "total" en el nombre
                    # Usar mention_text si money_value es 0.0 o no existe
                    if not total_amount:
                        for entity in document.entities:
                            if 'total' in entity.type_.lower():
                                # Intentar money_value primero
                                if hasattr(entity, 'normalized_value') and entity.normalized_value:
                                    norm_val = entity.normalized_value
                                    if hasattr(norm_val, 'money_value') and norm_val.money_value:
                                        money = norm_val.money_value
                                        unidades = float(money.units) if hasattr(money, 'units') else 0
                                        nanos = float(money.nanos) / 1e9 if hasattr(money, 'nanos') else 0
                                        valor_money = unidades + nanos
                                        if valor_money > 0:
                                            total_amount = valor_money
                                            break
                                
                                # Si money_value es 0.0 o no existe, usar mention_text
                                if not total_amount and hasattr(entity, 'mention_text') and entity.mention_text:
                                    total_amount = entity.mention_text
                                    break
                    
                    if total_amount:
                        try:
                            # Usar función de parsing español
                            if isinstance(total_amount, str):
                                total_con_impuestos = parsear_monto_espanol(total_amount)
                            elif isinstance(total_amount, (int, float)):
                                total_con_impuestos = float(total_amount)
                            else:
                                total_con_impuestos = parsear_monto_espanol(str(total_amount))
                            
                            if total_con_impuestos and total_con_impuestos > 0:
                                datos_extraidos['total_con_impuestos'] = round(total_con_impuestos, 2)
                                datos_extraidos['monto'] = round(total_con_impuestos, 2)
                                app.logger.info(f"✅ Total con Impuestos extraído: {datos_extraidos['total_con_impuestos']} (original: '{total_amount}')")
                            else:
                                app.logger.warning(f"⚠️ parsear_monto_inteligente retornó None para '{total_amount}'")
                        except (ValueError, AttributeError, TypeError) as e:
                            app.logger.warning(f"No se pudo convertir total_amount '{total_amount}': {str(e)}")
                    else:
                        app.logger.warning("⚠️ NO se encontró total_amount en las entidades de Document AI")
                        # RESPALDO: Buscar "Total" cerca de números en el texto
                        # Esto es lo que el usuario sugiere: buscar cerca del texto "Total"
                        import re
                        texto_lower = texto_completo.lower()
                        # Buscar patrones como "Total: 38,22" o "Total 38,22" o "TOTAL 38,22 €"
                        patrones_total = [
                            r'total[:\s]+([\d.,]+)',
                            r'total\s+factura[:\s]+([\d.,]+)',
                            r'importe\s+total[:\s]+([\d.,]+)',
                        ]
                        for patron in patrones_total:
                            match = re.search(patron, texto_lower, re.IGNORECASE)
                            if match:
                                try:
                                    valor_str = match.group(1)
                                    total_encontrado = parsear_monto_espanol(valor_str)
                                    if total_encontrado and total_encontrado > 0:
                                        datos_extraidos['total_con_impuestos'] = round(total_encontrado, 2)
                                        datos_extraidos['monto'] = round(total_encontrado, 2)
                                        app.logger.info(f"✅ Total encontrado por búsqueda de texto cerca de 'Total': {datos_extraidos['total_con_impuestos']} (original: '{valor_str}')")
                                        break
                                except (ValueError, AttributeError):
                                    continue
                    
                    # 9. IVA/IMPUESTOS: Usar mapeo configurado o buscar con lógica especial
                    tax_amount = extraer_campo_con_mapeo('impuestos', ['vat', 'tax_amount', 'total_tax_amount', 'vat_amount', 'tax'])
                    
                    # Si no encontramos con el mapeo, buscar cualquier entidad con "tax", "vat" o "iva" en el nombre
                    # Usar mention_text si money_value es 0.0 o no existe
                    if not tax_amount:
                        for entity in document.entities:
                            tipo_lower = entity.type_.lower()
                            if 'tax' in tipo_lower or 'vat' in tipo_lower or 'iva' in tipo_lower:
                                # Intentar money_value primero (si es > 0)
                                if hasattr(entity, 'normalized_value') and entity.normalized_value:
                                    norm_val = entity.normalized_value
                                    if hasattr(norm_val, 'money_value') and norm_val.money_value:
                                        money = norm_val.money_value
                                        unidades = float(money.units) if hasattr(money, 'units') else 0
                                        nanos = float(money.nanos) / 1e9 if hasattr(money, 'nanos') else 0
                                        valor_money = unidades + nanos
                                        if valor_money > 0:
                                            tax_amount = valor_money
                                            break
                                
                                # Si money_value es 0.0 o no existe, usar mention_text (más confiable en este caso)
                                if not tax_amount and hasattr(entity, 'mention_text') and entity.mention_text:
                                    tax_amount = entity.mention_text
                                    break
                    
                    if tax_amount:
                        try:
                            # Usar función de parsing español
                            if isinstance(tax_amount, str):
                                impuestos = parsear_monto_espanol(tax_amount)
                            elif isinstance(tax_amount, (int, float)):
                                impuestos = float(tax_amount)
                            else:
                                impuestos = parsear_monto_espanol(str(tax_amount))
                            
                            if impuestos and impuestos > 0:
                                # VALIDACIÓN: Si el valor es muy pequeño (< 1), probablemente sea porcentaje, no importe
                                if impuestos < 1:
                                    app.logger.warning(f"⚠️ tax_amount '{tax_amount}' es < 1 ({impuestos}), probablemente sea porcentaje, no importe. Se descarta.")
                                else:
                                    datos_extraidos['impuestos'] = round(impuestos, 2)
                                    datos_extraidos['iva'] = round(impuestos, 2)
                                    app.logger.info(f"✅ Impuestos/IVA extraído: {datos_extraidos['impuestos']} (original: '{tax_amount}')")
                        except (ValueError, AttributeError, TypeError) as e:
                            app.logger.warning(f"No se pudo convertir tax_amount '{tax_amount}': {str(e)}")
                    else:
                        app.logger.warning("⚠️ NO se encontró tax_amount en las entidades de Document AI")
                    
                    # 9.1. PORCENTAJES DE IVA: Detectar todos los tipos de IVA presentes en la factura
                    import re
                    porcentajes_iva_detectados = []
                    iva_porcentaje = None
                    
                    # Buscar en entidades de Document AI (puede haber múltiples)
                    for entity in document.entities:
                        tipo_lower = entity.type_.lower()
                        if 'vat_rate' in tipo_lower or 'tax_rate' in tipo_lower:
                            try:
                                valor = None
                                if hasattr(entity, 'normalized_value') and entity.normalized_value:
                                    if hasattr(entity.normalized_value, 'float_value'):
                                        valor = entity.normalized_value.float_value
                                    elif hasattr(entity, 'mention_text') and entity.mention_text:
                                        valor = entity.mention_text
                                elif hasattr(entity, 'mention_text') and entity.mention_text:
                                    valor = entity.mention_text
                                
                                if valor:
                                    if isinstance(valor, str):
                                        vat_str = str(valor).replace('%', '').replace(',', '.').strip()
                                        porcentaje = float(vat_str)
                                        if porcentaje < 1:
                                            porcentaje = porcentaje * 100
                                    else:
                                        porcentaje = float(valor)
                                        if porcentaje < 1:
                                            porcentaje = porcentaje * 100
                                    
                                    if porcentaje > 0 and porcentaje <= 100:
                                        porcentajes_iva_detectados.append(round(porcentaje, 2))
                            except (ValueError, AttributeError, TypeError):
                                continue
                    
                    # Buscar todos los porcentajes de IVA en el texto (puede haber múltiples)
                    texto_lower = texto_completo.lower()
                    patrones_iva_porcentaje = [
                        r'iva\s+(?:al\s+)?(\d+(?:[.,]\d+)?)\s*%',
                        r'(\d+(?:[.,]\d+)?)\s*%\s+iva',
                        r'iva\s*\((\d+(?:[.,]\d+)?)\s*%\)',
                        r'tipo\s+iva[:\s]+(\d+(?:[.,]\d+)?)\s*%',
                        r'(\d+(?:[.,]\d+)?)\s*%',  # Buscar cualquier porcentaje cerca de "IVA"
                    ]
                    
                    for patron in patrones_iva_porcentaje:
                        matches = re.finditer(patron, texto_lower, re.IGNORECASE)
                        for match in matches:
                            try:
                                porcentaje_str = match.group(1).replace(',', '.').strip()
                                porcentaje = float(porcentaje_str)
                                # Solo incluir porcentajes válidos en España (4%, 10%, 21%, y valores cercanos)
                                if porcentaje > 0 and porcentaje <= 100:
                                    porcentajes_iva_detectados.append(round(porcentaje, 2))
                            except (ValueError, AttributeError):
                                continue
                    
                    # Eliminar duplicados y ordenar
                    porcentajes_iva_detectados = sorted(list(set(porcentajes_iva_detectados)))
                    
                    # Filtrar solo porcentajes válidos en España (4%, 10%, 21% y valores cercanos)
                    porcentajes_validos_espana = [4, 10, 21]
                    porcentajes_iva_validos = []
                    for p in porcentajes_iva_detectados:
                        # Verificar si es un porcentaje válido o está cerca de uno válido
                        for pv in porcentajes_validos_espana:
                            if abs(p - pv) < 0.5:  # Tolerancia de 0.5 puntos porcentuales
                                porcentajes_iva_validos.append(pv)
                                break
                    
                    # Eliminar duplicados y ordenar
                    porcentajes_iva_validos = sorted(list(set(porcentajes_iva_validos)))
                    
                    if porcentajes_iva_validos:
                        datos_extraidos['iva_porcentajes_detectados'] = porcentajes_iva_validos
                        # Usar el porcentaje más común o el más alto como principal
                        iva_porcentaje = porcentajes_iva_validos[-1]  # El más alto por defecto
                        app.logger.info(f"✅ Porcentajes de IVA detectados: {porcentajes_iva_validos}%")
                    
                    # Calcular porcentaje efectivo si tenemos IVA total y Base Imponible total
                    porcentaje_efectivo = None
                    if 'impuestos' in datos_extraidos and 'total' in datos_extraidos:
                        try:
                            if datos_extraidos['total'] > 0:
                                porcentaje_efectivo = (datos_extraidos['impuestos'] / datos_extraidos['total']) * 100
                                porcentaje_efectivo = round(porcentaje_efectivo, 2)
                                datos_extraidos['iva_porcentaje_efectivo'] = porcentaje_efectivo
                                app.logger.info(f"✅ Porcentaje de IVA efectivo calculado: {porcentaje_efectivo}%")
                                
                                # Si no hay porcentajes detectados, usar el efectivo redondeado al más cercano
                                if not iva_porcentaje:
                                    porcentajes_validos = [4, 10, 21]
                                    porcentaje_mas_cercano = min(porcentajes_validos, key=lambda x: abs(x - porcentaje_efectivo))
                                    if abs(porcentaje_mas_cercano - porcentaje_efectivo) < 2:
                                        iva_porcentaje = porcentaje_mas_cercano
                                        datos_extraidos['iva_porcentaje'] = iva_porcentaje
                                        app.logger.info(f"✅ Porcentaje de IVA principal (calculado): {iva_porcentaje}%")
                        except (ValueError, TypeError, ZeroDivisionError) as e:
                            app.logger.warning(f"No se pudo calcular porcentaje efectivo de IVA: {str(e)}")
                    
                    # Si tenemos porcentajes detectados, usar el principal
                    if iva_porcentaje and 'iva_porcentaje' not in datos_extraidos:
                        datos_extraidos['iva_porcentaje'] = iva_porcentaje
                        app.logger.info(f"✅ Porcentaje de IVA principal: {iva_porcentaje}%")
                    
                    # Si tenemos TOTAL y Base Imponible pero no IVA ni porcentaje, calcular ambos
                    if 'iva_porcentaje' not in datos_extraidos and 'impuestos' not in datos_extraidos:
                        if 'total_con_impuestos' in datos_extraidos and 'total' in datos_extraidos:
                            try:
                                if datos_extraidos['total'] > 0:
                                    iva_calculado = datos_extraidos['total_con_impuestos'] - datos_extraidos['total']
                                    if iva_calculado > 0:
                                        datos_extraidos['impuestos'] = round(iva_calculado, 2)
                                        datos_extraidos['iva'] = round(iva_calculado, 2)
                                        # Calcular también el porcentaje efectivo
                                        porcentaje_efectivo = (iva_calculado / datos_extraidos['total']) * 100
                                        datos_extraidos['iva_porcentaje_efectivo'] = round(porcentaje_efectivo, 2)
                                        porcentajes_validos = [4, 10, 21]
                                        porcentaje_mas_cercano = min(porcentajes_validos, key=lambda x: abs(x - porcentaje_efectivo))
                                        if abs(porcentaje_mas_cercano - porcentaje_efectivo) < 2:
                                            datos_extraidos['iva_porcentaje'] = porcentaje_mas_cercano
                                        app.logger.info(f"✅ IVA calculado (TOTAL - Base Imponible): {datos_extraidos['impuestos']}, Porcentaje efectivo: {datos_extraidos.get('iva_porcentaje_efectivo', 'N/A')}%, Principal: {datos_extraidos.get('iva_porcentaje', 'N/A')}%")
                            except (ValueError, TypeError) as e:
                                app.logger.warning(f"No se pudo calcular IVA desde TOTAL y Base Imponible: {str(e)}")
                    
                    # 10. BASE IMPONIBLE: Usar mapeo configurado o buscar con lógica especial
                    subtotal_amount = extraer_campo_con_mapeo('base_imponible', ['net_amount', 'subtotal_amount', 'subtotal', 'line_item_amount', 'amount'])
                    
                    # Si no encontramos con el mapeo, buscar cualquier entidad con "subtotal" o "net" en el nombre
                    if not subtotal_amount:
                        for entity in document.entities:
                            tipo_lower = entity.type_.lower()
                            if ('subtotal' in tipo_lower or 'net' in tipo_lower) and hasattr(entity, 'normalized_value') and entity.normalized_value:
                                norm_val = entity.normalized_value
                                if hasattr(norm_val, 'money_value') and norm_val.money_value:
                                    money = norm_val.money_value
                                    unidades = float(money.units) if hasattr(money, 'units') else 0
                                    nanos = float(money.nanos) / 1e9 if hasattr(money, 'nanos') else 0
                                    valor_money = unidades + nanos
                                    if valor_money > 0:
                                        subtotal_amount = valor_money
                                        break
                                # Si no tiene money_value, usar mention_text
                                if not subtotal_amount and hasattr(entity, 'mention_text') and entity.mention_text:
                                    subtotal_amount = entity.mention_text
                                    break
                    
                    if subtotal_amount:
                        try:
                            # Usar función de parsing español
                            if isinstance(subtotal_amount, (int, float)):
                                base_imponible = float(subtotal_amount)
                            else:
                                base_imponible = parsear_monto_espanol(str(subtotal_amount))
                            
                            if base_imponible and base_imponible > 0:
                                datos_extraidos['total'] = round(base_imponible, 2)
                                app.logger.info(f"✅ Base Imponible extraída: {datos_extraidos['total']} (original: '{subtotal_amount}')")
                        except (ValueError, AttributeError, TypeError) as e:
                            app.logger.warning(f"No se pudo convertir base imponible '{subtotal_amount}': {str(e)}")
                    else:
                        app.logger.warning("⚠️ NO se encontró subtotal_amount en las entidades de Document AI")
                        # RESPALDO: Buscar "Base Imponible" o "Base" cerca de números en el texto
                        # Esto es lo que el usuario sugiere: buscar cerca del texto "Base Imponible"
                        import re
                        texto_lower = texto_completo.lower()
                        # Buscar patrones como "Base Imponible: 36,75" o "Base Imponible 36,75"
                        patrones_base = [
                            r'base\s+imponible[:\s]+([\d.,]+)',
                            r'base[:\s]+([\d.,]+)',
                            r'importe\s+bruto[:\s]+([\d.,]+)',
                        ]
                        for patron in patrones_base:
                            match = re.search(patron, texto_lower, re.IGNORECASE)
                            if match:
                                try:
                                    valor_str = match.group(1)
                                    base_imponible = parsear_monto_espanol(valor_str)
                                    if base_imponible and base_imponible > 0:
                                        datos_extraidos['total'] = round(base_imponible, 2)
                                        app.logger.info(f"✅ Base Imponible encontrada por búsqueda de texto cerca de 'Base Imponible': {datos_extraidos['total']} (original: '{valor_str}')")
                                        break
                                except (ValueError, AttributeError):
                                    continue
                    
                    # Si no tenemos base imponible pero sí tenemos total e IVA, calcularla
                    # Esto es válido porque: Base Imponible = Total con Impuestos - IVA
                    if 'total' not in datos_extraidos and 'total_con_impuestos' in datos_extraidos and 'impuestos' in datos_extraidos:
                        try:
                            base_calculada = datos_extraidos['total_con_impuestos'] - datos_extraidos['impuestos']
                            if base_calculada > 0:
                                datos_extraidos['total'] = round(base_calculada, 2)
                                app.logger.info(f"✅ Base Imponible calculada (Total - IVA): {datos_extraidos['total']}")
                        except (ValueError, TypeError) as e:
                            app.logger.warning(f"No se pudo calcular base imponible: {str(e)}")
                    
                    # VALIDACIÓN DE COHERENCIA: Verificar que Base + IVA = Total
                    if 'total' in datos_extraidos and 'impuestos' in datos_extraidos and 'total_con_impuestos' in datos_extraidos:
                        base = datos_extraidos['total']
                        iva = datos_extraidos['impuestos']
                        total = datos_extraidos['total_con_impuestos']
                        suma_calculada = round(base + iva, 2)
                        diferencia = abs(suma_calculada - total)
                        
                        if diferencia > 0.50:  # Tolerancia de 0.50€
                            app.logger.warning(f"⚠️ INCOHERENCIA DETECTADA: Base ({base}) + IVA ({iva}) = {suma_calculada}, pero Total extraído es {total}. Diferencia: {diferencia}€")
                            app.logger.warning(f"⚠️ Se recalcularán los valores para corregir el error...")
                            
                            # Si el total parece correcto pero la suma no coincide, buscar IVA en el texto
                            if 'impuestos' in datos_extraidos:
                                # Buscar el importe de IVA en el texto
                                import re
                                texto_lower = texto_completo.lower()
                                patrones_iva_importe = [
                                    r'imp\.?\s*iva[:\s]+([\d.,]+)',
                                    r'importe\s+iva[:\s]+([\d.,]+)',
                                    r'cuota\s+iva[:\s]+([\d.,]+)',
                                    r'iva[:\s]+([\d.,]+)\s*€',
                                ]
                                iva_encontrado = None
                                for patron in patrones_iva_importe:
                                    match = re.search(patron, texto_lower, re.IGNORECASE)
                                    if match:
                                        valor_str = match.group(1)
                                        iva_encontrado = parsear_monto_espanol(valor_str)
                                        if iva_encontrado and iva_encontrado > 0:
                                            app.logger.info(f"✅ IVA corregido por búsqueda en texto: {iva_encontrado} (original: '{valor_str}')")
                                            datos_extraidos['impuestos'] = round(iva_encontrado, 2)
                                            datos_extraidos['iva'] = round(iva_encontrado, 2)
                                            break
                                
                                # Recalcular para verificar coherencia
                                if iva_encontrado:
                                    suma_calculada = round(base + iva_encontrado, 2)
                                    diferencia_nueva = abs(suma_calculada - total)
                                    if diferencia_nueva <= 0.50:
                                        app.logger.info(f"✅ Coherencia restaurada: Base ({base}) + IVA ({iva_encontrado}) = {suma_calculada} ≈ Total ({total})")
                                    else:
                                        app.logger.warning(f"⚠️ Aún incoherente después de corrección: {suma_calculada} vs {total}, diferencia: {diferencia_nueva}€")
                        else:
                            app.logger.info(f"✅ Validación de coherencia: Base ({base}) + IVA ({iva}) = {suma_calculada} ≈ Total ({total}). Diferencia: {diferencia}€")
                    
                    # 10. Términos de pago / Método de pago
                    payment_terms = get_entity_value('payment_terms', ['payment_terms_payment_terms'])
                    if payment_terms:
                        terminos = str(payment_terms).strip().lower()
                        # Intentar inferir método de pago de los términos
                        if 'transferencia' in terminos or 'transfer' in terminos:
                            datos_extraidos['metodo_pago'] = 'transferencia'
                        elif 'remesa' in terminos:
                            datos_extraidos['metodo_pago'] = 'remesa'
                        elif 'metálico' in terminos or 'efectivo' in terminos or 'cash' in terminos or 'metalico' in terminos:
                            datos_extraidos['metodo_pago'] = 'metálico'
                        elif 'cheque' in terminos:
                            datos_extraidos['metodo_pago'] = 'cheque'
                        app.logger.info(f"✅ Términos de pago (Invoice Parser): {payment_terms}")
                    
                    # Solo extraemos datos básicos de factura (no artículos/líneas)
                    app.logger.info(f"📊 Resumen datos extraídos con Invoice Parser: {len([k for k in datos_extraidos.keys() if k != 'texto_completo'])} campos")
                
                # NO USAR REGEX - Confiar completamente en Invoice Parser
                # Si Invoice Parser no extrae algo, simplemente no lo tenemos
                # No hacer procesamiento manual adicional
                
                # DETECTAR RESIDENCIA AUTOMÁTICAMENTE desde el texto de la factura
                # Obtener residencias del usuario si aún no se han obtenido
                if residencias_usuario is None:
                    conn_residencias = get_db_connection()
                    cursor_residencias = conn_residencias.cursor()
                    try:
                        # Verificar qué columnas existen en la tabla residencia
                        cursor_residencias.execute("""
                            SELECT column_name 
                            FROM information_schema.columns 
                            WHERE table_name = 'residencia'
                            ORDER BY ordinal_position
                        """)
                        columnas_existentes = {row[0] for row in cursor_residencias.fetchall()}
                        
                        # Construir SELECT dinámicamente según las columnas disponibles
                        columnas_select_list = ['id_residencia', 'nombre']
                        columnas_adicionales = {
                            'nombre_fiscal': 'nombre_fiscal',
                            'nif': 'nif',
                            'direccion': 'direccion',
                            'ciudad': 'ciudad',
                            'provincia': 'provincia',
                            'codigo_postal': 'codigo_postal'
                        }
                        
                        for columna_db, columna_alias in columnas_adicionales.items():
                            if columna_db in columnas_existentes:
                                columnas_select_list.append(columna_db)
                            else:
                                columnas_select_list.append(f'NULL as {columna_alias}')
                        
                        columnas_select = ', '.join(columnas_select_list)
                        
                        if g.id_rol == ADMIN_ROLE_ID:
                            cursor_residencias.execute(f"""
                                SELECT {columnas_select}
                                FROM residencia
                                WHERE activa = TRUE
                                ORDER BY nombre
                            """)
                        else:
                            if g.residencias_acceso:
                                placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                                cursor_residencias.execute(f"""
                                    SELECT {columnas_select}
                                    FROM residencia
                                    WHERE activa = TRUE AND id_residencia IN ({placeholders})
                                    ORDER BY nombre
                                """, tuple(g.residencias_acceso))
                        residencias_usuario = cursor_residencias.fetchall()
                    except Exception as e:
                        app.logger.error(f"Error al obtener residencias para detección automática: {str(e)}")
                        residencias_usuario = []
                    finally:
                        cursor_residencias.close()
                        conn_residencias.close()
                
                texto_lower = texto_completo.lower()
                mejor_coincidencia = None
                mejor_puntuacion = 0
                
                for residencia_data in residencias_usuario:
                    # Manejar diferentes números de columnas según lo que exista
                    id_res = residencia_data[0]
                    nombre = residencia_data[1] if len(residencia_data) > 1 else None
                    nombre_fiscal = residencia_data[2] if len(residencia_data) > 2 else None
                    nif = residencia_data[3] if len(residencia_data) > 3 else None
                    direccion = residencia_data[4] if len(residencia_data) > 4 else None
                    ciudad = residencia_data[5] if len(residencia_data) > 5 else None
                    provincia = residencia_data[6] if len(residencia_data) > 6 else None
                    codigo_postal = residencia_data[7] if len(residencia_data) > 7 else None
                    puntuacion = 0
                    
                    # Buscar coincidencias con el nombre de la residencia
                    if nombre:
                        nombre_lower = nombre.lower()
                        if nombre_lower in texto_lower:
                            puntuacion += 10
                            # Si aparece completo, más puntos
                            if f' {nombre_lower} ' in f' {texto_lower} ':
                                puntuacion += 5
                    
                    # Buscar coincidencias con el nombre fiscal
                    if nombre_fiscal:
                        nombre_fiscal_lower = nombre_fiscal.lower()
                        if nombre_fiscal_lower in texto_lower:
                            puntuacion += 15  # Nombre fiscal es más específico
                    
                    # Buscar coincidencias con el NIF/CIF
                    if nif:
                        nif_limpio = nif.replace(' ', '').replace('-', '').upper()
                        nif_pattern = re.escape(nif_limpio).replace('\\', '')
                        if re.search(nif_pattern, texto_completo, re.IGNORECASE):
                            puntuacion += 20  # NIF es muy específico
                    
                    # Buscar coincidencias con ciudad
                    if ciudad:
                        ciudad_lower = ciudad.lower()
                        if ciudad_lower in texto_lower:
                            puntuacion += 3
                    
                    # Buscar coincidencias con código postal
                    if codigo_postal:
                        if codigo_postal in texto_completo:
                            puntuacion += 5
                    
                    # Si hay una buena coincidencia, guardarla
                    if puntuacion > mejor_puntuacion:
                        mejor_puntuacion = puntuacion
                        mejor_coincidencia = {
                            'id_residencia': id_res,
                            'nombre': nombre,
                            'puntuacion': puntuacion
                        }
                
                # Si encontramos una buena coincidencia (mínimo 10 puntos), usarla
                if mejor_coincidencia and mejor_puntuacion >= 10:
                    id_residencia_detectada = mejor_coincidencia['id_residencia']
                    app.logger.info(f"Residencia detectada automáticamente: {mejor_coincidencia['nombre']} (ID: {id_residencia_detectada}, Puntuación: {mejor_puntuacion})")
                    datos_extraidos['id_residencia_detectada'] = id_residencia_detectada
                    datos_extraidos['nombre_residencia_detectada'] = mejor_coincidencia['nombre']
                else:
                    app.logger.info(f"No se pudo detectar la residencia automáticamente (mejor puntuación: {mejor_puntuacion})")
                
                # Guardar texto_completo para incluirlo en la respuesta (solo para debugging/pruebas)
                datos_extraidos['texto_completo'] = texto_completo
                
                # MEJORA: Si faltan campos importantes, intentar extraerlos del texto completo como respaldo
                # Esto es especialmente útil cuando Document AI no detecta todos los campos
                import re
                texto_lower = texto_completo.lower()
                
                # Si no tenemos TOTAL, buscar en el texto
                if 'total_con_impuestos' not in datos_extraidos and 'monto' not in datos_extraidos:
                    patrones_total = [
                        r'total[:\s]+([\d.,]+)\s*€?',
                        r'total\s+factura[:\s]+([\d.,]+)\s*€?',
                        r'importe\s+total[:\s]+([\d.,]+)\s*€?',
                        r'total\s+a\s+pagar[:\s]+([\d.,]+)\s*€?',
                        r'total\s+incluido[:\s]+([\d.,]+)\s*€?',
                    ]
                    for patron in patrones_total:
                        match = re.search(patron, texto_completo, re.IGNORECASE)
                        if match:
                            try:
                                valor_str = match.group(1).replace('.', '').replace(',', '.').strip()
                                total_encontrado = float(valor_str)
                                if total_encontrado > 0:
                                    datos_extraidos['total_con_impuestos'] = round(total_encontrado, 2)
                                    datos_extraidos['monto'] = round(total_encontrado, 2)
                                    app.logger.info(f"✅ Total encontrado por búsqueda en texto: {datos_extraidos['total_con_impuestos']}")
                                    break
                            except (ValueError, AttributeError):
                                continue
                
                # Si no tenemos BASE IMPONIBLE, buscar en el texto
                if 'total' not in datos_extraidos:
                    patrones_base = [
                        r'base\s+imponible[:\s]+([\d.,]+)\s*€?',
                        r'base[:\s]+([\d.,]+)\s*€?',
                        r'importe\s+bruto[:\s]+([\d.,]+)\s*€?',
                        r'subtotal[:\s]+([\d.,]+)\s*€?',
                        r'importe\s+base[:\s]+([\d.,]+)\s*€?',
                    ]
                    for patron in patrones_base:
                        match = re.search(patron, texto_completo, re.IGNORECASE)
                        if match:
                            try:
                                valor_str = match.group(1).replace('.', '').replace(',', '.').strip()
                                base_encontrada = float(valor_str)
                                if base_encontrada > 0:
                                    datos_extraidos['total'] = round(base_encontrada, 2)
                                    app.logger.info(f"✅ Base Imponible encontrada por búsqueda en texto: {datos_extraidos['total']}")
                                    break
                            except (ValueError, AttributeError):
                                continue
                
                # Si no tenemos IVA, buscar en el texto
                if 'impuestos' not in datos_extraidos and 'iva' not in datos_extraidos:
                    patrones_iva = [
                        r'iva[:\s]+([\d.,]+)\s*€?',
                        r'impuestos[:\s]+([\d.,]+)\s*€?',
                        r'importe\s+iva[:\s]+([\d.,]+)\s*€?',
                        r'cuota\s+iva[:\s]+([\d.,]+)\s*€?',
                    ]
                    for patron in patrones_iva:
                        match = re.search(patron, texto_completo, re.IGNORECASE)
                        if match:
                            try:
                                valor_str = match.group(1).replace('.', '').replace(',', '.').strip()
                                iva_encontrado = float(valor_str)
                                if iva_encontrado > 0:
                                    datos_extraidos['impuestos'] = round(iva_encontrado, 2)
                                    datos_extraidos['iva'] = round(iva_encontrado, 2)
                                    app.logger.info(f"✅ IVA encontrado por búsqueda en texto: {datos_extraidos['impuestos']}")
                                    break
                            except (ValueError, AttributeError):
                                continue
                
                # Si no tenemos PROVEEDOR, buscar en el texto (buscar después de "proveedor", "emisor", etc.)
                if 'proveedor' not in datos_extraidos:
                    patrones_proveedor = [
                        r'(?:proveedor|emisor|facturador|vendedor)[:\s]+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ\s,\.]+?)(?:\n|€|NIF|CIF|IVA|Total|$)',
                        r'factura\s+a[:\s]+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ\s,\.]+?)(?:\n|€|NIF|CIF|IVA|Total|$)',
                    ]
                    for patron in patrones_proveedor:
                        match = re.search(patron, texto_completo, re.IGNORECASE | re.MULTILINE)
                        if match:
                            proveedor_encontrado = match.group(1).strip()
                            # Limpiar el nombre (quitar espacios extra, líneas nuevas, etc.)
                            proveedor_encontrado = ' '.join(proveedor_encontrado.split())
                            if len(proveedor_encontrado) > 3 and len(proveedor_encontrado) < 200:
                                datos_extraidos['proveedor'] = proveedor_encontrado
                                app.logger.info(f"✅ Proveedor encontrado por búsqueda en texto: {datos_extraidos['proveedor']}")
                                break
                
                # Si no tenemos NIF del proveedor, buscar en el texto
                if 'proveedor_nif' not in datos_extraidos:
                    patrones_nif = [
                        r'(?:nif|cif|n\.?i\.?f\.?|c\.?i\.?f\.?)[:\s]*([A-Z0-9]{8,10})',
                        r'([A-Z][0-9]{8})',  # Formato básico: A12345678
                        r'([A-Z]{2}[0-9]{7}[A-Z0-9])',  # Formato CIF: A1234567X
                    ]
                    for patron in patrones_nif:
                        matches = re.finditer(patron, texto_completo, re.IGNORECASE)
                        for match in matches:
                            nif_encontrado = match.group(1).upper().strip()
                            # Validar que parece un NIF/CIF válido
                            if len(nif_encontrado) >= 8 and len(nif_encontrado) <= 10:
                                datos_extraidos['proveedor_nif'] = nif_encontrado
                                app.logger.info(f"✅ NIF/CIF encontrado por búsqueda en texto: {datos_extraidos['proveedor_nif']}")
                                break
                        if 'proveedor_nif' in datos_extraidos:
                            break
                
                # Si no tenemos TELÉFONO del proveedor, buscar en el texto
                if 'proveedor_telefono' not in datos_extraidos:
                    patrones_telefono = [
                        r'(?:tel|teléfono|tlf|tfn|phone)[:\s]*([0-9]{9,12})',
                        r'([0-9]{3}[\s\-]?[0-9]{3}[\s\-]?[0-9]{3})',  # Formato: 666 555 444
                        r'([0-9]{9})',  # Formato: 666555444
                    ]
                    for patron in patrones_telefono:
                        matches = re.finditer(patron, texto_completo, re.IGNORECASE)
                        for match in matches:
                            telefono_encontrado = match.group(1).replace(' ', '').replace('-', '').strip()
                            if len(telefono_encontrado) >= 9:
                                datos_extraidos['proveedor_telefono'] = telefono_encontrado
                                app.logger.info(f"✅ Teléfono encontrado por búsqueda en texto: {datos_extraidos['proveedor_telefono']}")
                                break
                        if 'proveedor_telefono' in datos_extraidos:
                            break
                
                # Si no tenemos EMAIL del proveedor, buscar en el texto
                if 'proveedor_email' not in datos_extraidos:
                    patron_email = r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})'
                    match = re.search(patron_email, texto_completo, re.IGNORECASE)
                    if match:
                        email_encontrado = match.group(1).lower().strip()
                        datos_extraidos['proveedor_email'] = email_encontrado
                        app.logger.info(f"✅ Email encontrado por búsqueda en texto: {datos_extraidos['proveedor_email']}")
                
                # Si no tenemos RECEIVER (Sociedad/Pagador), buscar en el texto
                if 'receiver' not in datos_extraidos:
                    patrones_receiver = [
                        r'(?:receptor|pagador|cliente|sociedad|facturado\s+a)[:\s]+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ\s,\.S\.L\.]+?)(?:\n|€|NIF|CIF|IVA|Total|$)',
                        r'factura\s+a[:\s]+([A-ZÁÉÍÓÚÑ][A-ZÁÉÍÓÚÑa-záéíóúñ\s,\.S\.L\.]+?)(?:\n|€|NIF|CIF|IVA|Total|$)',
                    ]
                    for patron in patrones_receiver:
                        match = re.search(patron, texto_completo, re.IGNORECASE | re.MULTILINE)
                        if match:
                            receiver_encontrado = match.group(1).strip()
                            # Limpiar el nombre (quitar espacios extra, líneas nuevas, etc.)
                            receiver_encontrado = ' '.join(receiver_encontrado.split())
                            if len(receiver_encontrado) > 3 and len(receiver_encontrado) < 200:
                                datos_extraidos['receiver'] = receiver_encontrado
                                app.logger.info(f"✅ Receiver/Sociedad encontrado por búsqueda en texto: {datos_extraidos['receiver']}")
                                break
                
                # Recalcular porcentaje de IVA si ahora tenemos IVA y Base Imponible
                if 'impuestos' in datos_extraidos and 'total' in datos_extraidos and 'iva_porcentaje' not in datos_extraidos:
                    try:
                        if datos_extraidos['total'] > 0:
                            porcentaje_efectivo = (datos_extraidos['impuestos'] / datos_extraidos['total']) * 100
                            porcentajes_validos = [4, 10, 21]
                            porcentaje_mas_cercano = min(porcentajes_validos, key=lambda x: abs(x - porcentaje_efectivo))
                            if abs(porcentaje_mas_cercano - porcentaje_efectivo) < 2:
                                datos_extraidos['iva_porcentaje'] = porcentaje_mas_cercano
                                datos_extraidos['iva_porcentaje_efectivo'] = round(porcentaje_efectivo, 2)
                                app.logger.info(f"✅ Porcentaje de IVA calculado después de búsqueda en texto: {datos_extraidos['iva_porcentaje']}%")
                    except (ValueError, TypeError, ZeroDivisionError):
                        pass
                
                app.logger.info(f"📊 Datos extraídos después de búsqueda en texto: {[k for k in datos_extraidos.keys() if k != 'texto_completo']}")
                
                # Guardar texto_completo para la respuesta (fuera de datos_extraidos)
                texto_completo_respuesta = texto_completo
                
                app.logger.info(f"Datos extraídos para factura: {datos_extraidos}")
                
        except ImportError as e:
            # Si falta el paquete google-cloud-documentai
            app.logger.error(f"Paquete google-cloud-documentai no instalado: {str(e)}")
            app.logger.error("Instale el paquete con: pip install google-cloud-documentai")
            app.logger.warning("Continuando sin extracción automática de datos")
            texto_completo_respuesta = ''
        except Exception as e:
            # Si falla Document AI, continuar sin datos extraídos
            error_msg = str(e)
            
            # Detectar errores específicos de permisos
            if 'PermissionDenied' in error_msg or 'IAM_PERMISSION_DENIED' in error_msg or 'documentai.processors.processOnline' in error_msg:
                app.logger.error("Error de permisos en Document AI: La cuenta de servicio no tiene permisos para usar Document AI")
                app.logger.error("Asegúrese de que la cuenta de servicio tenga el rol 'Document AI API User' o 'Document AI API Editor'")
                app.logger.error("Puede otorgar permisos en: https://console.cloud.google.com/iam-admin/iam")
                app.logger.warning("Continuando sin extracción automática de datos")
            elif 'Invalid location' in error_msg or 'must match the server deployment' in error_msg:
                app.logger.error("Error de ubicación en Document AI: La ubicación del procesador no coincide con la del servidor")
                app.logger.error(f"El procesador está configurado para '{location}' pero el servidor está en otra región")
                app.logger.error("Configure DOCUMENT_AI_LOCATION en .env con la ubicación correcta (us, eu, etc.)")
                app.logger.warning("Continuando sin extracción automática de datos")
            elif 'processor' in error_msg.lower() and ('not exist' in error_msg.lower() or 'not found' in error_msg.lower() or 'does not exist' in error_msg.lower() or 'not available' in error_msg.lower()):
                app.logger.error("Error: El procesador de Document AI no existe o no está disponible")
                app.logger.error(f"PROCESSOR_ID configurado: {processor_id}")
                app.logger.error(f"LOCATION configurada: {location}")
                app.logger.error("Verifique que el PROCESSOR_ID y DOCUMENT_AI_LOCATION en las variables de entorno sean correctos")
                app.logger.error("Puede verificar los procesadores disponibles en: https://console.cloud.google.com/ai/document-ai/processors")
                app.logger.warning("Continuando sin extracción automática de datos")
            else:
                app.logger.warning(f"Error al procesar factura con Document AI: {error_msg}")
                app.logger.warning("Continuando sin extracción automática de datos")
                import traceback
                app.logger.debug(traceback.format_exc())
            
            texto_completo_respuesta = ''
        
        # Si se detectó una residencia automáticamente y el usuario tiene acceso, usarla
        if id_residencia_detectada:
            # Verificar que el usuario tiene acceso a la residencia detectada
            if g.id_rol == ADMIN_ROLE_ID or id_residencia_detectada in g.residencias_acceso:
                id_residencia = id_residencia_detectada
                app.logger.info(f"Usando residencia detectada automáticamente: {id_residencia}")
            else:
                app.logger.warning(f"Usuario no tiene acceso a la residencia detectada {id_residencia_detectada}, usando residencia proporcionada")
        
        # Determinar content_type para guardar en Cloud Storage
        nombre_archivo_lower = nombre_archivo.lower() if nombre_archivo else ''
        if es_imagen:
            if nombre_archivo_lower.endswith('.png') or (hasattr(archivo, 'content_type') and archivo.content_type == 'image/png'):
                content_type = 'image/png'
            else:
                content_type = 'image/jpeg'
        else:
            content_type = 'application/pdf'
        
        # Guardar archivo (PDF o imagen) en Cloud Storage (usar id_residencia detectada o proporcionada)
        from storage_manager import upload_document_unificado
        blob_path = upload_document_unificado(
            file_content, id_residencia, 'pago_proveedor', 0,  # id_entidad será 0 hasta que se cree el pago
            'Factura', nombre_archivo, content_type
        )
        
        if not blob_path:
            return jsonify({'error': 'Error al subir el archivo a Cloud Storage'}), 500
        
        app.logger.info(f"Retornando respuesta con datos_extraidos: {datos_extraidos}")
        
        # Generar URL firmada para el PDF (válida por 1 hora)
        pdf_url = None
        if blob_path:
            app.logger.info(f"📄 Generando URL firmada para PDF: {blob_path}")
            from storage_manager import get_document_url
            try:
                pdf_url = get_document_url(blob_path, expiration_minutes=60)
                if pdf_url:
                    app.logger.info(f"✅ PDF URL generada exitosamente: {pdf_url[:100]}...")
                else:
                    app.logger.warning(f"⚠️ get_document_url retornó None para blob_path: {blob_path}")
            except Exception as e:
                app.logger.error(f"❌ Error al generar URL firmada del PDF: {str(e)}")
                import traceback
                app.logger.error(traceback.format_exc())
        else:
            app.logger.warning("⚠️ blob_path es None o vacío, no se puede generar PDF URL")
        
        # Preparar entidades disponibles para el frontend (formato simplificado)
        entidades_disponibles_respuesta = {}
        if 'entidades_detalladas' in locals() and entidades_detalladas:
            for tipo_entidad, lista_entidades in entidades_detalladas.items():
                if lista_entidades:
                    # Tomar el primer ejemplo de cada tipo de entidad
                    primera_entidad = lista_entidades[0]
                    ejemplo_valor = None
                    confidence = 0.0
                    
                    # Obtener el valor de ejemplo
                    if 'money_value' in primera_entidad and primera_entidad['money_value']:
                        ejemplo_valor = str(primera_entidad['money_value'])
                    elif 'mention_text' in primera_entidad and primera_entidad['mention_text']:
                        ejemplo_valor = primera_entidad['mention_text']
                    
                    # Obtener confianza
                    if 'confidence' in primera_entidad and primera_entidad['confidence']:
                        confidence = float(primera_entidad['confidence'])
                    
                    if ejemplo_valor:
                        entidades_disponibles_respuesta[tipo_entidad] = {
                            'ejemplo_valor': ejemplo_valor,
                            'confidence': confidence
                        }
        
        # MEJORA: Añadir campos extraídos del texto como entidades disponibles si no fueron detectados por Document AI
        # Esto ayuda al usuario a ver qué campos se encontraron mediante búsqueda en texto
        if 'datos_extraidos' in locals() and datos_extraidos:
            campos_mapeo = {
                'total_con_impuestos': 'Total con Impuestos',
                'monto': 'Monto Total',
                'total': 'Base Imponible',
                'impuestos': 'IVA (€)',
                'iva': 'IVA',
                'iva_porcentaje': 'IVA (%)',
                'proveedor': 'Proveedor',
                'proveedor_nif': 'NIF/CIF Proveedor',
                'proveedor_telefono': 'Teléfono Proveedor',
                'proveedor_email': 'Email Proveedor',
                'proveedor_direccion': 'Dirección Proveedor',
                'numero_factura': 'Número de Factura',
                'fecha_pago': 'Fecha de Factura',
            }
            
            for campo_sistema, nombre_display in campos_mapeo.items():
                if campo_sistema in datos_extraidos:
                    valor = datos_extraidos[campo_sistema]
                    if valor:
                        # Crear una clave única para este campo
                        clave_entidad = f"extraido_{campo_sistema}"
                        # Solo añadir si no existe ya en las entidades detectadas por Document AI
                        if clave_entidad not in entidades_disponibles_respuesta:
                            entidades_disponibles_respuesta[clave_entidad] = {
                                'ejemplo_valor': str(valor),
                                'confidence': 85.0,  # Confianza media para campos extraídos del texto
                                'origen': 'búsqueda_en_texto'  # Indicar que viene de búsqueda en texto
                            }
        
        # Preparar respuesta con texto completo si está disponible
        respuesta = {
            'mensaje': 'Factura procesada exitosamente',
            'blob_path': blob_path,  # Ruta del PDF guardado en Cloud Storage (para asociar al registro)
            'pdf_url': pdf_url,  # URL firmada para visualizar el PDF
            'id_residencia_detectada': id_residencia_detectada,  # ID de residencia detectada automáticamente
            'datos_extraidos': datos_extraidos if datos_extraidos else {},
            'entidades_disponibles': entidades_disponibles_respuesta  # Entidades detectadas por la IA para mapeo
        }
        
        # Incluir texto completo si se extrajo (útil para debugging y pruebas)
        if 'texto_completo_respuesta' in locals() and texto_completo_respuesta:
            respuesta['texto_completo'] = texto_completo_respuesta[:10000]  # Limitar a 10000 caracteres para no sobrecargar la respuesta
        
        return jsonify(respuesta), 200
        
    except Exception as e:
        app.logger.error(f"Error al procesar factura: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error al procesar factura: {str(e)}'}), 500


@app.route('/api/v1/facturacion/proveedores/<int:id_pago>', methods=['PUT'])
@permiso_requerido('editar:pago_proveedor')
def actualizar_pago_proveedor(id_pago):
    """Actualiza un pago a proveedor. Acepta JSON o multipart/form-data (si hay archivo)."""
    try:
        # Detectar si hay archivo en la petición (multipart/form-data)
        archivo_factura = None
        factura_blob_path = None
        
        if 'factura' in request.files:
            archivo_factura = request.files['factura']
            if archivo_factura and archivo_factura.filename:
                # Validar que sea PDF
                if not archivo_factura.filename.lower().endswith('.pdf'):
                    return jsonify({'error': 'El archivo debe ser un PDF'}), 400
                
                # Obtener id_residencia del pago existente para subir el archivo
                conn_temp = get_db_connection()
                cursor_temp = conn_temp.cursor()
                try:
                    cursor_temp.execute("""
                        SELECT id_residencia FROM pago_proveedor WHERE id_pago = %s
                    """, (id_pago,))
                    pago_existente = cursor_temp.fetchone()
                    if not pago_existente:
                        return jsonify({'error': 'Pago no encontrado'}), 404
                    id_residencia = pago_existente[0]
                finally:
                    cursor_temp.close()
                    conn_temp.close()
                
                # Subir archivo a Cloud Storage
                from storage_manager import upload_document_unificado
                file_content = archivo_factura.read()
                nombre_archivo = archivo_factura.filename
                factura_blob_path = upload_document_unificado(
                    file_content, id_residencia, 'pago_proveedor', id_pago,
                    'Factura', nombre_archivo, 'application/pdf'
                )
                
                if not factura_blob_path:
                    return jsonify({'error': 'Error al subir el archivo a Cloud Storage'}), 500
        
        # Obtener datos (pueden venir de JSON o form)
        if request.is_json:
            data = request.get_json()
        else:
            # Si es multipart/form-data, construir data desde form
            data = {}
            if request.form.get('proveedor'):
                data['proveedor'] = request.form.get('proveedor')
            if request.form.get('concepto'):
                data['concepto'] = request.form.get('concepto')
            if request.form.get('monto'):
                data['monto'] = float(request.form.get('monto'))
            if request.form.get('id_residencia'):
                data['id_residencia'] = int(request.form.get('id_residencia'))
            if request.form.get('fecha_pago'):
                data['fecha_pago'] = request.form.get('fecha_pago')
            if request.form.get('fecha_prevista'):
                data['fecha_prevista'] = request.form.get('fecha_prevista')
            if request.form.get('metodo_pago'):
                data['metodo_pago'] = request.form.get('metodo_pago')
            if request.form.get('numero_factura'):
                data['numero_factura'] = request.form.get('numero_factura')
            if request.form.get('observaciones'):
                data['observaciones'] = request.form.get('observaciones')
            if request.form.get('estado'):
                data['estado'] = request.form.get('estado')
        
        # Si se subió un archivo, añadirlo a data
        if factura_blob_path:
            data['factura_blob_path'] = factura_blob_path
        
        if not data:
            return jsonify({'error': 'Datos requeridos'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el pago existe y pertenece a la residencia del usuario
            # Verificar que el pago existe y obtener su residencia
            cursor.execute("""
                SELECT id_pago, id_residencia FROM pago_proveedor
                WHERE id_pago = %s
            """, (id_pago,))
            
            pago_existente = cursor.fetchone()
            if not pago_existente:
                return jsonify({'error': 'Pago no encontrado'}), 404
            
            # Verificar acceso a la residencia del pago
            is_valid, error_response = validate_residencia_access(pago_existente[1])
            if not is_valid:
                return error_response
            
            # Si se está cambiando la residencia, validar acceso a la nueva residencia
            nueva_residencia = data.get('id_residencia')
            if nueva_residencia and nueva_residencia != pago_existente[1]:
                is_valid, error_response = validate_residencia_access(nueva_residencia)
            if not is_valid:
                return error_response
            
            # Preparar datos para actualizar
            updates = []
            valores = []
            
            if 'id_residencia' in data:
                updates.append('id_residencia = %s')
                valores.append(data['id_residencia'])
            
            if 'proveedor' in data:
                updates.append('proveedor = %s')
                valores.append(data['proveedor'])
            
            if 'concepto' in data:
                updates.append('concepto = %s')
                valores.append(data['concepto'])
            
            if 'monto' in data:
                updates.append('monto = %s')
                valores.append(data['monto'])
            
            if 'fecha_pago' in data:
                updates.append('fecha_pago = %s')
                valores.append(data['fecha_pago'] if data['fecha_pago'] else None)
            
            if 'fecha_prevista' in data:
                updates.append('fecha_prevista = %s')
                valores.append(data['fecha_prevista'] if data['fecha_prevista'] else None)
            
            if 'metodo_pago' in data:
                updates.append('metodo_pago = %s')
                valores.append(data['metodo_pago'])
            
            if 'numero_factura' in data:
                updates.append('numero_factura = %s')
                valores.append(data['numero_factura'])
            
            if 'observaciones' in data:
                updates.append('observaciones = %s')
                valores.append(data['observaciones'])
            
            # Verificar si existe la columna factura_blob_path y actualizarla si se proporciona
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'pago_proveedor' 
                  AND column_name = 'factura_blob_path'
            """)
            tiene_columna_factura = cursor.fetchone() is not None
            
            if tiene_columna_factura and 'factura_blob_path' in data:
                updates.append('factura_blob_path = %s')
                valores.append(data['factura_blob_path'])
            
            # Calcular estado automáticamente basado en las fechas
            if 'fecha_pago' in data or 'fecha_prevista' in data:
                fecha_pago = data.get('fecha_pago')
                fecha_prevista = data.get('fecha_prevista')
                
                if fecha_pago:
                    estado = 'pagado'
                elif fecha_prevista:
                    estado = 'pendiente'
                else:
                    estado = data.get('estado', 'pendiente')
                
                updates.append('estado = %s')
                valores.append(estado)
            elif 'estado' in data:
                updates.append('estado = %s')
                valores.append(data['estado'])
            
            if not updates:
                return jsonify({'error': 'No hay datos para actualizar'}), 400
            
            valores.append(id_pago)
            
            query = f"""
                UPDATE pago_proveedor
                SET {', '.join(updates)}
                WHERE id_pago = %s
                RETURNING id_pago
            """
            
            cursor.execute(query, valores)
            
            # Si se añadió una factura_blob_path, también guardar en documentación
            if tiene_columna_factura and 'factura_blob_path' in data and data.get('factura_blob_path'):
                factura_blob_path = data.get('factura_blob_path')
                try:
                    # Obtener datos del pago actualizado
                    cursor.execute("""
                        SELECT proveedor, id_residencia, concepto, numero_factura 
                        FROM pago_proveedor 
                        WHERE id_pago = %s
                    """, (id_pago,))
                    
                    pago_data = cursor.fetchone()
                    if pago_data:
                        proveedor_nombre = pago_data[0]
                        id_residencia_pago = pago_data[1]
                        concepto_pago = pago_data[2]
                        numero_factura_pago = pago_data[3]
                        
                        # Obtener id_proveedor del nombre del proveedor
                        cursor.execute("""
                            SELECT id_proveedor FROM proveedor 
                            WHERE nombre = %s AND id_residencia = %s AND activo = TRUE
                            LIMIT 1
                        """, (proveedor_nombre, id_residencia_pago))
                        
                        proveedor_result = cursor.fetchone()
                        if proveedor_result:
                            id_proveedor = proveedor_result[0]
                            
                            # Verificar si existe la tabla documento
                            cursor.execute("""
                                SELECT EXISTS (
                                    SELECT FROM information_schema.tables 
                                    WHERE table_schema = 'public' 
                                    AND table_name = 'documento'
                                )
                            """)
                            tabla_documento_existe = cursor.fetchone()[0]
                            
                            if tabla_documento_existe:
                                # Verificar si ya existe un documento para este pago
                                cursor.execute("""
                                    SELECT id_documento FROM documento
                                    WHERE tipo_entidad = 'proveedor' 
                                      AND id_entidad = %s
                                      AND url_archivo = %s
                                      AND activo = TRUE
                                    LIMIT 1
                                """, (id_proveedor, factura_blob_path))
                                
                                doc_existente = cursor.fetchone()
                                
                                if not doc_existente:
                                    # Obtener nombre del archivo del blob_path
                                    nombre_archivo = factura_blob_path.split('/')[-1] if '/' in factura_blob_path else f"Factura_{numero_factura_pago or id_pago}.pdf"
                                    
                                    # Crear documento en la tabla documento
                                    cursor.execute("""
                                        INSERT INTO documento (tipo_entidad, id_entidad, id_residencia, categoria_documento,
                                                              tipo_documento, nombre_archivo, descripcion, url_archivo,
                                                              tamaño_bytes, tipo_mime, id_usuario_subida, activo)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE)
                                        RETURNING id_documento
                                    """, (
                                        'proveedor',
                                        id_proveedor,
                                        id_residencia_pago,
                                        'fiscal',
                                        'Factura',
                                        nombre_archivo,
                                        f"Factura de pago: {proveedor_nombre} - {concepto_pago}",
                                        factura_blob_path,
                                        None,
                                        'application/pdf',
                                        g.id_usuario
                                    ))
                                    
                                    app.logger.info(f"Documento de factura creado para proveedor {id_proveedor}, pago {id_pago}")
                except Exception as doc_error:
                    # Si falla crear el documento, no fallar la actualización del pago
                    app.logger.warning(f"Error al crear documento de factura: {str(doc_error)}")
                    import traceback
                    app.logger.warning(traceback.format_exc())
            
            conn.commit()
            
            return jsonify({
                'mensaje': 'Pago a proveedor actualizado exitosamente',
                'id_pago': id_pago
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al actualizar pago a proveedor: {str(e)}")
            return jsonify({'error': 'Error al actualizar pago a proveedor'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============================================================================
# ENDPOINTS DE PROVEEDORES
# ============================================================================

@app.route('/api/v1/proveedores', methods=['GET'])
@permiso_requerido('leer:proveedor')
def listar_proveedores():
    """Lista los proveedores de las residencias del usuario."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Construir filtro de residencias
            where_clause, params = build_residencia_filter('', 'id_residencia')
            
            if where_clause:
                # Usuario normal: filtrar por residencias asignadas
                query = f"""
                    SELECT id_proveedor, id_residencia, nombre, nif_cif, direccion, telefono, email,
                           contacto, tipo_servicio, activo, observaciones, fecha_creacion
                    FROM proveedor
                    {where_clause}
                    ORDER BY nombre
                """
                cursor.execute(query, params)
            else:
                # Administrador: sin filtro (acceso total)
                cursor.execute("""
                    SELECT id_proveedor, id_residencia, nombre, nif_cif, direccion, telefono, email,
                       contacto, tipo_servicio, activo, observaciones, fecha_creacion
                FROM proveedor
                ORDER BY nombre
                """)
            
            proveedores = cursor.fetchall()
            
            resultado = []
            for prov in proveedores:
                resultado.append({
                    'id_proveedor': prov[0],
                    'id_residencia': prov[1],
                    'nombre': prov[2],
                    'nif_cif': prov[3],
                    'direccion': prov[4],
                    'telefono': prov[5],
                    'email': prov[6],
                    'contacto': prov[7],
                    'tipo_servicio': prov[8],
                    'activo': prov[9],
                    'observaciones': prov[10],
                    'fecha_creacion': prov[11].isoformat() if prov[11] else None
                })
            
            return jsonify({'proveedores': resultado, 'total': len(resultado)}), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al listar proveedores: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': 'Error al obtener proveedores', 'details': str(e)}), 500


@app.route('/api/v1/proveedores', methods=['POST'])
@permiso_requerido('crear:proveedor')
def crear_proveedor():
    """Crea un nuevo proveedor."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        nombre = data.get('nombre')
        valid, error = validate_text(nombre, 'Nombre del proveedor', min_length=2, max_length=255, required=True)
        if not valid:
            return jsonify({'error': error}), 400
        
        # Validar email si se proporciona
        if 'email' in data and data.get('email'):
            valid, error = validate_email(data.get('email'))
            if not valid:
                return jsonify({'error': error}), 400
        
        # Validar teléfono si se proporciona
        if 'telefono' in data and data.get('telefono'):
            valid, error = validate_phone(data.get('telefono'), 'Teléfono', required=False)
            if not valid:
                return jsonify({'error': error}), 400
        
        # Obtener id_residencia del request o usar la primera residencia asignada
        id_residencia = data.get('id_residencia')
        
        # Convertir a int si es string
        if id_residencia:
            try:
                id_residencia = int(id_residencia)
            except (ValueError, TypeError):
                return jsonify({'error': 'id_residencia debe ser un número válido'}), 400
        
        # Si no se proporciona id_residencia, usar la primera residencia asignada (si solo hay una)
        if not id_residencia:
            if g.id_rol == ADMIN_ROLE_ID:
                # Administrador debe especificar id_residencia explícitamente
                return jsonify({'error': 'id_residencia es requerido. Por favor, especifica la residencia (1 o 2)'}), 400
            elif not g.residencias_acceso or len(g.residencias_acceso) == 0:
                return jsonify({'error': 'Usuario sin residencias asignadas'}), 403
            elif len(g.residencias_acceso) == 1:
                id_residencia = g.residencias_acceso[0]
            else:
                return jsonify({'error': 'id_residencia es requerido cuando el usuario tiene acceso a múltiples residencias'}), 400
        
        # Validar que el usuario tiene acceso a la residencia especificada
        if g.id_rol != ADMIN_ROLE_ID:
            if id_residencia not in g.residencias_acceso:
                return jsonify({'error': 'No tienes permisos para crear proveedores en esta residencia'}), 403
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("""
                INSERT INTO proveedor (id_residencia, nombre, nif_cif, direccion, telefono,
                                     email, contacto, tipo_servicio, activo, observaciones)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id_proveedor
            """, (
                id_residencia,
                nombre,
                data.get('nif_cif'),
                data.get('direccion'),
                data.get('telefono'),
                data.get('email'),
                data.get('contacto'),
                data.get('tipo_servicio'),
                data.get('activo', True),
                data.get('observaciones')
            ))
            
            id_proveedor = cursor.fetchone()[0]
            conn.commit()
            
            return jsonify({
                'id_proveedor': id_proveedor,
                'mensaje': 'Proveedor creado exitosamente'
            }), 201
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al crear proveedor: {str(e)}")
            return jsonify({'error': 'Error al crear proveedor'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/proveedores/<int:id_proveedor>', methods=['GET'])
@permiso_requerido('leer:proveedor')
def obtener_proveedor(id_proveedor):
    """Obtiene un proveedor por su ID."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("""
                SELECT id_proveedor, id_residencia, nombre, nif_cif, direccion, telefono, email,
                       contacto, tipo_servicio, activo, observaciones, fecha_creacion
                FROM proveedor
                WHERE id_proveedor = %s
            """, (id_proveedor,))
            
            prov = cursor.fetchone()
            
            if not prov:
                app.logger.warning(f"Proveedor con ID {id_proveedor} no encontrado en la base de datos")
                return jsonify({'error': 'Proveedor no encontrado'}), 404
            
            # Verificar acceso a la residencia del proveedor
            is_valid, error_response = validate_residencia_access(prov[1])
            if not is_valid:
                app.logger.warning(f"Usuario no tiene acceso a la residencia {prov[1]} del proveedor {id_proveedor}")
                return error_response
            
            # Devolver resultado con id_residencia incluido
            resultado = {
                'id_proveedor': prov[0],
                'id_residencia': prov[1],
                'nombre': prov[2],
                'nif_cif': prov[3],
                'direccion': prov[4],
                'telefono': prov[5],
                'email': prov[6],
                'contacto': prov[7],
                'tipo_servicio': prov[8],
                'activo': prov[9],
                'observaciones': prov[10],
                'fecha_creacion': prov[11].isoformat() if prov[11] else None
            }
            app.logger.info(f"Proveedor {id_proveedor} obtenido exitosamente: {resultado['nombre']}")
            return jsonify(resultado), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al obtener proveedor {id_proveedor}: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error al obtener proveedor: {str(e)}'}), 500


@app.route('/api/v1/proveedores/<int:id_proveedor>/baja', methods=['POST'])
@permiso_requerido('editar:proveedor')
def dar_baja_proveedor(id_proveedor):
    """Da de baja a un proveedor con motivo y fecha."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        motivo_baja = data.get('motivo_baja')
        if not motivo_baja:
            return jsonify({'error': 'El motivo de baja es requerido'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el proveedor existe, está activo y pertenece a la residencia del usuario
            cursor.execute("""
                SELECT id_proveedor, activo, id_residencia FROM proveedor
                WHERE id_proveedor = %s
            """, (id_proveedor,))
            
            proveedor = cursor.fetchone()
            if not proveedor:
                return jsonify({'error': 'Proveedor no encontrado'}), 404
            
            # Verificar que el usuario tiene acceso a la residencia del proveedor
            is_valid, error_response = validate_residencia_access(proveedor[2])
            if not is_valid:
                return error_response
            
            if not proveedor[1]:  # Si ya está inactivo
                return jsonify({'error': 'El proveedor ya está dado de baja'}), 400
            
            # Verificar si las columnas de baja existen
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name = 'proveedor' 
                  AND column_name IN ('motivo_baja', 'fecha_baja')
            """)
            columnas_baja = {row[0] for row in cursor.fetchall()}
            
            # Actualizar el proveedor: activo = False, motivo_baja, fecha_baja = hoy
            from datetime import date
            fecha_baja = date.today()
            
            # Construir la consulta dinámicamente según las columnas que existan
            updates = ['activo = FALSE']
            valores = []
            
            if 'motivo_baja' in columnas_baja:
                updates.append('motivo_baja = %s')
                valores.append(motivo_baja)
            
            if 'fecha_baja' in columnas_baja:
                updates.append('fecha_baja = %s')
                valores.append(fecha_baja)
            
            valores.append(id_proveedor)
            
            query = f"""
                UPDATE proveedor
                SET {', '.join(updates)}
                WHERE id_proveedor = %s
                RETURNING id_proveedor
            """
            
            cursor.execute(query, tuple(valores))
            conn.commit()
            
            return jsonify({
                'mensaje': 'Proveedor dado de baja exitosamente',
                'id_proveedor': id_proveedor,
                'motivo_baja': motivo_baja if 'motivo_baja' in columnas_baja else None,
                'fecha_baja': str(fecha_baja) if 'fecha_baja' in columnas_baja else None
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al dar de baja al proveedor: {str(e)}")
            return jsonify({'error': 'Error al dar de baja al proveedor'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/proveedores/<int:id_proveedor>', methods=['PUT'])
@permiso_requerido('editar:proveedor')
def actualizar_proveedor(id_proveedor):
    """Actualiza un proveedor."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el proveedor existe y obtener su residencia
            cursor.execute("""
                SELECT id_proveedor, id_residencia FROM proveedor
                WHERE id_proveedor = %s
            """, (id_proveedor,))
            
            proveedor_existente = cursor.fetchone()
            if not proveedor_existente:
                return jsonify({'error': 'Proveedor no encontrado'}), 404
            
            # Verificar acceso a la residencia del proveedor
            is_valid, error_response = validate_residencia_access(proveedor_existente[1])
            if not is_valid:
                return error_response
            
            # Si se está cambiando la residencia, validar acceso a la nueva residencia
            nueva_residencia = data.get('id_residencia')
            if nueva_residencia and nueva_residencia != proveedor_existente[1]:
                is_valid, error_response = validate_residencia_access(nueva_residencia)
                if not is_valid:
                    return error_response
            
            # Campos actualizables (incluyendo id_residencia)
            campos_actualizables = [
                'id_residencia', 'nombre', 'nif_cif', 'direccion', 'telefono', 'email',
                'contacto', 'tipo_servicio', 'activo', 'observaciones'
            ]
            
            updates = []
            valores = []
            
            for campo in campos_actualizables:
                if campo in data:
                    updates.append(f"{campo} = %s")
                    valores.append(data[campo])
            
            if not updates:
                return jsonify({'error': 'No hay campos para actualizar'}), 400
            
            valores.append(id_proveedor)
            
            query = f"""
                UPDATE proveedor
                SET {', '.join(updates)}
                WHERE id_proveedor = %s
                RETURNING id_proveedor
            """
            
            cursor.execute(query, valores)
            conn.commit()
            
            return jsonify({'mensaje': 'Proveedor actualizado exitosamente'}), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al actualizar proveedor: {str(e)}")
            return jsonify({'error': 'Error al actualizar proveedor'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============================================================================
# ENDPOINTS DE PERSONAL
# ============================================================================

@app.route('/api/v1/personal', methods=['GET'])
def listar_personal():
    """Lista el personal de la residencia del usuario."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Construir query según acceso del usuario
            if g.id_rol == ADMIN_ROLE_ID:
                cursor.execute("""
                    SELECT id_personal, id_residencia, nombre, apellido, documento_identidad,
                       telefono, email, cargo, activo, fecha_contratacion, fecha_creacion
                FROM personal
                ORDER BY apellido, nombre
                """)
            else:
                if not g.residencias_acceso:
                    return jsonify({'personal': [], 'total': 0}), 200
                
                placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                cursor.execute(f"""
                    SELECT id_personal, id_residencia, nombre, apellido, documento_identidad,
                           telefono, email, cargo, activo, fecha_contratacion, fecha_creacion
                    FROM personal
                    WHERE id_residencia IN ({placeholders})
                    ORDER BY apellido, nombre
                """, tuple(g.residencias_acceso))
            
            personal_list = cursor.fetchall()
            
            resultado = []
            for p in personal_list:
                resultado.append({
                    'id_personal': p[0],
                    'id_residencia': p[1],
                    'nombre': p[2],
                    'apellido': p[3],
                    'documento_identidad': p[4],
                    'telefono': p[5],
                    'email': p[6],
                    'cargo': p[7],
                    'activo': p[8],
                    'fecha_contratacion': str(p[9]) if p[9] else None,
                    'fecha_creacion': p[10].isoformat() if p[10] else None
                })
            
            return jsonify({'personal': resultado, 'total': len(resultado)}), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al listar personal: {str(e)}")
        return jsonify({'error': 'Error al obtener personal'}), 500


@app.route('/api/v1/personal', methods=['POST'])
def crear_personal():
    """
    Crea un nuevo empleado/personal. Permite elegir la residencia (Violetas 1 o Violetas 2).
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        # Validar datos con el módulo de validación
        is_valid, errors = validate_personal_data(data, is_update=False)
        if not is_valid:
            return jsonify({'error': 'Errores de validación', 'detalles': errors}), 400
        
        nombre = data.get('nombre')
        apellido = data.get('apellido')
        id_residencia = data.get('id_residencia')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que la residencia existe
            cursor.execute("SELECT id_residencia FROM residencia WHERE id_residencia = %s", (id_residencia,))
            if not cursor.fetchone():
                return jsonify({'error': 'Residencia no encontrada'}), 404
            
            cursor.execute("""
                INSERT INTO personal (id_residencia, nombre, apellido, documento_identidad,
                                   telefono, email, cargo, activo, fecha_contratacion)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id_personal, fecha_creacion
            """, (
                id_residencia,
                nombre,
                apellido,
                data.get('documento_identidad'),
                data.get('telefono'),
                data.get('email'),
                data.get('cargo'),
                data.get('activo', True),
                data.get('fecha_contratacion')
            ))
            
            resultado = cursor.fetchone()
            id_personal = resultado[0]
            conn.commit()
            
            return jsonify({
                'id_personal': id_personal,
                'mensaje': 'Personal creado exitosamente'
            }), 201
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al crear personal: {str(e)}")
            return jsonify({'error': 'Error al crear personal'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/personal/<int:id_personal>', methods=['GET'])
def obtener_personal(id_personal):
    """Obtiene un empleado/personal específico por ID."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Construir query según acceso del usuario
            if g.id_rol == ADMIN_ROLE_ID:
                cursor.execute("""
                    SELECT id_personal, id_residencia, nombre, apellido, documento_identidad,
                           telefono, email, cargo, activo, fecha_contratacion, fecha_creacion
                    FROM personal
                    WHERE id_personal = %s
                """, (id_personal,))
            else:
                if not g.residencias_acceso:
                    return jsonify({'error': 'Personal no encontrado'}), 404
                
                placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                cursor.execute(f"""
                    SELECT id_personal, id_residencia, nombre, apellido, documento_identidad,
                           telefono, email, cargo, activo, fecha_contratacion, fecha_creacion
                    FROM personal
                    WHERE id_personal = %s AND id_residencia IN ({placeholders})
                """, (id_personal,) + tuple(g.residencias_acceso))
            
            personal = cursor.fetchone()
            
            if not personal:
                return jsonify({'error': 'Personal no encontrado'}), 404
            
            resultado = {
                'id_personal': personal[0],
                'id_residencia': personal[1],
                'nombre': personal[2],
                'apellido': personal[3],
                'documento_identidad': personal[4],
                'telefono': personal[5],
                'email': personal[6],
                'cargo': personal[7],
                'activo': personal[8],
                'fecha_contratacion': str(personal[9]) if personal[9] else None,
                'fecha_creacion': personal[10].isoformat() if personal[10] else None
            }
            
            return jsonify(resultado), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al obtener personal: {str(e)}")
        return jsonify({'error': 'Error al obtener personal'}), 500


@app.route('/api/v1/personal/<int:id_personal>', methods=['PUT'])
def actualizar_personal(id_personal):
    """Actualiza un empleado/personal existente."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        # Validar datos con el módulo de validación
        is_valid, errors = validate_personal_data(data, is_update=True)
        if not is_valid:
            return jsonify({'error': 'Errores de validación', 'detalles': errors}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el personal existe y el usuario tiene acceso
            if g.id_rol == ADMIN_ROLE_ID:
                cursor.execute("""
                    SELECT id_personal, id_residencia FROM personal WHERE id_personal = %s
                """, (id_personal,))
            else:
                if not g.residencias_acceso:
                    return jsonify({'error': 'Personal no encontrado'}), 404
                
                placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                cursor.execute(f"""
                    SELECT id_personal, id_residencia FROM personal 
                    WHERE id_personal = %s AND id_residencia IN ({placeholders})
                """, (id_personal,) + tuple(g.residencias_acceso))
            
            personal_existente = cursor.fetchone()
            
            if not personal_existente:
                return jsonify({'error': 'Personal no encontrado'}), 404
            
            personal_id_residencia = personal_existente[1]
            
            # Verificar que la residencia existe si se está cambiando
            id_residencia = data.get('id_residencia', personal_id_residencia)
            if id_residencia != personal_id_residencia:
                cursor.execute("SELECT id_residencia FROM residencia WHERE id_residencia = %s", (id_residencia,))
                if not cursor.fetchone():
                    return jsonify({'error': 'Residencia no encontrada'}), 404
            
            # Construir query de actualización
            updates = []
            valores = []
            
            if 'id_residencia' in data:
                updates.append('id_residencia = %s')
                valores.append(data['id_residencia'])
            
            if 'nombre' in data:
                updates.append('nombre = %s')
                valores.append(data['nombre'])
            
            if 'apellido' in data:
                updates.append('apellido = %s')
                valores.append(data['apellido'])
            
            if 'documento_identidad' in data:
                updates.append('documento_identidad = %s')
                valores.append(data['documento_identidad'] if data['documento_identidad'] else None)
            
            if 'telefono' in data:
                updates.append('telefono = %s')
                valores.append(data['telefono'] if data['telefono'] else None)
            
            if 'email' in data:
                updates.append('email = %s')
                valores.append(data['email'] if data['email'] else None)
            
            if 'cargo' in data:
                updates.append('cargo = %s')
                valores.append(data['cargo'] if data['cargo'] else None)
            
            if 'fecha_contratacion' in data:
                updates.append('fecha_contratacion = %s')
                valores.append(data['fecha_contratacion'] if data['fecha_contratacion'] else None)
            
            if 'activo' in data:
                updates.append('activo = %s')
                valores.append(data['activo'])
            
            if not updates:
                return jsonify({'error': 'No hay campos para actualizar'}), 400
            
            valores.append(id_personal)
            
            query = f"""
                UPDATE personal
                SET {', '.join(updates)}
                WHERE id_personal = %s
                RETURNING id_personal
            """
            
            cursor.execute(query, tuple(valores))
            conn.commit()
            
            return jsonify({
                'id_personal': id_personal,
                'mensaje': 'Personal actualizado exitosamente'
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al actualizar personal: {str(e)}")
            import traceback
            app.logger.error(traceback.format_exc())
            return jsonify({'error': 'Error al actualizar personal', 'details': str(e)}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============================================================================
# ENDPOINTS DE TURNOS EXTRA
# ============================================================================

@app.route('/api/v1/turnos-extra', methods=['GET'])
def listar_turnos_extra():
    """Lista los turnos extra de la residencia del usuario."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Obtener parámetros de filtro opcionales
            id_personal = request.args.get('id_personal', type=int)
            aprobado = request.args.get('aprobado', type=str)  # 'true', 'false', o None para todos
            
            # Construir query según acceso del usuario
            if g.id_rol == ADMIN_ROLE_ID:
                query = """
                    SELECT te.id_turno_extra, te.id_personal, te.id_residencia,
                           te.fecha, te.hora_entrada, te.hora_salida, te.motivo,
                           te.aprobado, te.fecha_creacion,
                           p.nombre || ' ' || p.apellido as nombre_personal,
                           p.cargo
                    FROM turno_extra te
                    JOIN personal p ON te.id_personal = p.id_personal
                    WHERE 1=1
                """
                params = []
            else:
                if not g.residencias_acceso:
                    return jsonify({'turnos_extra': [], 'total': 0}), 200
                
                placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                query = f"""
                    SELECT te.id_turno_extra, te.id_personal, te.id_residencia,
                           te.fecha, te.hora_entrada, te.hora_salida, te.motivo,
                           te.aprobado, te.fecha_creacion,
                           p.nombre || ' ' || p.apellido as nombre_personal,
                           p.cargo
                    FROM turno_extra te
                    JOIN personal p ON te.id_personal = p.id_personal
                    WHERE te.id_residencia IN ({placeholders})
                """
                params = list(g.residencias_acceso)
            
            if id_personal:
                query += " AND te.id_personal = %s"
                params.append(id_personal)
            
            if aprobado is not None:
                if aprobado.lower() == 'true':
                    query += " AND te.aprobado = TRUE"
                elif aprobado.lower() == 'false':
                    query += " AND te.aprobado = FALSE"
            
            query += " ORDER BY te.fecha DESC, te.hora_entrada DESC"
            
            cursor.execute(query, params)
            turnos = cursor.fetchall()
            
            resultado = []
            for t in turnos:
                resultado.append({
                    'id_turno_extra': t[0],
                    'id_personal': t[1],
                    'id_residencia': t[2],
                    'fecha': str(t[3]) if t[3] else None,
                    'hora_entrada': str(t[4]) if t[4] else None,
                    'hora_salida': str(t[5]) if t[5] else None,
                    'motivo': t[6],
                    'aprobado': t[7],
                    'fecha_creacion': t[8].isoformat() if t[8] else None,
                    'nombre_personal': t[9],
                    'cargo': t[10]
                })
            
            return jsonify({'turnos_extra': resultado, 'total': len(resultado)}), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al listar turnos extra: {str(e)}")
        return jsonify({'error': 'Error al obtener turnos extra'}), 500


@app.route('/api/v1/turnos-extra', methods=['POST'])
def crear_turno_extra():
    """Crea un nuevo turno extra."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        # Validar datos
        is_valid, errors = validate_turno_extra_data(data, is_update=False)
        if not is_valid:
            return jsonify({'error': 'Errores de validación', 'detalles': errors}), 400
        
        id_personal = data.get('id_personal')
        fecha = data.get('fecha')
        hora_entrada = data.get('hora_entrada')
        hora_salida = data.get('hora_salida')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el personal existe y obtener su residencia
            cursor.execute("""
                SELECT id_personal, id_residencia FROM personal
                WHERE id_personal = %s
            """, (id_personal,))
            
            personal = cursor.fetchone()
            if not personal:
                return jsonify({'error': 'Personal no encontrado'}), 404
            
            personal_id_residencia = personal[1]
            
            # Verificar acceso a la residencia del personal
            is_valid, error_response = validate_residencia_access(personal_id_residencia)
            if not is_valid:
                return error_response
            
            # Obtener id_residencia del request o usar la del personal
            id_residencia = data.get('id_residencia', personal_id_residencia)
            
            # Verificar que la residencia solicitada esté en la lista de acceso
            if g.id_rol != ADMIN_ROLE_ID and id_residencia not in g.residencias_acceso:
                return jsonify({'error': 'No tienes permisos para crear turnos en esta residencia'}), 403
            
            cursor.execute("""
                INSERT INTO turno_extra (id_personal, id_residencia, fecha, hora_entrada, hora_salida, motivo, aprobado)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING id_turno_extra, fecha_creacion
            """, (
                id_personal,
                id_residencia,
                fecha,
                hora_entrada,
                hora_salida,
                data.get('motivo'),
                data.get('aprobado', False)
            ))
            
            resultado = cursor.fetchone()
            id_turno_extra = resultado[0]
            conn.commit()
            
            return jsonify({
                'id_turno_extra': id_turno_extra,
                'mensaje': 'Turno extra creado exitosamente'
            }), 201
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al crear turno extra: {str(e)}")
            return jsonify({'error': 'Error al crear turno extra'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/turnos-extra/<int:id_turno_extra>', methods=['PUT'])
def actualizar_turno_extra(id_turno_extra):
    """Actualiza un turno extra (puede ser para aprobar/rechazar o editar)."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el turno existe y obtener su residencia
            cursor.execute("""
                SELECT id_turno_extra, id_residencia FROM turno_extra
                WHERE id_turno_extra = %s
            """, (id_turno_extra,))
            
            turno = cursor.fetchone()
            if not turno:
                return jsonify({'error': 'Turno extra no encontrado'}), 404
            
            # Verificar acceso a la residencia del turno
            is_valid, error_response = validate_residencia_access(turno[1])
            if not is_valid:
                return error_response
            
            # Construir la consulta de actualización dinámicamente
            updates = []
            params = []
            
            if 'fecha' in data:
                updates.append('fecha = %s')
                params.append(data.get('fecha'))
            
            if 'hora_entrada' in data:
                updates.append('hora_entrada = %s')
                params.append(data.get('hora_entrada'))
            
            if 'hora_salida' in data:
                updates.append('hora_salida = %s')
                params.append(data.get('hora_salida'))
            
            if 'motivo' in data:
                updates.append('motivo = %s')
                params.append(data.get('motivo'))
            
            if 'aprobado' in data:
                updates.append('aprobado = %s')
                params.append(data.get('aprobado'))
            
            if not updates:
                return jsonify({'error': 'No hay campos para actualizar'}), 400
            
            # Validar datos si se están actualizando campos críticos
            if 'fecha' in data or 'hora_entrada' in data or 'hora_salida' in data:
                is_valid, errors = validate_turno_extra_data(data, is_update=True)
                if not is_valid:
                    return jsonify({'error': 'Errores de validación', 'detalles': errors}), 400
            
            params.append(id_turno_extra)
            
            query = f"""
                UPDATE turno_extra
                SET {', '.join(updates)}
                WHERE id_turno_extra = %s
                RETURNING id_turno_extra
            """
            
            cursor.execute(query, params)
            
            if not cursor.fetchone():
                return jsonify({'error': 'Error al actualizar turno extra'}), 500
            
            conn.commit()
            
            return jsonify({
                'mensaje': 'Turno extra actualizado exitosamente'
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al actualizar turno extra: {str(e)}")
            return jsonify({'error': 'Error al actualizar turno extra'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/turnos-extra/<int:id_turno_extra>', methods=['DELETE'])
def eliminar_turno_extra(id_turno_extra):
    """Elimina un turno extra."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el turno existe y obtener su residencia
            cursor.execute("""
                SELECT id_turno_extra, id_residencia FROM turno_extra
                WHERE id_turno_extra = %s
            """, (id_turno_extra,))
            
            turno = cursor.fetchone()
            if not turno:
                return jsonify({'error': 'Turno extra no encontrado'}), 404
            
            # Verificar acceso a la residencia del turno
            is_valid, error_response = validate_residencia_access(turno[1])
            if not is_valid:
                return error_response
            
            cursor.execute("""
                DELETE FROM turno_extra
                WHERE id_turno_extra = %s
                RETURNING id_turno_extra
            """, (id_turno_extra,))
            
            conn.commit()
            
            return jsonify({
                'mensaje': 'Turno extra eliminado exitosamente'
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al eliminar turno extra: {str(e)}")
            return jsonify({'error': 'Error al eliminar turno extra'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============================================================================
# ENDPOINTS DE DOCUMENTACIÓN DE RESIDENTES
# ============================================================================

@app.route('/api/v1/residentes/<int:id_residente>/documentos', methods=['GET'])
def listar_documentos_residente(id_residente):
    """Lista los documentos de un residente. Permite ver documentos de residentes de cualquier residencia."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el residente existe (sin filtrar por residencia del usuario)
            cursor.execute("""
                SELECT id_residente, id_residencia FROM residente 
                WHERE id_residente = %s
            """, (id_residente,))
            
            residente = cursor.fetchone()
            if not residente:
                return jsonify({'error': 'Residente no encontrado'}), 404
            
            # Obtener la residencia real del residente
            id_residencia_residente = residente[1]
            
            cursor.execute("""
                SELECT id_documento, tipo_documento, nombre_archivo, descripcion,
                       fecha_subida, fecha_creacion, url_archivo, tamaño_bytes, tipo_mime
                FROM documento_residente
                WHERE id_residente = %s AND id_residencia = %s
                ORDER BY fecha_subida DESC
            """, (id_residente, id_residencia_residente))
            
            documentos = cursor.fetchall()
            
            resultado = []
            for doc in documentos:
                url_descarga = None
                if doc[6]:  # Si hay url_archivo
                    url_descarga = get_document_url(doc[6], expiration_minutes=60)
                
                resultado.append({
                    'id_documento': doc[0],
                    'tipo_documento': doc[1],
                    'nombre_archivo': doc[2],
                    'descripcion': doc[3],
                    'fecha_subida': doc[4].isoformat() if doc[4] else None,
                    'fecha_creacion': doc[5].isoformat() if doc[5] else None,
                    'url_archivo': doc[6],
                    'url_descarga': url_descarga,
                    'tamaño_bytes': doc[7],
                    'tipo_mime': doc[8]
                })
            
            return jsonify({'documentos': resultado, 'total': len(resultado)}), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al listar documentos: {str(e)}")
        return jsonify({'error': 'Error al obtener documentos'}), 500


@app.route('/api/v1/residentes/<int:id_residente>/documentos', methods=['POST'])
def crear_documento_residente(id_residente):
    """Crea un nuevo documento para un residente y lo sube a Cloud Storage."""
    try:
        # Verificar si hay archivo en la petición
        if 'archivo' not in request.files:
            # Si no hay archivo, crear solo el registro (modo compatible)
            data = request.get_json() if request.is_json else {}
            tipo_documento = data.get('tipo_documento') or request.form.get('tipo_documento')
            nombre_archivo = data.get('nombre_archivo') or request.form.get('nombre_archivo')
            descripcion = data.get('descripcion') or request.form.get('descripcion')
            
            if not tipo_documento or not nombre_archivo:
                return jsonify({'error': 'tipo_documento y nombre_archivo son requeridos'}), 400
            
            conn = get_db_connection()
            cursor = conn.cursor()
            
            try:
                # Verificar que el residente existe (sin filtrar por residencia del usuario)
                cursor.execute("""
                    SELECT id_residente, id_residencia FROM residente 
                    WHERE id_residente = %s
                """, (id_residente,))
                
                residente = cursor.fetchone()
                if not residente:
                    return jsonify({'error': 'Residente no encontrado'}), 404
                
                id_residencia = residente[1]
                
                cursor.execute("""
                    INSERT INTO documento_residente (id_residente, id_residencia, tipo_documento,
                                                    nombre_archivo, descripcion)
                    VALUES (%s, %s, %s, %s, %s)
                    RETURNING id_documento, fecha_subida
                """, (id_residente, id_residencia, tipo_documento, nombre_archivo, descripcion))
                
                resultado = cursor.fetchone()
                id_documento = resultado[0]
                conn.commit()
                
                return jsonify({
                    'id_documento': id_documento,
                    'mensaje': 'Documento creado exitosamente'
                }), 201
                
            except Exception as e:
                conn.rollback()
                app.logger.error(f"Error al crear documento: {str(e)}")
                return jsonify({'error': 'Error al crear documento'}), 500
            finally:
                cursor.close()
                conn.close()
        
        # Si hay archivo, subirlo a Cloud Storage
        archivo = request.files['archivo']
        tipo_documento = request.form.get('tipo_documento')
        descripcion = request.form.get('descripcion')
        
        if not tipo_documento:
            return jsonify({'error': 'tipo_documento es requerido'}), 400
        
        if archivo.filename == '':
            return jsonify({'error': 'No se seleccionó ningún archivo'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el residente existe (sin filtrar por residencia del usuario)
            cursor.execute("""
                SELECT id_residente, id_residencia FROM residente 
                WHERE id_residente = %s
            """, (id_residente,))
            
            residente = cursor.fetchone()
            if not residente:
                return jsonify({'error': 'Residente no encontrado'}), 404
            
            id_residencia = residente[1]
            
            # Leer contenido del archivo
            file_content = archivo.read()
            nombre_archivo = secure_filename(archivo.filename)
            tipo_mime = archivo.content_type or mimetypes.guess_type(nombre_archivo)[0] or 'application/octet-stream'
            tamaño_bytes = len(file_content)
            
            # Subir a Cloud Storage
            try:
                blob_path = upload_document(
                    file_content=file_content,
                    id_residencia=id_residencia,
                    id_residente=id_residente,
                    tipo_documento=tipo_documento,
                    nombre_archivo=nombre_archivo,
                    content_type=tipo_mime
                )
                
                if not blob_path:
                    app.logger.error("upload_document retornó None")
                    return jsonify({
                        'error': 'Error al subir el archivo a Cloud Storage. Verifique la configuración de GOOGLE_APPLICATION_CREDENTIALS y GCS_BUCKET_NAME'
                    }), 500
            except Exception as upload_error:
                app.logger.error(f"Error al subir documento a Cloud Storage: {str(upload_error)}")
                return jsonify({
                    'error': f'Error al subir el archivo: {str(upload_error)}'
                }), 500
            
            # Guardar en base de datos
            cursor.execute("""
                INSERT INTO documento_residente (id_residente, id_residencia, tipo_documento,
                                                nombre_archivo, descripcion, url_archivo,
                                                tamaño_bytes, tipo_mime)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id_documento, fecha_subida
            """, (
                id_residente,
                id_residencia,
                tipo_documento,
                nombre_archivo,
                descripcion,
                blob_path,
                tamaño_bytes,
                tipo_mime
            ))
            
            resultado = cursor.fetchone()
            id_documento = resultado[0]
            conn.commit()
            
            return jsonify({
                'id_documento': id_documento,
                'mensaje': 'Documento subido exitosamente',
                'url_archivo': blob_path
            }), 201
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al crear documento: {str(e)}", exc_info=True)
            return jsonify({'error': f'Error al crear documento: {str(e)}'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/documentos/<int:id_documento>', methods=['DELETE'])
def eliminar_documento(id_documento):
    """Elimina un documento de la base de datos y de Cloud Storage."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Obtener información del documento incluyendo url_archivo (sin filtrar por residencia del usuario)
            cursor.execute("""
                SELECT id_documento, url_archivo FROM documento_residente 
                WHERE id_documento = %s
            """, (id_documento,))
            
            documento = cursor.fetchone()
            if not documento:
                return jsonify({'error': 'Documento no encontrado'}), 404
            
            url_archivo = documento[1]
            
            # Eliminar de Cloud Storage si existe
            if url_archivo:
                try:
                    delete_document(url_archivo)
                except Exception as e:
                    app.logger.warning(f"No se pudo eliminar archivo de Cloud Storage: {str(e)}")
            
            # Eliminar de base de datos
            cursor.execute("""
                DELETE FROM documento_residente 
                WHERE id_documento = %s
            """, (id_documento,))
            
            conn.commit()
            
            return jsonify({'mensaje': 'Documento eliminado exitosamente'}), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al eliminar documento: {str(e)}")
            return jsonify({'error': 'Error al eliminar documento'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/documentos/<int:id_documento>/descargar', methods=['GET'])
def descargar_documento(id_documento):
    """Genera una URL firmada para descargar un documento."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el documento existe (sin filtrar por residencia del usuario)
            cursor.execute("""
                SELECT url_archivo FROM documento_residente 
                WHERE id_documento = %s
            """, (id_documento,))
            
            documento = cursor.fetchone()
            if not documento:
                return jsonify({'error': 'Documento no encontrado'}), 404
            
            url_archivo = documento[0]
            
            if not url_archivo:
                return jsonify({'error': 'El documento no tiene archivo asociado'}), 404
            
            # Generar URL firmada válida por 1 hora
            url_descarga = get_document_url(url_archivo, expiration_minutes=60)
            
            if not url_descarga:
                return jsonify({'error': 'Error al generar URL de descarga'}), 500
            
            return jsonify({'url_descarga': url_descarga}), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


# ============================================================================
# ENDPOINTS DE USUARIOS Y CONFIGURACIÓN
# ============================================================================

@app.route('/api/v1/roles', methods=['GET'])
def listar_roles():
    """Lista roles disponibles. Administrador ve todos, admin ve desde admin hacia abajo."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Asegurar que el rol Administrador existe (id_rol = 2)
            cursor.execute("""
                SELECT id_rol FROM rol WHERE id_rol = 2
            """)
            if not cursor.fetchone():
                # Crear rol Administrador si no existe
                try:
                    cursor.execute("""
                        INSERT INTO rol (id_rol, nombre, descripcion, activo)
                        VALUES (2, 'Administrador', 'Administrador con acceso a todos los módulos y permisos de todas las residencias', TRUE)
                    """)
                    conn.commit()
                except Exception as e:
                    # Si falla, intentar actualizar el rol existente
                    app.logger.warning(f"No se pudo crear rol Administrador: {str(e)}")
                    cursor.execute("""
                        UPDATE rol 
                        SET nombre = 'Administrador', 
                            descripcion = 'Administrador con acceso a todos los módulos y permisos de todas las residencias',
                            activo = TRUE
                        WHERE id_rol = 2
                    """)
                    conn.commit()
            
            # Filtrar roles según el rol del usuario actual
            # Mostrar: Administrador (id_rol=2), Director (id_rol=3), Personal (id_rol=4)
            filtro_rol = "AND id_rol IN (2, 3, 4)"
            
            cursor.execute(f"""
                SELECT id_rol, nombre, descripcion, activo
                FROM rol
                WHERE activo = TRUE {filtro_rol}
                ORDER BY id_rol
            """)
            
            roles = cursor.fetchall()
            
            return jsonify({
                'roles': [
                    {
                        'id_rol': r[0],
                        'nombre': r[1],
                        'descripcion': r[2],
                        'activo': r[3]
                    }
                    for r in roles
                ]
            }), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al listar roles: {str(e)}")
        return jsonify({'error': 'Error al obtener roles'}), 500


@app.route('/api/v1/modulos', methods=['GET'])
def listar_modulos():
    """Lista los módulos/permisos disponibles en el sistema."""
    modulos = [
        {
            'id_modulo': 'residentes',
            'nombre': 'Residentes',
            'permisos': [
                {'id': 'leer:residente', 'nombre': 'Leer', 'descripcion': 'Ver lista de residentes'},
                {'id': 'crear:residente', 'nombre': 'Crear', 'descripcion': 'Agregar nuevos residentes'},
                {'id': 'editar:residente', 'nombre': 'Editar', 'descripcion': 'Modificar residentes existentes'},
                {'id': 'eliminar:residente', 'nombre': 'Eliminar', 'descripcion': 'Dar de baja residentes'}
            ]
        },
        {
            'id_modulo': 'personal',
            'nombre': 'Personal',
            'permisos': [
                {'id': 'leer:personal', 'nombre': 'Leer', 'descripcion': 'Ver lista de personal'},
                {'id': 'crear:personal', 'nombre': 'Crear', 'descripcion': 'Agregar nuevo personal'},
                {'id': 'editar:personal', 'nombre': 'Editar', 'descripcion': 'Modificar personal existente'},
                {'id': 'eliminar:personal', 'nombre': 'Eliminar', 'descripcion': 'Dar de baja personal'}
            ]
        },
        {
            'id_modulo': 'proveedores',
            'nombre': 'Proveedores',
            'permisos': [
                {'id': 'leer:proveedor', 'nombre': 'Leer', 'descripcion': 'Ver lista de proveedores'},
                {'id': 'crear:proveedor', 'nombre': 'Crear', 'descripcion': 'Agregar nuevos proveedores'},
                {'id': 'editar:proveedor', 'nombre': 'Editar', 'descripcion': 'Modificar proveedores existentes'},
                {'id': 'eliminar:proveedor', 'nombre': 'Eliminar', 'descripcion': 'Eliminar proveedores'}
            ]
        },
        {
            'id_modulo': 'facturacion',
            'nombre': 'Facturación',
            'permisos': [
                {'id': 'leer:cobro', 'nombre': 'Leer Cobros', 'descripcion': 'Ver cobros de residentes'},
                {'id': 'crear:cobro', 'nombre': 'Crear Cobros', 'descripcion': 'Registrar nuevos cobros'},
                {'id': 'editar:cobro', 'nombre': 'Editar Cobros', 'descripcion': 'Modificar cobros existentes'},
                {'id': 'eliminar:cobro', 'nombre': 'Eliminar Cobros', 'descripcion': 'Eliminar cobros pendientes'},
                {'id': 'leer:pago_proveedor', 'nombre': 'Leer Pagos', 'descripcion': 'Ver pagos a proveedores'},
                {'id': 'crear:pago_proveedor', 'nombre': 'Crear Pagos', 'descripcion': 'Registrar nuevos pagos'},
                {'id': 'editar:pago_proveedor', 'nombre': 'Editar Pagos', 'descripcion': 'Modificar pagos existentes'},
                {'id': 'eliminar:pago_proveedor', 'nombre': 'Eliminar Pagos', 'descripcion': 'Eliminar pagos a proveedores'}
            ]
        },
        {
            'id_modulo': 'receiver',
            'nombre': 'Entidades Fiscales',
            'permisos': [
                {'id': 'leer:receiver', 'nombre': 'Leer Entidades', 'descripcion': 'Ver entidades fiscales'},
                {'id': 'escribir:receiver', 'nombre': 'Gestionar Entidades', 'descripcion': 'Crear y modificar entidades fiscales'}
            ]
        },
        {
            'id_modulo': 'documentacion',
            'nombre': 'Documentación',
            'permisos': [
                {'id': 'leer:documento', 'nombre': 'Leer', 'descripcion': 'Ver documentos'},
                {'id': 'crear:documento', 'nombre': 'Subir', 'descripcion': 'Subir nuevos documentos'},
                {'id': 'eliminar:documento', 'nombre': 'Eliminar', 'descripcion': 'Eliminar documentos'}
            ]
        },
        {
            'id_modulo': 'turnos',
            'nombre': 'Turnos',
            'permisos': [
                {'id': 'leer:turno', 'nombre': 'Leer', 'descripcion': 'Ver turnos del personal'},
                {'id': 'crear:turno', 'nombre': 'Crear', 'descripcion': 'Registrar nuevos turnos'},
                {'id': 'editar:turno', 'nombre': 'Editar', 'descripcion': 'Modificar turnos existentes'}
            ]
        },
        {
            'id_modulo': 'historicos',
            'nombre': 'Históricos',
            'permisos': [
                {'id': 'leer:registro_asistencial', 'nombre': 'Leer Registros Asistenciales', 'descripcion': 'Ver registros asistenciales históricos'},
                {'id': 'crear:registro_asistencial', 'nombre': 'Crear Registros Asistenciales', 'descripcion': 'Registrar eventos asistenciales históricos'},
                {'id': 'editar:registro_asistencial', 'nombre': 'Editar Registros Asistenciales', 'descripcion': 'Modificar registros asistenciales históricos'},
                {'id': 'eliminar:registro_asistencial', 'nombre': 'Eliminar Registros', 'descripcion': 'Eliminar registros asistenciales'}
            ]
        },
        {
            'id_modulo': 'configuracion',
            'nombre': 'Configuración',
            'permisos': [
                {'id': 'leer:usuario', 'nombre': 'Leer Usuarios', 'descripcion': 'Ver lista de usuarios'},
                {'id': 'crear:usuario', 'nombre': 'Crear Usuarios', 'descripcion': 'Crear nuevos usuarios'},
                {'id': 'editar:usuario', 'nombre': 'Editar Usuarios', 'descripcion': 'Modificar usuarios existentes'},
                {'id': 'eliminar:usuario', 'nombre': 'Eliminar Usuarios', 'descripcion': 'Eliminar usuarios'},
                {'id': 'leer:residencia', 'nombre': 'Leer Residencias', 'descripcion': 'Ver información de residencias'},
                {'id': 'editar:residencia', 'nombre': 'Editar Residencias', 'descripcion': 'Modificar información de residencias'}
            ]
        }
    ]
    
    return jsonify({'modulos': modulos}), 200


@app.route('/api/v1/residencias', methods=['GET'])
def listar_residencias():
    """Lista las residencias a las que el usuario tiene acceso."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Si es Administrador, mostrar todas las residencias activas (para asignación de usuarios)
            # Si no es Administrador, solo mostrar las residencias activas a las que tiene acceso
            # IMPORTANTE: Solo se muestran residencias activas para evitar asignar usuarios a residencias desactivadas
            if g.id_rol == ADMIN_ROLE_ID:
                cursor.execute("""
                    SELECT id_residencia, nombre, direccion, telefono, activa, fecha_creacion
                    FROM residencia
                    WHERE activa = TRUE
                    ORDER BY id_residencia
                """)
            else:
                # Usuario normal: solo mostrar residencias a las que tiene acceso
                if not hasattr(g, 'residencias_acceso') or not g.residencias_acceso:
                    return jsonify({'residencias': [], 'total': 0}), 200
                
                placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                cursor.execute(f"""
                    SELECT id_residencia, nombre, direccion, telefono, activa, fecha_creacion
                    FROM residencia
                    WHERE id_residencia IN ({placeholders})
                      AND activa = TRUE
                    ORDER BY id_residencia
                """, tuple(g.residencias_acceso))
            
            residencias = cursor.fetchall()
            
            return jsonify({
                'residencias': [
                    {
                        'id_residencia': r[0],
                        'nombre': r[1],
                        'direccion': r[2],
                        'telefono': r[3],
                        'activa': r[4],
                        'fecha_creacion': r[5].isoformat() if r[5] else None
                    }
                    for r in residencias
                ],
                'total': len(residencias)
            }), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al listar residencias: {str(e)}")
        return jsonify({'error': 'Error al obtener residencias'}), 500


@app.route('/api/v1/residencias', methods=['POST'])
@permiso_requerido('escribir:residencia')
def crear_residencia():
    """Crea una nueva residencia. Solo Administrador."""
    # Verificar que es Administrador
    if g.id_rol != ADMIN_ROLE_ID:
        return jsonify({'error': 'Solo el Administradoristrador puede crear residencias'}), 403
    
    try:
        data = request.get_json()
        
        # Validar datos requeridos
        if not data or 'nombre' not in data:
            return jsonify({'error': 'El nombre es requerido'}), 400
        
        nombre = data.get('nombre', '').strip()
        nombre_fiscal = data.get('nombre_fiscal', '').strip() if data.get('nombre_fiscal') else None
        nif = data.get('nif', '').strip() if data.get('nif') else None
        direccion = data.get('direccion', '').strip() if data.get('direccion') else None
        codigo_postal = data.get('codigo_postal', '').strip() if data.get('codigo_postal') else None
        ciudad = data.get('ciudad', '').strip() if data.get('ciudad') else None
        provincia = data.get('provincia', '').strip() if data.get('provincia') else None
        telefono = data.get('telefono', '').strip() if data.get('telefono') else None
        email = data.get('email', '').strip() if data.get('email') else None
        web = data.get('web', '').strip() if data.get('web') else None
        cuenta_bancaria = data.get('cuenta_bancaria', '').strip() if data.get('cuenta_bancaria') else None
        observaciones = data.get('observaciones', '').strip() if data.get('observaciones') else None
        activa = data.get('activa', True)
        
        # Validar nombre
        if not nombre or len(nombre) < 2:
            return jsonify({'error': 'El nombre debe tener al menos 2 caracteres'}), 400
        
        if len(nombre) > 255:
            return jsonify({'error': 'El nombre es demasiado largo (máximo 255 caracteres)'}), 400
        
        # Validar campos opcionales
        if nombre_fiscal and len(nombre_fiscal) > 255:
            return jsonify({'error': 'El nombre fiscal es demasiado largo (máximo 255 caracteres)'}), 400
        
        if nif and len(nif) > 20:
            return jsonify({'error': 'El NIF es demasiado largo (máximo 20 caracteres)'}), 400
        
        if telefono and len(telefono) > 50:
            return jsonify({'error': 'El teléfono es demasiado largo (máximo 50 caracteres)'}), 400
        
        if direccion and len(direccion) > 500:
            return jsonify({'error': 'La dirección es demasiado larga (máximo 500 caracteres)'}), 400
        
        if codigo_postal and len(codigo_postal) > 10:
            return jsonify({'error': 'El código postal es demasiado largo (máximo 10 caracteres)'}), 400
        
        if ciudad and len(ciudad) > 100:
            return jsonify({'error': 'La ciudad es demasiado larga (máximo 100 caracteres)'}), 400
        
        if provincia and len(provincia) > 100:
            return jsonify({'error': 'La provincia es demasiado larga (máximo 100 caracteres)'}), 400
        
        if email and len(email) > 255:
            return jsonify({'error': 'El email es demasiado largo (máximo 255 caracteres)'}), 400
        
        if web and len(web) > 255:
            return jsonify({'error': 'La web es demasiado larga (máximo 255 caracteres)'}), 400
        
        if cuenta_bancaria and len(cuenta_bancaria) > 34:
            return jsonify({'error': 'La cuenta bancaria es demasiado larga (máximo 34 caracteres)'}), 400
        
        if observaciones and len(observaciones) > 1000:
            return jsonify({'error': 'Las observaciones son demasiado largas (máximo 1000 caracteres)'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que no exista una residencia con el mismo nombre
            cursor.execute("""
                SELECT id_residencia FROM residencia WHERE LOWER(nombre) = LOWER(%s)
            """, (nombre,))
            
            if cursor.fetchone():
                return jsonify({'error': 'Ya existe una residencia con ese nombre'}), 400
            
            # Crear la residencia
            cursor.execute("""
                INSERT INTO residencia (nombre, nombre_fiscal, nif, direccion, codigo_postal, ciudad, provincia, 
                                       telefono, email, web, cuenta_bancaria, observaciones, activa)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id_residencia, nombre, nombre_fiscal, nif, direccion, codigo_postal, ciudad, provincia, 
                          telefono, email, web, cuenta_bancaria, observaciones, activa, fecha_creacion
            """, (nombre, nombre_fiscal, nif, direccion, codigo_postal, ciudad, provincia, 
                  telefono, email, web, cuenta_bancaria, observaciones, activa))
            
            residencia = cursor.fetchone()
            conn.commit()
            
            return jsonify({
                'mensaje': 'Residencia creada exitosamente',
                'residencia': {
                    'id_residencia': residencia[0],
                    'nombre': residencia[1],
                    'nombre_fiscal': residencia[2],
                    'nif': residencia[3],
                    'direccion': residencia[4],
                    'codigo_postal': residencia[5],
                    'ciudad': residencia[6],
                    'provincia': residencia[7],
                    'telefono': residencia[8],
                    'email': residencia[9],
                    'web': residencia[10],
                    'cuenta_bancaria': residencia[11],
                    'observaciones': residencia[12],
                    'activa': residencia[13],
                    'fecha_creacion': residencia[14].isoformat() if residencia[14] else None
                }
            }), 201
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al crear residencia: {str(e)}")
            return jsonify({'error': 'Error al crear la residencia'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al crear residencia: {str(e)}")
        return jsonify({'error': 'Error al procesar la solicitud'}), 500


@app.route('/api/v1/residencias/<int:id_residencia>', methods=['PUT'])
@permiso_requerido('escribir:residencia')
def actualizar_residencia(id_residencia):
    """Actualiza una residencia existente. Solo Administrador."""
    # Verificar que es Administrador
    if g.id_rol != ADMIN_ROLE_ID:
        return jsonify({'error': 'Solo el Administradoristrador puede actualizar residencias'}), 403
    
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No se proporcionaron datos'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que la residencia existe
            cursor.execute("""
                SELECT id_residencia, nombre FROM residencia WHERE id_residencia = %s
            """, (id_residencia,))
            
            residencia_existente = cursor.fetchone()
            if not residencia_existente:
                return jsonify({'error': 'Residencia no encontrada'}), 404
            
            # Preparar campos a actualizar
            updates = []
            params = []
            
            if 'nombre' in data:
                nombre = data.get('nombre', '').strip()
                if not nombre or len(nombre) < 2:
                    return jsonify({'error': 'El nombre debe tener al menos 2 caracteres'}), 400
                if len(nombre) > 255:
                    return jsonify({'error': 'El nombre es demasiado largo (máximo 255 caracteres)'}), 400
                
                # Verificar que no exista otra residencia con el mismo nombre
                cursor.execute("""
                    SELECT id_residencia FROM residencia 
                    WHERE LOWER(nombre) = LOWER(%s) AND id_residencia != %s
                """, (nombre, id_residencia))
                
                if cursor.fetchone():
                    return jsonify({'error': 'Ya existe otra residencia con ese nombre'}), 400
                
                updates.append("nombre = %s")
                params.append(nombre)
            
            if 'direccion' in data:
                direccion = data.get('direccion', '').strip() if data.get('direccion') else None
                if direccion and len(direccion) > 500:
                    return jsonify({'error': 'La dirección es demasiado larga (máximo 500 caracteres)'}), 400
                updates.append("direccion = %s")
                params.append(direccion)
            
            if 'telefono' in data:
                telefono = data.get('telefono', '').strip() if data.get('telefono') else None
                if telefono and len(telefono) > 50:
                    return jsonify({'error': 'El teléfono es demasiado largo (máximo 50 caracteres)'}), 400
                updates.append("telefono = %s")
                params.append(telefono)
            
            if 'nombre_fiscal' in data:
                nombre_fiscal = data.get('nombre_fiscal', '').strip() if data.get('nombre_fiscal') else None
                if nombre_fiscal and len(nombre_fiscal) > 255:
                    return jsonify({'error': 'El nombre fiscal es demasiado largo (máximo 255 caracteres)'}), 400
                updates.append("nombre_fiscal = %s")
                params.append(nombre_fiscal)
            
            if 'nif' in data:
                nif = data.get('nif', '').strip() if data.get('nif') else None
                if nif and len(nif) > 20:
                    return jsonify({'error': 'El NIF es demasiado largo (máximo 20 caracteres)'}), 400
                updates.append("nif = %s")
                params.append(nif)
            
            if 'codigo_postal' in data:
                codigo_postal = data.get('codigo_postal', '').strip() if data.get('codigo_postal') else None
                if codigo_postal and len(codigo_postal) > 10:
                    return jsonify({'error': 'El código postal es demasiado largo (máximo 10 caracteres)'}), 400
                updates.append("codigo_postal = %s")
                params.append(codigo_postal)
            
            if 'ciudad' in data:
                ciudad = data.get('ciudad', '').strip() if data.get('ciudad') else None
                if ciudad and len(ciudad) > 100:
                    return jsonify({'error': 'La ciudad es demasiado larga (máximo 100 caracteres)'}), 400
                updates.append("ciudad = %s")
                params.append(ciudad)
            
            if 'provincia' in data:
                provincia = data.get('provincia', '').strip() if data.get('provincia') else None
                if provincia and len(provincia) > 100:
                    return jsonify({'error': 'La provincia es demasiado larga (máximo 100 caracteres)'}), 400
                updates.append("provincia = %s")
                params.append(provincia)
            
            if 'email' in data:
                email = data.get('email', '').strip() if data.get('email') else None
                if email and len(email) > 255:
                    return jsonify({'error': 'El email es demasiado largo (máximo 255 caracteres)'}), 400
                updates.append("email = %s")
                params.append(email)
            
            if 'web' in data:
                web = data.get('web', '').strip() if data.get('web') else None
                if web and len(web) > 255:
                    return jsonify({'error': 'La web es demasiado larga (máximo 255 caracteres)'}), 400
                updates.append("web = %s")
                params.append(web)
            
            if 'cuenta_bancaria' in data:
                cuenta_bancaria = data.get('cuenta_bancaria', '').strip() if data.get('cuenta_bancaria') else None
                if cuenta_bancaria and len(cuenta_bancaria) > 34:
                    return jsonify({'error': 'La cuenta bancaria es demasiado larga (máximo 34 caracteres)'}), 400
                updates.append("cuenta_bancaria = %s")
                params.append(cuenta_bancaria)
            
            if 'observaciones' in data:
                observaciones = data.get('observaciones', '').strip() if data.get('observaciones') else None
                if observaciones and len(observaciones) > 1000:
                    return jsonify({'error': 'Las observaciones son demasiado largas (máximo 1000 caracteres)'}), 400
                updates.append("observaciones = %s")
                params.append(observaciones)
            
            if 'activa' in data:
                activa = bool(data.get('activa'))
                updates.append("activa = %s")
                params.append(activa)
            
            if not updates:
                return jsonify({'error': 'No se proporcionaron campos para actualizar'}), 400
            
            # Ejecutar actualización
            params.append(id_residencia)
            query = f"""
                UPDATE residencia 
                SET {', '.join(updates)}
                WHERE id_residencia = %s
                RETURNING id_residencia, nombre, nombre_fiscal, nif, direccion, codigo_postal, ciudad, provincia, 
                         telefono, email, web, cuenta_bancaria, observaciones, activa, fecha_creacion
            """
            
            cursor.execute(query, params)
            residencia = cursor.fetchone()
            conn.commit()
            
            return jsonify({
                'mensaje': 'Residencia actualizada exitosamente',
                'residencia': {
                    'id_residencia': residencia[0],
                    'nombre': residencia[1],
                    'nombre_fiscal': residencia[2],
                    'nif': residencia[3],
                    'direccion': residencia[4],
                    'codigo_postal': residencia[5],
                    'ciudad': residencia[6],
                    'provincia': residencia[7],
                    'telefono': residencia[8],
                    'email': residencia[9],
                    'web': residencia[10],
                    'cuenta_bancaria': residencia[11],
                    'observaciones': residencia[12],
                    'activa': residencia[13],
                    'fecha_creacion': residencia[14].isoformat() if residencia[14] else None
                }
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al actualizar residencia: {str(e)}")
            return jsonify({'error': 'Error al actualizar la residencia'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al actualizar residencia: {str(e)}")
        return jsonify({'error': 'Error al procesar la solicitud'}), 500


# ==================== ENDPOINTS DE RECEIVERS (ENTIDADES FISCALES) ====================

@app.route('/api/v1/receivers', methods=['GET'])
@permiso_requerido('leer:receiver')
def listar_receivers():
    """Lista todas las entidades fiscales (receivers/sociedades)."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("""
                SELECT id_receiver, nombre, nif_cif, direccion, telefono, email, cuenta_bancaria, activo, fecha_creacion
                FROM receiver
                ORDER BY nombre
            """)
            
            receivers = cursor.fetchall()
            
            resultado = []
            for r in receivers:
                resultado.append({
                    'id_receiver': r[0],
                    'nombre': r[1],
                    'nif_cif': r[2],
                    'direccion': r[3],
                    'telefono': r[4],
                    'email': r[5],
                    'cuenta_bancaria': r[6],
                    'activo': r[7],
                    'fecha_creacion': r[8].isoformat() if r[8] else None
                })
            
            return jsonify({'receivers': resultado, 'total': len(resultado)}), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al listar receivers: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error al obtener entidades fiscales: {str(e)}'}), 500


@app.route('/api/v1/receivers', methods=['POST'])
@permiso_requerido('escribir:receiver')
def crear_receiver():
    """Crea una nueva entidad fiscal (receiver/sociedad)."""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        nombre = data.get('nombre')
        if not nombre:
            return jsonify({'error': 'nombre es requerido'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("""
                INSERT INTO receiver (nombre, nif_cif, direccion, telefono, email, cuenta_bancaria, activo)
                VALUES (%s, %s, %s, %s, %s, %s, TRUE)
                RETURNING id_receiver
            """, (
                nombre,
                data.get('nif_cif'),
                data.get('direccion'),
                data.get('telefono'),
                data.get('email'),
                data.get('cuenta_bancaria')
            ))
            
            id_receiver = cursor.fetchone()[0]
            conn.commit()
            
            return jsonify({
                'id_receiver': id_receiver,
                'mensaje': 'Entidad fiscal creada exitosamente'
            }), 201
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al crear receiver: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error al crear entidad fiscal: {str(e)}'}), 500


@app.route('/api/v1/residencias/<int:id_residencia>/receivers', methods=['GET'])
@permiso_requerido('leer:residencia')
def listar_receivers_residencia(id_residencia):
    """Lista las entidades fiscales asociadas a una residencia."""
    try:
        # Verificar acceso a la residencia
        is_valid, error_response = validate_residencia_access(id_residencia)
        if not is_valid:
            return error_response
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute("""
                SELECT r.id_receiver, r.nombre, r.nif_cif, r.direccion, r.telefono, r.email, r.cuenta_bancaria, r.activo
                FROM receiver r
                INNER JOIN residencia_receiver rr ON r.id_receiver = rr.id_receiver
                WHERE rr.id_residencia = %s AND rr.activo = TRUE AND r.activo = TRUE
                ORDER BY r.nombre
            """, (id_residencia,))
            
            receivers = cursor.fetchall()
            
            resultado = []
            for r in receivers:
                resultado.append({
                    'id_receiver': r[0],
                    'nombre': r[1],
                    'nif_cif': r[2],
                    'direccion': r[3],
                    'telefono': r[4],
                    'email': r[5],
                    'cuenta_bancaria': r[6],
                    'activo': r[7]
                })
            
            return jsonify({'receivers': resultado, 'total': len(resultado)}), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al listar receivers de residencia: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error al obtener entidades fiscales: {str(e)}'}), 500


@app.route('/api/v1/residencias/<int:id_residencia>/receivers/<int:id_receiver>', methods=['POST'])
@permiso_requerido('escribir:residencia')
def asociar_receiver_residencia(id_residencia, id_receiver):
    """Asocia una entidad fiscal a una residencia."""
    try:
        # Verificar acceso a la residencia
        is_valid, error_response = validate_residencia_access(id_residencia)
        if not is_valid:
            return error_response
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el receiver existe y está activo
            cursor.execute("SELECT activo FROM receiver WHERE id_receiver = %s", (id_receiver,))
            receiver = cursor.fetchone()
            
            if not receiver:
                return jsonify({'error': 'Entidad fiscal no encontrada'}), 404
            
            if not receiver[0]:
                return jsonify({'error': 'La entidad fiscal está inactiva'}), 400
            
            # Verificar que la residencia existe
            cursor.execute("SELECT activa FROM residencia WHERE id_residencia = %s", (id_residencia,))
            residencia = cursor.fetchone()
            
            if not residencia:
                return jsonify({'error': 'Residencia no encontrada'}), 404
            
            # Asociar (o reactivar si ya existe)
            cursor.execute("""
                INSERT INTO residencia_receiver (id_residencia, id_receiver, activo)
                VALUES (%s, %s, TRUE)
                ON CONFLICT (id_residencia, id_receiver) 
                DO UPDATE SET activo = TRUE, fecha_asignacion = CURRENT_TIMESTAMP
            """, (id_residencia, id_receiver))
            
            conn.commit()
            
            return jsonify({'mensaje': 'Entidad fiscal asociada exitosamente'}), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al asociar receiver a residencia: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error al asociar entidad fiscal: {str(e)}'}), 500


@app.route('/api/v1/residencias/<int:id_residencia>/receivers/<int:id_receiver>', methods=['DELETE'])
@permiso_requerido('escribir:residencia')
def desasociar_receiver_residencia(id_residencia, id_receiver):
    """Desasocia una entidad fiscal de una residencia."""
    try:
        # Verificar acceso a la residencia
        is_valid, error_response = validate_residencia_access(id_residencia)
        if not is_valid:
            return error_response
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Desactivar la relación (no eliminar físicamente)
            cursor.execute("""
                UPDATE residencia_receiver 
                SET activo = FALSE 
                WHERE id_residencia = %s AND id_receiver = %s
            """, (id_residencia, id_receiver))
            
            if cursor.rowcount == 0:
                return jsonify({'error': 'La entidad fiscal no está asociada a esta residencia'}), 404
            
            conn.commit()
            
            return jsonify({'mensaje': 'Entidad fiscal desasociada exitosamente'}), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al desasociar receiver de residencia: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error al desasociar entidad fiscal: {str(e)}'}), 500


@app.route('/api/v1/residencias/<int:id_residencia>', methods=['DELETE'])
@permiso_requerido('escribir:residencia')
def eliminar_residencia(id_residencia):
    """Elimina (desactiva) una residencia. Solo Administrador."""
    # Verificar que es Administrador
    if g.id_rol != ADMIN_ROLE_ID:
        return jsonify({'error': 'Solo el Administradoristrador puede eliminar residencias'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que la residencia existe
            cursor.execute("""
                SELECT id_residencia, nombre FROM residencia WHERE id_residencia = %s
            """, (id_residencia,))
            
            residencia = cursor.fetchone()
            if not residencia:
                return jsonify({'error': 'Residencia no encontrada'}), 404
            
            # Verificar si hay residentes activos en esta residencia
            cursor.execute("""
                SELECT COUNT(*) FROM residente 
                WHERE id_residencia = %s AND activo = TRUE
            """, (id_residencia,))
            
            residentes_activos = cursor.fetchone()[0]
            
            # Verificar si hay usuarios asignados a esta residencia
            cursor.execute("""
                SELECT COUNT(*) FROM usuario_residencia 
                WHERE id_residencia = %s
            """, (id_residencia,))
            
            usuarios_asignados = cursor.fetchone()[0]
            
            # En lugar de eliminar físicamente, desactivamos la residencia
            cursor.execute("""
                UPDATE residencia 
                SET activa = FALSE
                WHERE id_residencia = %s
                RETURNING id_residencia, nombre, activa
            """, (id_residencia,))
            
            residencia_actualizada = cursor.fetchone()
            conn.commit()
            
            return jsonify({
                'mensaje': 'Residencia desactivada exitosamente',
                'residencia': {
                    'id_residencia': residencia_actualizada[0],
                    'nombre': residencia_actualizada[1],
                    'activa': residencia_actualizada[2]
                },
                'advertencias': {
                    'residentes_activos': residentes_activos,
                    'usuarios_asignados': usuarios_asignados
                }
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al eliminar residencia: {str(e)}")
            return jsonify({'error': 'Error al eliminar la residencia'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al eliminar residencia: {str(e)}")
        return jsonify({'error': 'Error al procesar la solicitud'}), 500


@app.route('/api/v1/usuarios/me', methods=['GET'])
def obtener_usuario_actual():
    """Obtiene la información del usuario actual."""
    try:
        # Verificar que g.id_usuario existe (debería estar establecido por before_request)
        if not hasattr(g, 'id_usuario') or not g.id_usuario:
            app.logger.error("No se encontró id_usuario en g. Verificar autenticación.")
            return jsonify({'error': 'Usuario no autenticado'}), 401
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Consulta que maneja tanto usuarios con id_residencia como Administrador sin id_residencia
            # Verificar primero si las columnas nombre y apellido existen
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_schema = 'public' 
                AND table_name = 'usuario'
                AND column_name IN ('nombre', 'apellido')
            """)
            columnas_existentes = [row[0] for row in cursor.fetchall()]
            tiene_nombre = 'nombre' in columnas_existentes
            tiene_apellido = 'apellido' in columnas_existentes
            
            # Construir SELECT dinámicamente según columnas disponibles
            campos_select = ['u.id_usuario', 'u.email', 'u.id_rol', 'u.id_residencia', 'r.nombre as nombre_rol']
            campos_index = {'id_usuario': 0, 'email': 1, 'id_rol': 2, 'id_residencia': 3, 'nombre_rol': 4}
            indice_actual = 5
            
            if tiene_nombre:
                campos_select.append('u.nombre')
                campos_index['nombre'] = indice_actual
                indice_actual += 1
            else:
                campos_index['nombre'] = None
            
            if tiene_apellido:
                campos_select.append('u.apellido')
                campos_index['apellido'] = indice_actual
                indice_actual += 1
            else:
                campos_index['apellido'] = None
            
            campos_select.extend(['res.nombre as nombre_residencia', 'u.activo'])
            campos_index['nombre_residencia'] = indice_actual
            campos_index['activo'] = indice_actual + 1
            indice_actual += 2
            
            # Verificar si existe requiere_cambio_clave
            cursor.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_schema = 'public' 
                AND table_name = 'usuario'
                AND column_name = 'requiere_cambio_clave'
            """)
            tiene_requiere_cambio = cursor.fetchone() is not None
            
            if tiene_requiere_cambio:
                campos_select.append('u.requiere_cambio_clave')
                campos_index['requiere_cambio_clave'] = indice_actual
            else:
                campos_index['requiere_cambio_clave'] = None
            
            query = f"""
                SELECT {', '.join(campos_select)}
                FROM usuario u
                JOIN rol r ON u.id_rol = r.id_rol
                LEFT JOIN residencia res ON u.id_residencia = res.id_residencia
                WHERE u.id_usuario = %s
            """
            
            cursor.execute(query, (g.id_usuario,))
            usuario = cursor.fetchone()
            
            if not usuario:
                app.logger.warning(f"Usuario con id {g.id_usuario} no encontrado en la base de datos")
                return jsonify({'error': 'Usuario no encontrado'}), 404
            
            # Extraer valores de forma segura
            nombre = usuario[campos_index['nombre']] if campos_index['nombre'] is not None and len(usuario) > campos_index['nombre'] else None
            apellido = usuario[campos_index['apellido']] if campos_index['apellido'] is not None and len(usuario) > campos_index['apellido'] else None
            id_residencia = usuario[campos_index['id_residencia']] if len(usuario) > campos_index['id_residencia'] else None
            nombre_residencia = usuario[campos_index['nombre_residencia']] if len(usuario) > campos_index['nombre_residencia'] else None
            requiere_cambio_clave = usuario[campos_index['requiere_cambio_clave']] if campos_index['requiere_cambio_clave'] is not None and len(usuario) > campos_index['requiere_cambio_clave'] else False
            
            # Obtener residencias asignadas (SOLO desde usuario_residencia - igual que permisos)
            residencias = []
            try:
                # Verificar si la tabla usuario_residencia existe
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables 
                        WHERE table_schema = 'public'
                        AND table_name = 'usuario_residencia'
                    )
                """)
                resultado = cursor.fetchone()
                tabla_existe = resultado[0] if resultado else False
                
                if tabla_existe:
                    # Usar usuario_residencia (modo nuevo)
                    if g.id_rol == ADMIN_ROLE_ID:
                        # Administrador: acceso total
                        cursor.execute("SELECT id_residencia, nombre FROM residencia WHERE activa = TRUE ORDER BY nombre")
                        todas_residencias = cursor.fetchall()
                        residencias = [{'id_residencia': r[0], 'nombre': r[1]} for r in todas_residencias]
                    else:
                        # Usuario normal: obtener residencias asignadas ACTIVAS
                        cursor.execute("""
                            SELECT ur.id_residencia, res.nombre
                            FROM usuario_residencia ur
                            JOIN residencia res ON ur.id_residencia = res.id_residencia
                            WHERE ur.id_usuario = %s AND res.activa = TRUE
                            ORDER BY res.nombre
                        """, (g.id_usuario,))
                        residencias_data = cursor.fetchall()
                        residencias = [{'id_residencia': r[0], 'nombre': r[1]} for r in residencias_data]
                else:
                    # Modo legacy: usar id_residencia de usuario
                    if id_residencia:
                        cursor.execute("SELECT id_residencia, nombre FROM residencia WHERE id_residencia = %s AND activa = TRUE", (id_residencia,))
                        res_data = cursor.fetchone()
                        if res_data:
                            residencias = [{'id_residencia': res_data[0], 'nombre': res_data[1]}]
            except Exception as e:
                app.logger.error(f"Error al obtener residencias del usuario: {str(e)}")
                import traceback
                app.logger.error(traceback.format_exc())
                # Si falla, usar id_residencia del usuario si existe
                if id_residencia:
                    residencias = [{'id_residencia': id_residencia, 'nombre': nombre_residencia if nombre_residencia else 'N/A'}]
            
            # Obtener permisos del usuario (individuales + heredados del rol)
            permisos = []
            try:
                id_rol_actual = usuario[campos_index['id_rol']]
                if id_rol_actual == ADMIN_ROLE_ID:
                    # Administrador tiene todos los permisos
                    cursor.execute("""
                        SELECT nombre_permiso
                        FROM permiso
                        WHERE activo = TRUE
                        ORDER BY nombre_permiso
                    """)
                    permisos_data = cursor.fetchall()
                    permisos = [p[0] for p in permisos_data]
                else:
                    # Combinar permisos individuales + heredados del rol
                    cursor.execute("""
                        SELECT DISTINCT nombre_permiso
                        FROM (
                            SELECT nombre_permiso FROM usuario_permiso WHERE id_usuario = %s
                            UNION
                            SELECT nombre_permiso FROM rol_permiso WHERE id_rol = %s
                        ) AS permisos_combinados
                        ORDER BY nombre_permiso
                    """, (g.id_usuario, id_rol_actual))
                    permisos_data = cursor.fetchall()
                    permisos = [p[0] for p in permisos_data]
            except Exception as e:
                app.logger.warning(f"Error al obtener permisos del usuario (tabla puede no existir): {str(e)}")
                # Si no existe la tabla o hay error, lista vacía (usuario sin permisos)
                permisos = []
            
            respuesta = {
                'id_usuario': usuario[campos_index['id_usuario']],
                'email': usuario[campos_index['email']],
                'nombre': nombre,
                'apellido': apellido,
                'id_rol': usuario[campos_index['id_rol']],
                'id_residencia': id_residencia,
                'nombre_rol': usuario[campos_index['nombre_rol']],
                'nombre_residencia': nombre_residencia,
                'activo': usuario[campos_index['activo']] if len(usuario) > campos_index['activo'] else True,
                'requiere_cambio_clave': requiere_cambio_clave,
                'residencias': residencias,
                'permisos': permisos  # Permisos individuales del usuario
            }
            
            return jsonify(respuesta), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al obtener usuario actual: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({'error': f'Error al obtener información del usuario: {str(e)}'}), 500


@app.route('/api/v1/usuarios', methods=['GET'])
def listar_usuarios():
    """Lista usuarios del sistema. Administrador ve todos."""
    # Solo Administrador (id_rol = 2) puede acceder
    if g.id_rol != ADMIN_ROLE_ID:
        return jsonify({'error': 'No tienes permisos para acceder a esta información'}), 403
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar si la tabla usuario_residencia existe
            tabla_existe = False
            try:
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.tables 
                        WHERE table_schema = 'public'
                        AND table_name = 'usuario_residencia'
                    )
                """)
                resultado = cursor.fetchone()
                tabla_existe = resultado[0] if resultado else False
            except Exception as e:
                app.logger.warning(f"No se pudo verificar existencia de tabla usuario_residencia: {str(e)}")
                tabla_existe = False
            
            # Verificar si la columna id_residencia existe en usuario (modo legacy)
            columna_id_residencia_existe = False
            try:
                cursor.execute("""
                    SELECT EXISTS (
                        SELECT FROM information_schema.columns 
                        WHERE table_schema = 'public'
                        AND table_name = 'usuario'
                        AND column_name = 'id_residencia'
                    )
                """)
                resultado = cursor.fetchone()
                columna_id_residencia_existe = resultado[0] if resultado else False
            except Exception as e:
                app.logger.warning(f"No se pudo verificar existencia de columna id_residencia: {str(e)}")
                columna_id_residencia_existe = False
            
            if not tabla_existe:
                # Si no existe usuario_residencia, intentar usar id_residencia de usuario (modo legacy)
                if columna_id_residencia_existe:
                    try:
                        # Administrador ve todos los usuarios (sin filtro de rol)
                        cursor.execute("""
                            SELECT u.id_usuario, u.email, u.nombre, u.apellido, u.id_rol, u.activo, u.fecha_creacion,
                                   r.nombre as nombre_rol, u.id_residencia
                            FROM usuario u
                            JOIN rol r ON u.id_rol = r.id_rol
                            ORDER BY u.fecha_creacion DESC
                        """)
                        
                        usuarios = cursor.fetchall()
                        
                        usuarios_con_residencias = []
                        for u in usuarios:
                            id_residencia = u[8] if len(u) > 8 else None
                            # Obtener nombre de residencia
                            nombre_residencia = None
                            if id_residencia:
                                cursor.execute("SELECT nombre FROM residencia WHERE id_residencia = %s", (id_residencia,))
                                res_result = cursor.fetchone()
                                nombre_residencia = res_result[0] if res_result else None
                            
                            usuarios_con_residencias.append({
                                'id_usuario': u[0],
                                'email': u[1],
                                'nombre': u[2],
                                'apellido': u[3],
                                'id_rol': u[4],
                                'activo': u[5],
                                'fecha_creacion': u[6].isoformat() if u[6] else None,
                                'nombre_rol': u[7],
                                'residencias': [
                                    {
                                        'id_residencia': id_residencia,
                                        'nombre': nombre_residencia
                                    }
                                ] if id_residencia and nombre_residencia else []
                            })
                    except Exception as e:
                        app.logger.error(f"Error al obtener usuarios con id_residencia: {str(e)}")
                        # Si falla, usar modo sin residencias
                        # Administrador ve todos los usuarios (sin filtro de rol)
                        cursor.execute("""
                            SELECT u.id_usuario, u.email, u.nombre, u.apellido, u.id_rol, u.activo, u.fecha_creacion,
                                   r.nombre as nombre_rol
                            FROM usuario u
                            JOIN rol r ON u.id_rol = r.id_rol
                            ORDER BY u.fecha_creacion DESC
                        """)
                        
                        usuarios = cursor.fetchall()
                        
                        usuarios_con_residencias = []
                        for u in usuarios:
                            usuarios_con_residencias.append({
                        'id_usuario': u[0],
                        'email': u[1],
                        'nombre': u[2],
                        'apellido': u[3],
                        'id_rol': u[4],
                                'activo': u[5],
                                'fecha_creacion': u[6].isoformat() if u[6] else None,
                                'nombre_rol': u[7],
                                'residencias': []
                            })
                else:
                    # No existe ni usuario_residencia ni id_residencia, devolver usuarios sin residencias
                    # Administrador ve todos los usuarios (sin filtro de rol)
                    cursor.execute("""
                        SELECT u.id_usuario, u.email, u.nombre, u.apellido, u.id_rol, u.activo, u.fecha_creacion,
                               r.nombre as nombre_rol
                        FROM usuario u
                        JOIN rol r ON u.id_rol = r.id_rol
                        ORDER BY u.fecha_creacion DESC
                    """)
                    
                    usuarios = cursor.fetchall()
                    
                    usuarios_con_residencias = []
                    for u in usuarios:
                        usuarios_con_residencias.append({
                            'id_usuario': u[0],
                            'email': u[1],
                            'nombre': u[2],
                            'apellido': u[3],
                            'id_rol': u[4],
                            'activo': u[5],
                            'fecha_creacion': u[6].isoformat() if u[6] else None,
                            'nombre_rol': u[7],
                            'residencias': []
                        })
            else:
                # Usar usuario_residencia (modo nuevo)
                # Administrador ve todos los usuarios (sin filtro de rol)
                cursor.execute("""
                    SELECT u.id_usuario, u.email, u.nombre, u.apellido, u.id_rol, u.activo, u.fecha_creacion,
                           r.nombre as nombre_rol
                    FROM usuario u
                    JOIN rol r ON u.id_rol = r.id_rol
                    ORDER BY u.fecha_creacion DESC
                """)
                
                usuarios = cursor.fetchall()
                
                # Obtener residencias y permisos para cada usuario
                usuarios_con_residencias = []
                for u in usuarios:
                    id_usuario = u[0]
                    # Obtener residencias ACTIVAS del usuario desde usuario_residencia
                    cursor.execute("""
                        SELECT ur.id_residencia, res.nombre
                        FROM usuario_residencia ur
                        JOIN residencia res ON ur.id_residencia = res.id_residencia
                        WHERE ur.id_usuario = %s AND res.activa = TRUE
                        ORDER BY res.nombre
                    """, (id_usuario,))
                    residencias_usuario = cursor.fetchall()
                    
                    # Obtener permisos del usuario (individuales + heredados del rol)
                    try:
                        id_rol_usuario = u[4]  # u[4] es id_rol
                        if id_rol_usuario == ADMIN_ROLE_ID:
                            # Administrador tiene todos los permisos
                            cursor.execute("""
                                SELECT nombre_permiso
                                FROM permiso
                                WHERE activo = TRUE
                                ORDER BY nombre_permiso
                            """)
                            permisos_usuario = [row[0] for row in cursor.fetchall()]
                        else:
                            # Combinar permisos individuales + heredados del rol
                            cursor.execute("""
                                SELECT DISTINCT nombre_permiso
                                FROM (
                                    SELECT nombre_permiso FROM usuario_permiso WHERE id_usuario = %s
                                    UNION
                                    SELECT nombre_permiso FROM rol_permiso WHERE id_rol = %s
                                ) AS permisos_combinados
                                ORDER BY nombre_permiso
                            """, (id_usuario, id_rol_usuario))
                        permisos_usuario = [row[0] for row in cursor.fetchall()]
                    except Exception:
                        # Si la tabla no existe aún, usar lista vacía
                        permisos_usuario = []
                    
                    usuarios_con_residencias.append({
                        'id_usuario': u[0],
                        'email': u[1],
                        'nombre': u[2],
                        'apellido': u[3],
                        'id_rol': u[4],
                        'activo': u[5],
                        'fecha_creacion': u[6].isoformat() if u[6] else None,
                        'nombre_rol': u[7],
                        'residencias': [
                            {
                                'id_residencia': r[0],
                                'nombre': r[1]
                            }
                            for r in residencias_usuario
                        ],
                        'permisos': permisos_usuario
                    })
            
            return jsonify({
                'usuarios': usuarios_con_residencias
            }), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        app.logger.error(f"Error al listar usuarios: {str(e)}\n{error_trace}")
        return jsonify({'error': f'Error al obtener usuarios: {str(e)}'}), 500


@app.route('/api/v1/usuarios/<int:id_usuario>', methods=['PUT'])
def actualizar_usuario(id_usuario):
    """Actualiza un usuario. Los usuarios pueden actualizar su propia información, los Administradores pueden actualizar otros usuarios."""
    # Verificar permisos: el propio usuario o Administrador pueden actualizar
    if g.id_rol != ADMIN_ROLE_ID and g.id_usuario != id_usuario:
        return jsonify({'error': 'No tienes permisos para actualizar este usuario'}), 403
    
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'Datos JSON requeridos'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que el usuario existe
            cursor.execute("SELECT id_usuario, id_rol FROM usuario WHERE id_usuario = %s", (id_usuario,))
            usuario_existente = cursor.fetchone()
            
            if not usuario_existente:
                return jsonify({'error': 'Usuario no encontrado'}), 404
            
            # Si no es Administrador, solo puede actualizar ciertos campos (y solo su propia cuenta)
            if g.id_rol != ADMIN_ROLE_ID:
                # Usuario normal solo puede actualizar nombre, apellido y contraseña (solo su propia cuenta)
                updates = []
                params = []
                
                if 'nombre' in data:
                    updates.append("nombre = %s")
                    params.append(data.get('nombre'))
                
                if 'apellido' in data:
                    updates.append("apellido = %s")
                    params.append(data.get('apellido'))
                
                if 'password' in data:
                    password = data.get('password')
                    # Validar fuerza de contraseña
                    is_valid, error_msg = validate_password_strength(password)
                    if not is_valid:
                        return jsonify({'error': error_msg}), 400
                    password_hash = generate_password_hash(password, method='pbkdf2:sha256')
                    updates.append("password_hash = %s")
                    updates.append("requiere_cambio_clave = FALSE")  # Marcar que ya cambió la contraseña
                    params.append(password_hash)
                
                if not updates:
                    return jsonify({'error': 'No hay campos para actualizar'}), 400
                
                params.append(id_usuario)
                cursor.execute(f"""
                    UPDATE usuario
                    SET {', '.join(updates)}
                    WHERE id_usuario = %s
                """, params)
                conn.commit()
                
                return jsonify({'mensaje': 'Información actualizada exitosamente'}), 200
                
            else:
                # Admin puede actualizar todos los campos excepto la contraseña (se maneja por separado)
                updates = []
                params = []
                
                if 'email' in data:
                    email = data.get('email')
                    valid, error = validate_email(email)
                    if not valid:
                        return jsonify({'error': error}), 400
                    # Verificar que el email no esté en uso por otro usuario
                    cursor.execute("SELECT id_usuario FROM usuario WHERE email = %s AND id_usuario != %s", (email, id_usuario))
                    if cursor.fetchone():
                        return jsonify({'error': 'El email ya está registrado'}), 400
                    updates.append("email = %s")
                    params.append(email)
                
                if 'nombre' in data:
                    updates.append("nombre = %s")
                    params.append(data.get('nombre'))
                
                if 'apellido' in data:
                    updates.append("apellido = %s")
                    params.append(data.get('apellido'))
                
                if 'id_rol' in data:
                    nuevo_id_rol = data.get('id_rol')
                    # Prevenir que admin asigne rol Administrador
                    if nuevo_id_rol == ADMIN_ROLE_ID and g.id_rol != ADMIN_ROLE_ID:
                        return jsonify({'error': 'No tienes permisos para asignar el rol Administrador'}), 403
                    # Solo Administradores pueden asignar rol Administrador
                    if nuevo_id_rol == ADMIN_ROLE_ID and g.id_rol != ADMIN_ROLE_ID:
                        return jsonify({'error': 'Solo los Administradores pueden asignar el rol Administrador'}), 403
                    # Verificar que el rol existe
                    cursor.execute("SELECT id_rol FROM rol WHERE id_rol = %s AND activo = TRUE", (nuevo_id_rol,))
                    if not cursor.fetchone():
                        return jsonify({'error': 'Rol no válido'}), 400
                    updates.append("id_rol = %s")
                    params.append(nuevo_id_rol)
                
                # Manejar residencias múltiples (array)
                if 'residencias' in data:
                    residencias = data.get('residencias', [])
                    if not residencias or len(residencias) == 0:
                        return jsonify({'error': 'Debe asignar al menos una residencia'}), 400
                    
                    # Verificar que todas las residencias existen y están activas
                    placeholders = ','.join(['%s'] * len(residencias))
                    cursor.execute(f"""
                        SELECT id_residencia FROM residencia 
                        WHERE id_residencia IN ({placeholders}) AND activa = TRUE
                    """, tuple(residencias))
                    
                    residencias_validas = [row[0] for row in cursor.fetchall()]
                    
                    if len(residencias_validas) == 0:
                        return jsonify({
                            'error': 'Debe asignar al menos una residencia activa. Ninguna de las residencias seleccionadas está activa.'
                        }), 400
                    
                    if len(residencias_validas) != len(residencias):
                        residencias_invalidas = [r for r in residencias if r not in residencias_validas]
                        return jsonify({
                            'error': f'Una o más residencias no existen o están inactivas: {residencias_invalidas}. Solo se pueden asignar residencias activas.',
                            'residencias_invalidas': residencias_invalidas
                        }), 400
                    
                    # Eliminar residencias actuales del usuario
                    cursor.execute("DELETE FROM usuario_residencia WHERE id_usuario = %s", (id_usuario,))
                    
                    # Asignar nuevas residencias (solo activas)
                    for id_residencia in residencias_validas:
                        cursor.execute("""
                            INSERT INTO usuario_residencia (id_usuario, id_residencia)
                            VALUES (%s, %s)
                            ON CONFLICT DO NOTHING
                        """, (id_usuario, id_residencia))
                    
                    # Verificar que el usuario tiene al menos una residencia activa después de la actualización
                    cursor.execute("""
                        SELECT COUNT(*) 
                        FROM usuario_residencia ur
                        JOIN residencia r ON ur.id_residencia = r.id_residencia
                        WHERE ur.id_usuario = %s AND r.activa = TRUE
                    """, (id_usuario,))
                    total_residencias_activas = cursor.fetchone()[0]
                    
                    if total_residencias_activas == 0:
                        conn.rollback()
                        return jsonify({
                            'error': 'Error: El usuario debe tener al menos una residencia activa. No se puede completar la actualización.'
                        }), 400
                
                # Manejar permisos personalizados (array)
                if 'permisos' in data:
                    permisos = data.get('permisos', [])
                    
                    # Crear tabla usuario_permiso si no existe
                    cursor.execute("""
                        CREATE TABLE IF NOT EXISTS usuario_permiso (
                            id_usuario INTEGER NOT NULL,
                            nombre_permiso VARCHAR(255) NOT NULL,
                            PRIMARY KEY (id_usuario, nombre_permiso),
                            FOREIGN KEY (id_usuario) REFERENCES usuario(id_usuario) ON DELETE CASCADE
                        )
                    """)
                    
                    # Eliminar permisos actuales del usuario
                    cursor.execute("DELETE FROM usuario_permiso WHERE id_usuario = %s", (id_usuario,))
                    
                    # Asignar nuevos permisos
                    if permisos and len(permisos) > 0:
                        for nombre_permiso in permisos:
                            # Verificar que el permiso existe antes de insertarlo
                            cursor.execute("""
                                SELECT COUNT(*) FROM permiso WHERE nombre_permiso = %s AND activo = TRUE
                            """, (nombre_permiso,))
                            permiso_existe = cursor.fetchone()[0] > 0
                            
                            if not permiso_existe:
                                app.logger.warning(f"Permiso '{nombre_permiso}' no existe en la tabla permiso. Se omitirá.")
                                continue
                            
                            cursor.execute("""
                                INSERT INTO usuario_permiso (id_usuario, nombre_permiso)
                                VALUES (%s, %s)
                                ON CONFLICT DO NOTHING
                            """, (id_usuario, nombre_permiso))
                
                if 'activo' in data:
                    updates.append("activo = %s")
                    params.append(data.get('activo'))
                
                if 'password' in data:
                    password = data.get('password')
                    # Validar fuerza de contraseña
                    is_valid, error_msg = validate_password_strength(password)
                    if not is_valid:
                        return jsonify({'error': error_msg}), 400
                    password_hash = generate_password_hash(password, method='pbkdf2:sha256')
                    updates.append("password_hash = %s")
                    updates.append("requiere_cambio_clave = FALSE")  # Marcar que ya cambió la contraseña
                    params.append(password_hash)
                
                # Si hay updates en la tabla usuario, aplicarlos
                if updates:
                    params.append(id_usuario)
                    cursor.execute(f"""
                        UPDATE usuario
                        SET {', '.join(updates)}
                        WHERE id_usuario = %s
                    """, params)
            
            conn.commit()
            
            return jsonify({
                'mensaje': 'Usuario actualizado exitosamente'
            }), 200
            
        except Exception as e:
            conn.rollback()
            import traceback
            error_trace = traceback.format_exc()
            app.logger.error(f"Error al actualizar usuario: {str(e)}\n{error_trace}")
            # En desarrollo, incluir más detalles del error
            if app.config.get('DEBUG'):
                return jsonify({
                    'error': 'Error al actualizar usuario',
                    'detalle': str(e),
                    'traceback': error_trace.split('\n')[-5:]  # Últimas 5 líneas del traceback
                }), 500
            return jsonify({'error': 'Error al actualizar usuario'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.errorhandler(404)
def not_found(error):
    """Manejo de errores 404."""
    return jsonify({'error': 'Endpoint no encontrado'}), 404


@app.errorhandler(500)
def internal_error(error):
    """Manejo de errores 500."""
    return jsonify({'error': 'Error interno del servidor'}), 500


# ============================================================================
# ENDPOINTS DE DOCUMENTACIÓN UNIFICADA
# ============================================================================

@app.route('/api/v1/documentacion', methods=['GET'])
def listar_documentacion():
    """
    Lista todos los documentos con filtros opcionales por tipo de entidad, categoría y residencia.
    
    Query params:
    - tipo_entidad: 'residente', 'proveedor', 'personal' (opcional)
    - categoria: 'medica', 'fiscal', 'sanitaria', 'laboral', 'otra' (opcional)
    - id_residencia: ID de residencia (opcional, se filtra automáticamente por permisos)
    - id_entidad: ID específico de la entidad (opcional)
    """
    try:
        tipo_entidad = request.args.get('tipo_entidad')
        categoria = request.args.get('categoria')
        id_entidad = request.args.get('id_entidad', type=int)
        id_residencia_filtro = request.args.get('id_residencia', type=int)

        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Construir filtro de residencias
            residencias_filtro = ""
            params = []
            
            if g.id_rol != ADMIN_ROLE_ID:
                if g.residencias_acceso:
                    placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                    residencias_filtro = f" AND d.id_residencia IN ({placeholders})"
                    params.extend(g.residencias_acceso)
                else:
                    # Usuario sin residencias
                    app.logger.warning(f"📍 Usuario sin residencias - retornando vacío")
                    return jsonify({'documentos': [], 'total': 0}), 200
            
            # Si hay filtro específico por residencia, aplicarlo
            if id_residencia_filtro:
                # Verificar que el usuario tiene acceso a esa residencia
                if g.id_rol != ADMIN_ROLE_ID:
                    if id_residencia_filtro not in g.residencias_acceso:
                        return jsonify({'error': 'No tienes acceso a esta residencia'}), 403
                # Aplicar filtro específico
                residencias_filtro = " AND d.id_residencia = %s"
                params = [id_residencia_filtro]
            
            # Verificar si la tabla 'documento' existe
            cursor.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_schema = 'public' 
                    AND table_name = 'documento'
                )
            """)
            tabla_documento_existe = cursor.fetchone()[0]
            
            # Construir query unificada: documentos unificados + documentos legacy de residentes
            # Primero: documentos de la tabla unificada 'documento' (solo si existe)
            query_unificados = None
            params_unificados = []
            
            if tabla_documento_existe:
                query_unificados = f"""
                SELECT d.id_documento, d.tipo_entidad, d.id_entidad, d.id_residencia,
                       d.categoria_documento, d.tipo_documento, d.nombre_archivo, d.descripcion,
                       d.fecha_subida, d.url_archivo, d.tamaño_bytes, d.tipo_mime,
                       d.id_usuario_subida, d.activo,
                       res.nombre as nombre_residencia
                FROM documento d
                JOIN residencia res ON d.id_residencia = res.id_residencia
                WHERE d.activo = TRUE
                {residencias_filtro}
            """
            
            params_unificados = params.copy()
            
            # Filtros opcionales para documentos unificados
            if tipo_entidad:
                query_unificados += " AND d.tipo_entidad = %s"
                params_unificados.append(tipo_entidad)
            
            if categoria:
                query_unificados += " AND d.categoria_documento = %s"
                params_unificados.append(categoria)
            
            if id_entidad:
                query_unificados += " AND d.id_entidad = %s"
                params_unificados.append(id_entidad)
                
                # El filtro de residencia ya está aplicado en residencias_filtro
            
            # Segundo: documentos legacy de residentes (tabla documento_residente)
            # Construir filtro de residencias para documento_residente
            residencias_filtro_legacy = ""
            params_legacy_residencias = []
            if id_residencia_filtro:
                # Filtro específico por residencia
                residencias_filtro_legacy = " AND dr.id_residencia = %s"
                params_legacy_residencias = [id_residencia_filtro]
            elif residencias_filtro:
                # Construir el filtro con el alias correcto
                if g.id_rol != ADMIN_ROLE_ID and g.residencias_acceso:
                    placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                    residencias_filtro_legacy = f" AND dr.id_residencia IN ({placeholders})"
                    params_legacy_residencias = g.residencias_acceso.copy()
            else:
                # Si no hay filtro de residencias (Administrador sin filtro), no aplicar filtro
                residencias_filtro_legacy = ""
                params_legacy_residencias = []
            
            # Verificar si existe la columna categoria_documento en documento_residente
            tiene_categoria = False
            try:
                cursor.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'documento_residente' 
                      AND column_name = 'categoria_documento'
                """)
                tiene_categoria = cursor.fetchone() is not None
            except Exception as e:
                app.logger.warning(f"Error al verificar columna categoria_documento: {str(e)}")
                tiene_categoria = False
            
            if tiene_categoria:
                query_legacy = f"""
                SELECT dr.id_documento, 'residente' as tipo_entidad, dr.id_residente as id_entidad, 
                       dr.id_residencia,
                       COALESCE(dr.categoria_documento, 'otra') as categoria_documento, 
                       dr.tipo_documento, dr.nombre_archivo, dr.descripcion,
                       dr.fecha_subida, dr.url_archivo, dr.tamaño_bytes, dr.tipo_mime,
                       NULL as id_usuario_subida, TRUE as activo,
                       res.nombre as nombre_residencia
                FROM documento_residente dr
                JOIN residencia res ON dr.id_residencia = res.id_residencia
                WHERE 1=1
                {residencias_filtro_legacy}
            """
            else:
                # Si no tiene columna categoria_documento, usar 'otra' por defecto
                query_legacy = f"""
                    SELECT dr.id_documento, 'residente' as tipo_entidad, dr.id_residente as id_entidad, 
                           dr.id_residencia,
                           'otra' as categoria_documento, 
                           dr.tipo_documento, dr.nombre_archivo, dr.descripcion,
                           dr.fecha_subida, dr.url_archivo, dr.tamaño_bytes, dr.tipo_mime,
                           NULL as id_usuario_subida, TRUE as activo,
                           res.nombre as nombre_residencia
                    FROM documento_residente dr
                    JOIN residencia res ON dr.id_residencia = res.id_residencia
                    WHERE 1=1
                    {residencias_filtro_legacy}
                """
            
            params_legacy = params_legacy_residencias.copy()
            
            # Filtros opcionales para documentos legacy
            if tipo_entidad:
                if tipo_entidad == 'residente':
                    # Solo incluir documentos legacy si el filtro es 'residente'
                    pass  # Ya está incluido
                else:
                    # Si el filtro es otro tipo, no incluir documentos legacy
                    query_legacy = None
            else:
                # Si no hay filtro de tipo, incluir documentos legacy
                pass
            
            if categoria and query_legacy:
                # Para documentos legacy, mapear categorías si es necesario
                if tiene_categoria:
                    # Si tiene columna categoria_documento, filtrar por ella
                    if categoria == 'otra':
                        # Incluir documentos sin categoría o con categoría 'otra'
                        query_legacy += " AND (dr.categoria_documento IS NULL OR dr.categoria_documento = 'otra')"
                    else:
                        query_legacy += " AND dr.categoria_documento = %s"
                    params_legacy.append(categoria)
                else:
                    # Si no tiene columna categoria_documento, solo mostrar si el filtro es 'otra'
                    if categoria != 'otra':
                        # Si el filtro no es 'otra', no mostrar documentos legacy sin categoría
                        query_legacy = None
            
            if id_entidad and query_legacy:
                query_legacy += " AND dr.id_residente = %s"
                params_legacy.append(id_entidad)
            
            # Ejecutar consultas y combinar resultados
            documentos = []
            
            # Consultar documentos unificados (solo si la tabla existe y hay query)
            if query_unificados:
                try:
                    cursor.execute(query_unificados, params_unificados)
                    documentos_unificados = cursor.fetchall()
                    documentos.extend(documentos_unificados)
                except Exception as e:
                    app.logger.error(f"Error al consultar documentos unificados: {str(e)}")
                    app.logger.error(f"Query: {query_unificados}")
                    app.logger.error(f"Params: {params_unificados}")
                    # Continuar con documentos legacy si hay error
            
            # Consultar documentos legacy (solo si no hay filtro de tipo o si es 'residente')
            if query_legacy and (not tipo_entidad or tipo_entidad == 'residente'):
                try:
                    cursor.execute(query_legacy, params_legacy)
                    documentos_legacy = cursor.fetchall()
                    documentos.extend(documentos_legacy)
                except Exception as e:
                    app.logger.error(f"Error al consultar documentos legacy: {str(e)}")
                    app.logger.error(f"Query: {query_legacy}")
                    app.logger.error(f"Params: {params_legacy}")
                    # Continuar sin documentos legacy
            
            # Consultar facturas de pagos a proveedores (si no hay filtro de tipo o si es 'proveedor')
            if not tipo_entidad or tipo_entidad == 'proveedor':
                try:
                    # Verificar si existe la columna factura_blob_path
                    cursor.execute("""
                        SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_name = 'pago_proveedor' 
                          AND column_name = 'factura_blob_path'
                    """)
                    tiene_columna_factura = cursor.fetchone() is not None
                    
                    if tiene_columna_factura:
                        query_facturas = f"""
                            SELECT pp.id_pago, 'pago_proveedor' as tipo_entidad, 
                                   pp.id_residencia as id_entidad, pp.id_residencia,
                                   'fiscal' as categoria_documento,
                                   'Factura' as tipo_documento,
                                   CONCAT('Factura_', COALESCE(pp.numero_factura, pp.id_pago::text), '.pdf') as nombre_archivo,
                                   CONCAT('Factura de pago: ', pp.proveedor, ' - ', pp.concepto) as descripcion,
                                   pp.fecha_creacion as fecha_subida,
                                   pp.factura_blob_path as url_archivo,
                                   NULL as tamaño_bytes,
                                   'application/pdf' as tipo_mime,
                                   NULL as id_usuario_subida,
                                   TRUE as activo,
                                   res.nombre as nombre_residencia
                            FROM pago_proveedor pp
                            JOIN residencia res ON pp.id_residencia = res.id_residencia
                            WHERE pp.factura_blob_path IS NOT NULL 
                              AND pp.factura_blob_path != ''
                        """
                        
                        params_facturas = []
                        if id_residencia_filtro:
                            # Filtro específico por residencia
                            query_facturas += " AND pp.id_residencia = %s"
                            params_facturas.append(id_residencia_filtro)
                        elif residencias_filtro:
                            if g.id_rol != ADMIN_ROLE_ID and g.residencias_acceso:
                                placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                                query_facturas += f" AND pp.id_residencia IN ({placeholders})"
                                params_facturas.extend(g.residencias_acceso)
                        
                        cursor.execute(query_facturas, params_facturas)
                        facturas = cursor.fetchall()
                        documentos.extend(facturas)
                except Exception as e:
                    app.logger.error(f"Error al consultar facturas de pagos: {str(e)}")
                    # Continuar sin facturas
            
            # Ordenar por fecha_subida descendente
            try:
                documentos.sort(key=lambda x: x[8] if x[8] else datetime.min, reverse=True)
            except Exception as e:
                app.logger.error(f"Error al ordenar documentos: {str(e)}")
                # Si falla el ordenamiento, continuar sin ordenar
            
            # Obtener nombres de las entidades
            resultado = []
            for doc in documentos:
                try:
                    # Verificar que el documento tiene suficientes campos
                    if len(doc) < 15:
                        app.logger.warning(f"Documento con menos campos de los esperados: {len(doc)} campos")
                        continue
                    
                    tipo_ent = doc[1]
                    id_ent = doc[2]
                    nombre_entidad = None
                    
                    # Obtener nombre de la entidad
                    try:
                        if tipo_ent == 'residente':
                            cursor.execute("SELECT nombre, apellido FROM residente WHERE id_residente = %s", (id_ent,))
                            ent = cursor.fetchone()
                            if ent:
                                nombre_entidad = f"{ent[0]} {ent[1]}"
                        elif tipo_ent == 'proveedor':
                            cursor.execute("SELECT nombre FROM proveedor WHERE id_proveedor = %s", (id_ent,))
                            ent = cursor.fetchone()
                            if ent:
                                nombre_entidad = ent[0]
                        elif tipo_ent == 'personal':
                            cursor.execute("SELECT nombre, apellido FROM personal WHERE id_personal = %s", (id_ent,))
                            ent = cursor.fetchone()
                            if ent:
                                nombre_entidad = f"{ent[0]} {ent[1]}"
                        elif tipo_ent == 'pago_proveedor':
                            # Para facturas de pagos, obtener el nombre del proveedor del pago
                            cursor.execute("SELECT proveedor FROM pago_proveedor WHERE id_pago = %s", (doc[0],))
                            ent = cursor.fetchone()
                            if ent:
                                nombre_entidad = ent[0]
                    except Exception as e:
                        app.logger.warning(f"Error al obtener nombre de entidad {tipo_ent} {id_ent}: {str(e)}")
                    
                    url_descarga = None
                    if doc[9]:  # Si hay url_archivo
                        try:
                            url_descarga = get_document_url(doc[9], expiration_minutes=60)
                        except Exception as e:
                            app.logger.warning(f"Error al generar URL de descarga: {str(e)}")
                    
                    resultado.append({
                        'id_documento': doc[0],
                        'tipo_entidad': doc[1],
                        'id_entidad': doc[2],
                        'nombre_entidad': nombre_entidad,
                        'id_residencia': doc[3],
                        'nombre_residencia': doc[14] if len(doc) > 14 else None,
                        'categoria_documento': doc[4],
                        'tipo_documento': doc[5],
                        'nombre_archivo': doc[6],
                        'descripcion': doc[7],
                        'fecha_subida': doc[8].isoformat() if doc[8] else None,
                        'url_archivo': doc[9],
                        'url_descarga': url_descarga,
                        'tamaño_bytes': doc[10],
                        'tipo_mime': doc[11],
                        'id_usuario_subida': doc[12]
                    })
                except Exception as e:
                    app.logger.error(f"Error al procesar documento: {str(e)}")
                    app.logger.error(f"Documento: {doc}")
                    continue
            
            return jsonify({
                'documentos': resultado,
                'total': len(resultado),
                'filtros': {
                    'tipo_entidad': tipo_entidad,
                    'categoria': categoria,
                    'id_entidad': id_entidad
                }
            }), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al listar documentación: {str(e)}")
        return jsonify({'error': 'Error al obtener documentación'}), 500


@app.route('/api/v1/documentacion', methods=['POST'])
def crear_documento_unificado():
    """
    Crea un nuevo documento y lo sube a Cloud Storage.
    
    Request (multipart/form-data):
    - archivo: Archivo a subir
    - tipo_entidad: 'residente', 'proveedor', 'personal'
    - id_entidad: ID de la entidad
    - categoria_documento: 'medica', 'fiscal', 'sanitaria', 'laboral', 'otra'
    - tipo_documento: Tipo específico del documento
    - descripcion: Descripción opcional
    """
    try:
        if 'archivo' not in request.files:
            return jsonify({'error': 'Archivo requerido'}), 400
        
        archivo = request.files['archivo']
        if archivo.filename == '':
            return jsonify({'error': 'Nombre de archivo vacío'}), 400
        
        tipo_entidad = request.form.get('tipo_entidad')
        id_entidad = request.form.get('id_entidad', type=int)
        categoria_documento = request.form.get('categoria_documento')
        tipo_documento = request.form.get('tipo_documento')
        descripcion = request.form.get('descripcion', '')
        
        if not all([tipo_entidad, id_entidad, categoria_documento, tipo_documento]):
            return jsonify({'error': 'tipo_entidad, id_entidad, categoria_documento y tipo_documento son requeridos'}), 400
        
        if tipo_entidad not in ['residente', 'proveedor', 'personal']:
            return jsonify({'error': 'tipo_entidad debe ser: residente, proveedor o personal'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Verificar que la entidad existe y obtener su residencia
            id_residencia = None
            if tipo_entidad == 'residente':
                cursor.execute("SELECT id_residencia FROM residente WHERE id_residente = %s", (id_entidad,))
                res = cursor.fetchone()
                if not res:
                    return jsonify({'error': 'Residente no encontrado'}), 404
                id_residencia = res[0]
            elif tipo_entidad == 'proveedor':
                cursor.execute("SELECT id_residencia FROM proveedor WHERE id_proveedor = %s", (id_entidad,))
                res = cursor.fetchone()
                if not res:
                    return jsonify({'error': 'Proveedor no encontrado'}), 404
                id_residencia = res[0]
            elif tipo_entidad == 'personal':
                cursor.execute("SELECT id_residencia FROM personal WHERE id_personal = %s", (id_entidad,))
                res = cursor.fetchone()
                if not res:
                    return jsonify({'error': 'Personal no encontrado'}), 404
                id_residencia = res[0]
            
            # Verificar permisos de acceso a la residencia
            is_valid, error_response = validate_residencia_access(id_residencia)
            if not is_valid:
                return error_response
            
            # Leer contenido del archivo
            file_content = archivo.read()
            nombre_archivo = archivo.filename
            content_type = archivo.content_type
            
            # Subir a Cloud Storage usando función unificada
            from storage_manager import upload_document_unificado
            blob_path = upload_document_unificado(
                file_content, id_residencia, tipo_entidad, id_entidad,
                tipo_documento, nombre_archivo, content_type
            )
            
            if not blob_path:
                return jsonify({'error': 'Error al subir el archivo a Cloud Storage'}), 500
            
            # Guardar en base de datos
            cursor.execute("""
                INSERT INTO documento (tipo_entidad, id_entidad, id_residencia, categoria_documento,
                                      tipo_documento, nombre_archivo, descripcion, url_archivo,
                                      tamaño_bytes, tipo_mime, id_usuario_subida, activo)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE)
                RETURNING id_documento, fecha_subida
            """, (
                tipo_entidad, id_entidad, id_residencia, categoria_documento,
                tipo_documento, nombre_archivo, descripcion, blob_path,
                len(file_content), content_type, g.id_usuario
            ))
            
            resultado = cursor.fetchone()
            id_documento = resultado[0]
            conn.commit()
            
            return jsonify({
                'id_documento': id_documento,
                'mensaje': 'Documento subido exitosamente',
                'url_archivo': blob_path
            }), 201
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al crear documento: {str(e)}")
            return jsonify({'error': f'Error al crear documento: {str(e)}'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/documentacion/<int:id_documento>', methods=['DELETE'])
def eliminar_documento_unificado(id_documento):
    """Elimina un documento unificado (marca como inactivo y elimina de Cloud Storage)."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Buscar primero en la tabla unificada 'documento'
            cursor.execute("""
                SELECT id_documento, url_archivo, id_residencia, tipo_entidad, id_entidad, 'documento' as tabla_origen
                FROM documento 
                WHERE id_documento = %s AND activo = TRUE
            """, (id_documento,))
            
            doc = cursor.fetchone()
            tabla_origen = 'documento'
            
            # Si no se encuentra, buscar en documento_residente (legacy)
            if not doc:
                cursor.execute("""
                    SELECT id_documento, url_archivo, id_residencia, 'residente' as tipo_entidad, id_residente as id_entidad, 'documento_residente' as tabla_origen
                    FROM documento_residente 
                    WHERE id_documento = %s
                """, (id_documento,))
                
                doc = cursor.fetchone()
                tabla_origen = 'documento_residente'
            
            if not doc:
                return jsonify({'error': 'Documento no encontrado'}), 404
            
            id_residencia = doc[2]
            
            # Verificar permisos
            is_valid, error_response = validate_residencia_access(id_residencia)
            if not is_valid:
                return error_response
            
            # Eliminar de Cloud Storage
            if doc[1]:  # Si hay url_archivo
                delete_document(doc[1])
            
            # Eliminar según la tabla de origen
            if tabla_origen == 'documento':
                # Marcar como inactivo (soft delete)
                cursor.execute("""
                    UPDATE documento
                    SET activo = FALSE
                    WHERE id_documento = %s
                    RETURNING id_documento
                """, (id_documento,))
            else:
                # Eliminar físicamente de documento_residente
                cursor.execute("""
                    DELETE FROM documento_residente
                    WHERE id_documento = %s
                    RETURNING id_documento
                """, (id_documento,))
            
            conn.commit()
            
            return jsonify({
                'mensaje': 'Documento eliminado exitosamente'
            }), 200
            
        except Exception as e:
            conn.rollback()
            app.logger.error(f"Error al eliminar documento: {str(e)}")
            return jsonify({'error': 'Error al eliminar documento'}), 500
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500


@app.route('/api/v1/documentacion/<int:id_documento>/descargar', methods=['GET'])
def descargar_documento_unificado(id_documento):
    """
    Genera una URL firmada para descargar un documento.
    Soporta documentos de la tabla 'documento', 'documento_residente' y facturas de 'pago_proveedor'.
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            # Buscar primero en pago_proveedor (facturas)
            doc = None
            try:
                cursor.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_name = 'pago_proveedor' 
                      AND column_name = 'factura_blob_path'
                """)
                tiene_columna_factura = cursor.fetchone() is not None
                
                if tiene_columna_factura:
                    cursor.execute("""
                        SELECT factura_blob_path, id_residencia, proveedor FROM pago_proveedor
                        WHERE id_pago = %s AND factura_blob_path IS NOT NULL AND factura_blob_path != ''
                    """, (id_documento,))
                    
                    factura = cursor.fetchone()
                    if factura:
                        # Verificar permisos
                        is_valid, error_response = validate_residencia_access(factura[1])
                        if not is_valid:
                            return error_response
                        
                        # Generar URL de descarga y devolver inmediatamente
                        url_descarga = get_document_url(factura[0], expiration_minutes=60)
                        if url_descarga:
                            return jsonify({
                                'url_descarga': url_descarga,
                                'nombre_archivo': f"Factura_{factura[2] or id_documento}.pdf"
                            }), 200
                        else:
                            return jsonify({'error': 'Error al generar URL de descarga'}), 500
            except Exception as e:
                app.logger.warning(f"Error al buscar factura: {str(e)}")
            
            # Si no se encuentra en facturas, buscar en la tabla unificada 'documento'
            if not doc:
                cursor.execute("""
                    SELECT url_archivo, id_residencia FROM documento
                    WHERE id_documento = %s AND activo = TRUE
                """, (id_documento,))
                
                doc = cursor.fetchone()
            
            # Si no se encuentra, buscar en documento_residente (legacy)
            if not doc:
                cursor.execute("""
                    SELECT url_archivo, id_residencia FROM documento_residente
                    WHERE id_documento = %s
                """, (id_documento,))
                
                doc = cursor.fetchone()
            
            if not doc:
                return jsonify({'error': 'Documento no encontrado'}), 404
            
            # Verificar permisos
            is_valid, error_response = validate_residencia_access(doc[1])
            if not is_valid:
                return error_response
            
            url_descarga = get_document_url(doc[0], expiration_minutes=60)
            
            if not url_descarga:
                return jsonify({'error': 'Error al generar URL de descarga'}), 500
            
            return jsonify({
                'url_descarga': url_descarga,
                'expira_en_minutos': 60
            }), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al generar URL de descarga: {str(e)}")
        return jsonify({'error': 'Error al generar URL de descarga'}), 500


@app.route('/api/v1/documentacion/categorias', methods=['GET'])
def listar_categorias():
    """Lista las categorías de documentos disponibles."""
    categorias = [
        {'valor': 'medica', 'nombre': 'Médica'},
        {'valor': 'fiscal', 'nombre': 'Fiscal'},
        {'valor': 'sanitaria', 'nombre': 'Sanitaria'},
        {'valor': 'laboral', 'nombre': 'Laboral'},
        {'valor': 'otra', 'nombre': 'Otra'}
    ]
    return jsonify({'categorias': categorias}), 200


@app.route('/api/v1/documentacion/tipos-entidad', methods=['GET'])
def listar_tipos_entidad():
    """Lista los tipos de entidades disponibles."""
    tipos = [
        {'valor': 'residente', 'nombre': 'Residente'},
        {'valor': 'proveedor', 'nombre': 'Proveedor'},
        {'valor': 'personal', 'nombre': 'Personal'}
    ]
    return jsonify({'tipos': tipos}), 200


# ============================================================================
# ENDPOINTS DE HISTÓRICOS - COBROS Y PAGOS
# ============================================================================

@app.route('/api/v1/historicos', methods=['GET'])
def listar_historicos():
    """
    Lista todos los cobros y pagos históricos con filtros opcionales.
    
    Query params:
    - tipo: 'cobros', 'pagos', 'todos' (default: 'todos')
    - fecha_desde: Fecha inicio (YYYY-MM-DD)
    - fecha_hasta: Fecha fin (YYYY-MM-DD)
    - id_residencia: ID de residencia (opcional)
    - estado: 'pendiente', 'cobrado', 'pagado' (opcional)
    - exportar: 'pdf', 'excel' (opcional, si se especifica genera el archivo)
    """
    try:
        tipo = request.args.get('tipo', 'todos')  # 'cobros', 'pagos', 'todos'
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        id_residencia_filtro = request.args.get('id_residencia', type=int)
        estado_filtro = request.args.get('estado')
        exportar = request.args.get('exportar')  # 'pdf' o 'excel'
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cobros = []
            pagos = []
            
            # Obtener cobros (pago_residente)
            if tipo in ['cobros', 'todos']:
                query_cobros = """
                    SELECT p.id_pago, p.id_residente, r.nombre || ' ' || r.apellido as residente,
                           p.monto, p.fecha_pago, p.fecha_prevista, p.mes_pagado, p.concepto,
                           p.metodo_pago, p.estado, p.es_cobro_previsto, p.observaciones, 
                           p.fecha_creacion, res.id_residencia, res.nombre as nombre_residencia
                    FROM pago_residente p
                    JOIN residente r ON p.id_residente = r.id_residente
                    JOIN residencia res ON p.id_residencia = res.id_residencia
                    WHERE 1=1
                """
                params_cobros = []
                
                # Filtrar por residencias de acceso
                if g.id_rol != ADMIN_ROLE_ID:
                    if g.residencias_acceso:
                        placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                        query_cobros += f" AND p.id_residencia IN ({placeholders})"
                        params_cobros.extend(g.residencias_acceso)
                    else:
                        query_cobros += " AND FALSE"  # Sin acceso
                
                if id_residencia_filtro:
                    query_cobros += " AND p.id_residencia = %s"
                    params_cobros.append(id_residencia_filtro)
                
                if fecha_desde:
                    query_cobros += " AND (p.fecha_pago >= %s OR p.fecha_prevista >= %s OR (p.fecha_pago IS NULL AND p.fecha_prevista IS NULL AND p.fecha_creacion::date >= %s))"
                    params_cobros.extend([fecha_desde, fecha_desde, fecha_desde])
                
                if fecha_hasta:
                    query_cobros += " AND (p.fecha_pago <= %s OR p.fecha_prevista <= %s OR (p.fecha_pago IS NULL AND p.fecha_prevista IS NULL AND p.fecha_creacion::date <= %s))"
                    params_cobros.extend([fecha_hasta, fecha_hasta, fecha_hasta])
                
                if estado_filtro:
                    query_cobros += " AND p.estado = %s"
                    params_cobros.append(estado_filtro)
                
                query_cobros += " ORDER BY COALESCE(p.fecha_pago, p.fecha_prevista, p.fecha_creacion::date) DESC, p.fecha_creacion DESC"
                
                cursor.execute(query_cobros, params_cobros)
                cobros_raw = cursor.fetchall()
                
                for cobro in cobros_raw:
                    cobros.append({
                        'id_pago': cobro[0],
                        'id_residente': cobro[1],
                        'residente': cobro[2],
                        'monto': float(cobro[3]) if cobro[3] else 0,
                        'fecha_pago': cobro[4].isoformat() if cobro[4] else None,
                        'fecha_prevista': cobro[5].isoformat() if cobro[5] else None,
                        'mes_pagado': cobro[6],
                        'concepto': cobro[7],
                        'metodo_pago': cobro[8],
                        'estado': cobro[9],
                        'es_cobro_previsto': cobro[10],
                        'observaciones': cobro[11],
                        'fecha_creacion': cobro[12].isoformat() if cobro[12] else None,
                        'id_residencia': cobro[13],
                        'nombre_residencia': cobro[14],
                        'tipo': 'cobro'
                    })
            
            # Obtener pagos (pago_proveedor)
            if tipo in ['pagos', 'todos']:
                query_pagos = """
                    SELECT p.id_pago, p.proveedor, p.concepto, p.monto, p.fecha_pago, 
                           p.fecha_prevista, p.metodo_pago, p.estado, p.numero_factura,
                           p.es_estimacion, p.observaciones, p.fecha_creacion,
                           res.id_residencia, res.nombre as nombre_residencia
                    FROM pago_proveedor p
                    JOIN residencia res ON p.id_residencia = res.id_residencia
                    WHERE 1=1
                """
                params_pagos = []
                
                # Filtrar por residencias de acceso
                if g.id_rol != ADMIN_ROLE_ID:
                    if g.residencias_acceso:
                        placeholders = ','.join(['%s'] * len(g.residencias_acceso))
                        query_pagos += f" AND p.id_residencia IN ({placeholders})"
                        params_pagos.extend(g.residencias_acceso)
                    else:
                        query_pagos += " AND FALSE"  # Sin acceso
                
                if id_residencia_filtro:
                    query_pagos += " AND p.id_residencia = %s"
                    params_pagos.append(id_residencia_filtro)
                
                if fecha_desde:
                    query_pagos += " AND (p.fecha_pago >= %s OR p.fecha_prevista >= %s OR (p.fecha_pago IS NULL AND p.fecha_prevista IS NULL AND p.fecha_creacion::date >= %s))"
                    params_pagos.extend([fecha_desde, fecha_desde, fecha_desde])
                
                if fecha_hasta:
                    query_pagos += " AND (p.fecha_pago <= %s OR p.fecha_prevista <= %s OR (p.fecha_pago IS NULL AND p.fecha_prevista IS NULL AND p.fecha_creacion::date <= %s))"
                    params_pagos.extend([fecha_hasta, fecha_hasta, fecha_hasta])
                
                if estado_filtro:
                    query_pagos += " AND p.estado = %s"
                    params_pagos.append(estado_filtro)
                
                query_pagos += " ORDER BY COALESCE(p.fecha_pago, p.fecha_prevista, p.fecha_creacion::date) DESC, p.fecha_creacion DESC"
                
                cursor.execute(query_pagos, params_pagos)
                pagos_raw = cursor.fetchall()
                
                for pago in pagos_raw:
                    pagos.append({
                        'id_pago': pago[0],
                        'proveedor': pago[1],
                        'concepto': pago[2],
                        'monto': float(pago[3]) if pago[3] else 0,
                        'fecha_pago': pago[4].isoformat() if pago[4] else None,
                        'fecha_prevista': pago[5].isoformat() if pago[5] else None,
                        'metodo_pago': pago[6],
                        'estado': pago[7],
                        'numero_factura': pago[8],
                        'es_estimacion': pago[9],
                        'observaciones': pago[10],
                        'fecha_creacion': pago[11].isoformat() if pago[11] else None,
                        'id_residencia': pago[12],
                        'nombre_residencia': pago[13],
                        'tipo': 'pago'
                    })
            
            # Si se solicita exportación, generar archivo
            if exportar == 'pdf':
                return generar_pdf_historicos(cobros, pagos, fecha_desde, fecha_hasta)
            elif exportar == 'excel':
                return generar_excel_historicos(cobros, pagos, fecha_desde, fecha_hasta)
            
            # Retornar JSON
            return jsonify({
                'cobros': cobros,
                'pagos': pagos,
                'total_cobros': len(cobros),
                'total_pagos': len(pagos),
                'total_cobros_monto': sum(c['monto'] for c in cobros),
                'total_pagos_monto': sum(p['monto'] for p in pagos),
                'filtros': {
                    'tipo': tipo,
                    'fecha_desde': fecha_desde,
                    'fecha_hasta': fecha_hasta,
                    'id_residencia': id_residencia_filtro,
                    'estado': estado_filtro
                }
            }), 200
            
        finally:
            cursor.close()
            conn.close()
            
    except Exception as e:
        app.logger.error(f"Error al listar históricos: {str(e)}")
        return jsonify({'error': 'Error al obtener históricos'}), 500


def generar_pdf_historicos(cobros, pagos, fecha_desde=None, fecha_hasta=None):
    """Genera un PDF con los históricos de cobros y pagos."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=1  # Centrado
        )
        elements.append(Paragraph("Histórico de Cobros y Pagos", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Información de filtros
        if fecha_desde or fecha_hasta:
            filtro_text = "Período: "
            if fecha_desde:
                filtro_text += f"Desde {fecha_desde}"
            if fecha_hasta:
                filtro_text += f" Hasta {fecha_hasta}"
            elements.append(Paragraph(filtro_text, styles['Normal']))
            elements.append(Spacer(1, 0.1*inch))
        
        # Cobros
        if cobros:
            elements.append(Paragraph(f"<b>COBROS A RESIDENTES ({len(cobros)})</b>", styles['Heading2']))
            elements.append(Spacer(1, 0.1*inch))
            
            data_cobros = [['Residente', 'Monto', 'Fecha Pago', 'Estado', 'Concepto']]
            for cobro in cobros:
                fecha = cobro.get('fecha_pago') or cobro.get('fecha_prevista') or '-'
                if fecha and fecha != '-':
                    fecha = fecha.split('T')[0] if 'T' in str(fecha) else fecha
                data_cobros.append([
                    cobro.get('residente', '-'),
                    f"€{cobro.get('monto', 0):.2f}",
                    fecha,
                    cobro.get('estado', '-'),
                    cobro.get('concepto', '-')[:50]
                ])
            
            table_cobros = Table(data_cobros, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 2*inch])
            table_cobros.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            elements.append(table_cobros)
            elements.append(Spacer(1, 0.2*inch))
            
            # Total cobros
            total_cobros = sum(c.get('monto', 0) for c in cobros)
            elements.append(Paragraph(f"<b>Total Cobros: €{total_cobros:.2f}</b>", styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
        
        # Pagos
        if pagos:
            if cobros:
                elements.append(PageBreak())
            elements.append(Paragraph(f"<b>PAGOS A PROVEEDORES ({len(pagos)})</b>", styles['Heading2']))
            elements.append(Spacer(1, 0.1*inch))
            
            data_pagos = [['Proveedor', 'Monto', 'Fecha Pago', 'Estado', 'Concepto']]
            for pago in pagos:
                fecha = pago.get('fecha_pago') or pago.get('fecha_prevista') or '-'
                if fecha and fecha != '-':
                    fecha = fecha.split('T')[0] if 'T' in str(fecha) else fecha
                data_pagos.append([
                    pago.get('proveedor', '-'),
                    f"€{pago.get('monto', 0):.2f}",
                    fecha,
                    pago.get('estado', '-'),
                    pago.get('concepto', '-')[:50]
                ])
            
            table_pagos = Table(data_pagos, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 2*inch])
            table_pagos.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            elements.append(table_pagos)
            elements.append(Spacer(1, 0.2*inch))
            
            # Total pagos
            total_pagos = sum(p.get('monto', 0) for p in pagos)
            elements.append(Paragraph(f"<b>Total Pagos: €{total_pagos:.2f}</b>", styles['Normal']))
        
        # Pie de página
        elements.append(Spacer(1, 0.3*inch))
        fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M')
        elements.append(Paragraph(f"Generado el {fecha_generacion}", styles['Normal']))
        
        doc.build(elements)
        buffer.seek(0)
        
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename=historicos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            }
        )
        
    except ImportError:
        app.logger.error("reportlab no está instalado. Instale con: pip install reportlab")
        return jsonify({'error': 'Generación de PDF no disponible. Instale reportlab.'}), 500
    except Exception as e:
        app.logger.error(f"Error al generar PDF: {str(e)}")
        return jsonify({'error': f'Error al generar PDF: {str(e)}'}), 500


def generar_excel_historicos(cobros, pagos, fecha_desde=None, fecha_hasta=None):
    """Genera un archivo Excel con los históricos de cobros y pagos."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Históricos"
        
        # Estilos
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Título
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = "Histórico de Cobros y Pagos"
        cell.font = Font(bold=True, size=16, color="667eea")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Filtros
        if fecha_desde or fecha_hasta:
            filtro_text = "Período: "
            if fecha_desde:
                filtro_text += f"Desde {fecha_desde}"
            if fecha_hasta:
                filtro_text += f" Hasta {fecha_hasta}"
            ws[f'A{row}'] = filtro_text
            row += 2
        
        # Cobros
        if cobros:
            ws[f'A{row}'] = f"COBROS A RESIDENTES ({len(cobros)})"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            # Encabezados
            headers = ['Residente', 'Monto', 'Fecha Pago', 'Estado', 'Concepto', 'Método Pago', 'Residencia']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Datos
            for cobro in cobros:
                fecha = cobro.get('fecha_pago') or cobro.get('fecha_prevista') or ''
                if fecha and 'T' in str(fecha):
                    fecha = fecha.split('T')[0]
                
                ws.cell(row=row, column=1, value=cobro.get('residente', ''))
                ws.cell(row=row, column=2, value=f"€{cobro.get('monto', 0):.2f}")
                ws.cell(row=row, column=3, value=fecha)
                ws.cell(row=row, column=4, value=cobro.get('estado', ''))
                ws.cell(row=row, column=5, value=cobro.get('concepto', ''))
                ws.cell(row=row, column=6, value=cobro.get('metodo_pago', ''))
                ws.cell(row=row, column=7, value=cobro.get('nombre_residencia', ''))
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = border
                row += 1
            
            # Total
            total_cobros = sum(c.get('monto', 0) for c in cobros)
            ws.cell(row=row, column=1, value="TOTAL COBROS").font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"€{total_cobros:.2f}").font = Font(bold=True)
            row += 2
        
        # Pagos
        if pagos:
            ws[f'A{row}'] = f"PAGOS A PROVEEDORES ({len(pagos)})"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            # Encabezados
            headers = ['Proveedor', 'Monto', 'Fecha Pago', 'Estado', 'Concepto', 'Nº Factura', 'Residencia']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Datos
            for pago in pagos:
                fecha = pago.get('fecha_pago') or pago.get('fecha_prevista') or ''
                if fecha and 'T' in str(fecha):
                    fecha = fecha.split('T')[0]
                
                ws.cell(row=row, column=1, value=pago.get('proveedor', ''))
                ws.cell(row=row, column=2, value=f"€{pago.get('monto', 0):.2f}")
                ws.cell(row=row, column=3, value=fecha)
                ws.cell(row=row, column=4, value=pago.get('estado', ''))
                ws.cell(row=row, column=5, value=pago.get('concepto', ''))
                ws.cell(row=row, column=6, value=pago.get('numero_factura', ''))
                ws.cell(row=row, column=7, value=pago.get('nombre_residencia', ''))
                
                for col in range(1, 8):
                    ws.cell(row=row, column=col).border = border
                row += 1
            
            # Total
            total_pagos = sum(p.get('monto', 0) for p in pagos)
            ws.cell(row=row, column=1, value="TOTAL PAGOS").font = Font(bold=True)
            ws.cell(row=row, column=2, value=f"€{total_pagos:.2f}").font = Font(bold=True)
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return Response(
            buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=historicos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            }
        )
        
    except ImportError:
        app.logger.error("openpyxl no está instalado. Instale con: pip install openpyxl")
        return jsonify({'error': 'Generación de Excel no disponible. Instale openpyxl.'}), 500
    except Exception as e:
        app.logger.error(f"Error al generar Excel: {str(e)}")
        return jsonify({'error': f'Error al generar Excel: {str(e)}'}), 500


if __name__ == '__main__':
    # Configurar logging para mostrar en consola
    import logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    
    # Para desarrollo local
    # Usar puerto 5001 porque 5000 está ocupado por AirPlay en macOS
    port = int(os.getenv('PORT', 5001))
    
    print("\n" + "="*50)
    print("  Servidor Flask Violetas iniciado")
    print("="*50)
    print(f"  URL: http://localhost:{port}")
    print(f"  Modo: DEBUG")
    print(f"  Presiona Ctrl+C para detener")
    print("="*50 + "\n")
    
    app.run(debug=True, host='0.0.0.0', port=port)

